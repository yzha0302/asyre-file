#!/usr/bin/env python3
"""Asyre File - self-hosted markdown workspace for humans and AI agents.
Originally: Markdown editor v8 — v7 + drag-drop move + reference detection + smart new file.
Usage:
  python3 server.py [port]              # Full mode (Asher, all files)
  python3 server.py [port] --file PATH  # Share mode (single file only)
  python3 server.py [port] --dir PATH   # Scoped mode (only show files under PATH)
"""
__version__ = "1.0.0"

import json, os, re, sys, time, secrets, threading, mimetypes, base64, hashlib
from http.server import HTTPServer, SimpleHTTPRequestHandler
from urllib.parse import parse_qs, urlparse, quote as url_quote, unquote as url_unquote
from http.cookies import SimpleCookie

# Configuration
import config as _cfg
# Setup wizard
import setup_wizard as _setup

# Export module (optional)
try:
    from export import export_pdf, export_word, start_cleanup_thread, get_config
    _HAS_EXPORT = True
except ImportError:
    _HAS_EXPORT = False
    def export_pdf(*a, **kw): raise RuntimeError("Export not available. pip install weasyprint python-docx")
    def export_word(*a, **kw): raise RuntimeError("Export not available. pip install weasyprint python-docx")
    def start_cleanup_thread(): pass
    def get_config(): return {}

# Track pending downloads: {downloadId: {path, filename, created}}
_downloads = {}
_downloads_lock = threading.Lock()

# ==================== AUTH SYSTEM ====================
_EDITOR_DIR = os.path.dirname(os.path.abspath(__file__))
_USERS_PATH = os.path.join(_EDITOR_DIR, 'users.json')
_users_cache = None
_users_mtime = 0
_sessions = {}  # {token: {user, role, paths, name, expires}}
_sessions_lock = threading.Lock()
_SESSIONS_FILE = os.path.join(_EDITOR_DIR, '.sessions.json')
_DEVICE_COOKIE_MAX_AGE = 365 * 24 * 3600  # 1 year


def _save_sessions_to_file():
    """Atomically write sessions dict to JSON file. Must be called with _sessions_lock held."""
    try:
        tmp_path = _SESSIONS_FILE + '.tmp'
        with open(tmp_path, 'w') as f:
            json.dump(_sessions, f, ensure_ascii=False)
        os.replace(tmp_path, _SESSIONS_FILE)
    except Exception:
        try:
            os.unlink(tmp_path)
        except Exception:
            pass


def _load_sessions_from_file():
    """Load sessions from JSON file, filtering expired ones. Called at startup."""
    global _sessions
    try:
        with open(_SESSIONS_FILE, 'r') as f:
            data = json.load(f)
        if not isinstance(data, dict):
            data = {}
        now = time.time()
        _sessions = {k: v for k, v in data.items() if isinstance(v, dict) and v.get('expires', 0) > now}
    except (FileNotFoundError, json.JSONDecodeError, Exception):
        _sessions = {}

def get_users():
    """Load users.json with mtime-based cache — hot-reloadable."""
    global _users_cache, _users_mtime
    try:
        mt = os.path.getmtime(_USERS_PATH)
        if _users_cache is None or mt > _users_mtime:
            with open(_USERS_PATH, 'r') as f:
                _users_cache = json.load(f)
            _users_mtime = mt
    except:
        if _users_cache is None:
            _users_cache = {"users": {}, "settings": {"sessionExpireHours": 72}}
    return _users_cache

def _save_users(data):
    global _users_cache, _users_mtime
    # Re-read from disk first to avoid overwriting external edits
    try:
        mt = os.path.getmtime(_USERS_PATH)
        if mt > _users_mtime:
            with open(_USERS_PATH, 'r') as f:
                fresh = json.load(f)
            # Merge: external changes win for fields we didn't modify
            for uname, uinfo in fresh.get('users', {}).items():
                if uname not in data.get('users', {}):
                    data.setdefault('users', {})[uname] = uinfo
                else:
                    # Preserve password_hash from disk if not explicitly changed in this save
                    disk_hash = uinfo.get('password_hash', '')
                    mem_hash = data['users'][uname].get('password_hash', '')
                    if disk_hash and disk_hash != mem_hash:
                        data['users'][uname]['password_hash'] = disk_hash
    except:
        pass
    with open(_USERS_PATH, 'w') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    _users_cache = data
    _users_mtime = os.path.getmtime(_USERS_PATH)

def hash_password(password):
    salt = secrets.token_hex(16)
    h = hashlib.sha256((salt + password).encode()).hexdigest()
    return f"{salt}:{h}"

def verify_password(password, stored):
    try:
        salt, h = stored.split(':', 1)
        return hashlib.sha256((salt + password).encode()).hexdigest() == h
    except:
        return False

def create_session(username, user_info):
    token = secrets.token_urlsafe(32)
    users_data = get_users()
    expire_hours = users_data.get('settings', {}).get('sessionExpireHours', 72)
    with _sessions_lock:
        _sessions[token] = {
            'user': username,
            'name': user_info.get('name', username),
            'role': user_info.get('role', 'viewer'),
            'paths': user_info.get('paths', []),
            'expires': time.time() + expire_hours * 3600
        }
        _save_sessions_to_file()
    return token

def get_session(token):
    if not token:
        return None
    with _sessions_lock:
        sess = _sessions.get(token)
        if sess and sess['expires'] > time.time():
            return sess
        elif sess:
            del _sessions[token]
    return None

def delete_session(token):
    with _sessions_lock:
        _sessions.pop(token, None)
        _save_sessions_to_file()

def cleanup_sessions():
    now = time.time()
    with _sessions_lock:
        expired = [k for k, v in _sessions.items() if v['expires'] <= now]
        for k in expired:
            del _sessions[k]
        if expired:
            _save_sessions_to_file()

def create_device_token(username):
    """Create a persistent device token and store in users.json."""
    token = secrets.token_urlsafe(32)
    users_data = get_users()
    user = users_data.get('users', {}).get(username)
    if not user:
        return None
    devices = user.get('devices', [])
    devices.append({'token': token, 'created': time.strftime('%Y-%m-%d %H:%M')})
    # Keep max 10 devices per user
    if len(devices) > 10:
        devices = devices[-10:]
    user['devices'] = devices
    _save_users(users_data)
    return token

def verify_device_token(token):
    """Check if device token is valid, return (username, user_info) or (None, None)."""
    if not token:
        return None, None
    users_data = get_users()
    for uname, uinfo in users_data.get('users', {}).items():
        for dev in uinfo.get('devices', []):
            if dev.get('token') == token:
                return uname, uinfo
    return None, None

def verify_pin(username, pin):
    """Verify user's PIN."""
    users_data = get_users()
    user = users_data.get('users', {}).get(username)
    if not user:
        return False
    stored_pin = user.get('pin')
    if not stored_pin:
        return True  # No PIN set = auto-pass
    return str(pin) == str(stored_pin)

def set_user_pin(username, pin):
    """Set or clear PIN for a user. Empty/None = no PIN (auto-login)."""
    users_data = get_users()
    user = users_data.get('users', {}).get(username)
    if not user:
        return False
    if pin:
        user['pin'] = str(pin)
    else:
        user.pop('pin', None)
    _save_users(users_data)
    return True

def check_path_access(user_paths, file_path):
    """Check if a file path is allowed for the user's path list."""
    if '*' in user_paths:
        return True
    for allowed in user_paths:
        if file_path.startswith(allowed):
            return True
    return False

def _safe_redirect(path):
    """Convert absolute path to relative for reverse proxy safety."""
    if not path or path == '/':
        return './'
    if path.startswith('/'):
        return '.' + path
    return path


def _get_auth_cookie(handler):
    """Extract auth_token from cookies."""
    cookie_str = handler.headers.get('Cookie', '')
    if not cookie_str:
        return None
    cookies = SimpleCookie()
    try:
        cookies.load(cookie_str)
    except:
        return None
    morsel = cookies.get('auth_token')
    return morsel.value if morsel else None

WORKSPACE = _cfg.get('workspace.path')
ANN_DIR = os.path.join(WORKSPACE, 'annotations')
TRASH_DIR = os.path.join(WORKSPACE, '.trash')
os.makedirs(ANN_DIR, exist_ok=True)
os.makedirs(TRASH_DIR, exist_ok=True)
PORT = int(sys.argv[1]) if len(sys.argv) > 1 and sys.argv[1].isdigit() else _cfg.get('server.port')

# ==================== SHARE SYSTEM ====================
SHARES_FILE = os.path.join(WORKSPACE, 'shares.json')
_shares_lock = threading.Lock()

def _load_shares():
    try:
        with open(SHARES_FILE, 'r') as f:
            return json.load(f)
    except:
        return {}

def _save_shares(shares):
    os.makedirs(os.path.dirname(SHARES_FILE), exist_ok=True)
    with open(SHARES_FILE, 'w') as f:
        json.dump(shares, f, ensure_ascii=False, indent=2)

def create_share(file_path, mode='readonly', label='', share_type='file', files=None):
    """Create a share token. mode: 'readonly' or 'editable'. share_type: 'file' or 'folder'."""
    with _shares_lock:
        shares = _load_shares()
        token = secrets.token_urlsafe(8)
        entry = {
            'file': file_path,
            'mode': mode,
            'type': share_type,
            'label': label or os.path.basename(file_path),
            'created': time.strftime('%Y-%m-%d %H:%M:%S UTC', time.gmtime()),
        }
        if share_type == 'folder' and files:
            entry['files'] = files  # list of relative paths within folder
        shares[token] = entry
        _save_shares(shares)
    return token

def get_share(token):
    shares = _load_shares()
    return shares.get(token)

def delete_share(token):
    with _shares_lock:
        shares = _load_shares()
        if token in shares:
            del shares[token]
            _save_shares(shares)
            return True
    return False

import shutil

# ==================== TRASH SYSTEM ====================
TRASH_META_FILE = os.path.join(TRASH_DIR, '.trash_meta.json')
_trash_lock = threading.Lock()

def _load_trash_meta():
    try:
        with open(TRASH_META_FILE, 'r') as f:
            return json.load(f)
    except:
        return []

def _save_trash_meta(meta):
    with open(TRASH_META_FILE, 'w') as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)

def trash_item(rel_path):
    """Move a file or directory to .trash/ with metadata for restore."""
    full = os.path.normpath(os.path.join(WORKSPACE, rel_path))
    if not full.startswith(WORKSPACE) or not os.path.exists(full):
        return None
    # Don't allow trashing .trash itself or critical files
    if '.trash' in rel_path.split(os.sep):
        return None
    with _trash_lock:
        meta = _load_trash_meta()
        trash_id = f"{int(time.time())}_{secrets.token_hex(4)}"
        basename = os.path.basename(full)
        trash_name = f"{trash_id}_{basename}"
        trash_dest = os.path.join(TRASH_DIR, trash_name)
        shutil.move(full, trash_dest)
        entry = {
            'id': trash_id,
            'original': rel_path,
            'name': basename,
            'trash_name': trash_name,
            'deleted_at': time.strftime('%Y-%m-%d %H:%M:%S UTC', time.gmtime()),
            'is_dir': os.path.isdir(trash_dest)
        }
        meta.append(entry)
        _save_trash_meta(meta)
    return entry

def list_trash():
    """List items in trash."""
    return _load_trash_meta()

def restore_item(trash_id):
    """Restore an item from trash to its original location."""
    with _trash_lock:
        meta = _load_trash_meta()
        entry = None
        for m in meta:
            if m['id'] == trash_id:
                entry = m
                break
        if not entry:
            return None, 'Item not found in trash'
        trash_path = os.path.join(TRASH_DIR, entry['trash_name'])
        if not os.path.exists(trash_path):
            meta = [m for m in meta if m['id'] != trash_id]
            _save_trash_meta(meta)
            return None, 'File missing from trash'
        orig = os.path.normpath(os.path.join(WORKSPACE, entry['original']))
        if os.path.exists(orig):
            return None, f'Original path already exists: {entry["original"]}'
        os.makedirs(os.path.dirname(orig), exist_ok=True)
        shutil.move(trash_path, orig)
        meta = [m for m in meta if m['id'] != trash_id]
        _save_trash_meta(meta)
    return entry, None

def empty_trash():
    """Permanently delete all items in trash."""
    with _trash_lock:
        for item in os.listdir(TRASH_DIR):
            if item == '.trash_meta.json':
                continue
            p = os.path.join(TRASH_DIR, item)
            if os.path.isdir(p):
                shutil.rmtree(p)
            else:
                os.remove(p)
        _save_trash_meta([])

SHARE_FILE = None
SCOPED_DIR = None
SCOPED_DIRS = []  # Multiple dirs support
for i, arg in enumerate(sys.argv):
    if arg == '--file' and i+1 < len(sys.argv):
        SHARE_FILE = sys.argv[i+1]
        if not os.path.isabs(SHARE_FILE):
            SHARE_FILE = os.path.join(WORKSPACE, SHARE_FILE)
    elif arg == '--dir' and i+1 < len(sys.argv):
        raw = sys.argv[i+1]
        # Support comma-separated dirs
        for d in raw.split(','):
            d = d.strip()
            if not os.path.isabs(d):
                d = os.path.join(WORKSPACE, d)
            d = os.path.normpath(d)
            if os.path.isdir(d):
                SCOPED_DIRS.append(d)
        if SCOPED_DIRS:
            SCOPED_DIR = SCOPED_DIRS[0]  # backward compat

FULL_MODE = SHARE_FILE is None

LOGIN_HTML = r'''<!DOCTYPE html>
<html lang="zh" data-theme="dark">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>登录 — 修荷 Editor</title>
<link href="https://cdn.jsdelivr.net/npm/daisyui@5/themes.css" rel="stylesheet" type="text/css" />
<link href="https://cdn.jsdelivr.net/npm/daisyui@5" rel="stylesheet" type="text/css" />
<style>
body{font-family:-apple-system,system-ui,'Segoe UI',sans-serif;min-height:100vh;display:flex;align-items:center;justify-content:center;background:oklch(var(--b1));margin:0}
.error{display:__ERROR_DISPLAY__}
</style>
</head>
<body>
<div class="card bg-base-200 shadow-2xl w-96 max-w-[90vw]">
<div class="card-body gap-6">

<div class="text-center">
  <svg class="mx-auto mb-3 opacity-60" width="40" height="40" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.5" stroke-linecap="round" stroke-linejoin="round"><path d="M15 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V7z"/><polyline points="14 2 14 8 20 8"/><line x1="16" y1="13" x2="8" y2="13"/><line x1="16" y1="17" x2="8" y2="17"/><line x1="10" y1="9" x2="8" y2="9"/></svg>
  <h2 class="text-xl font-bold text-primary">修荷 Editor</h2>
  <p class="text-sm text-base-content/50 mt-1">请登录以继续</p>
</div>

<div class="error">
  <div class="alert alert-error text-sm">__ERROR_MSG__</div>
</div>

<form method="POST" action="__LOGIN_ACTION__">
  <input type="hidden" name="redirect" value="__REDIRECT__">

  <fieldset class="fieldset gap-4">
    <label class="fieldset-label text-xs font-semibold text-base-content/60">用户名</label>
    <input type="text" name="username" class="input input-bordered w-full" autofocus autocomplete="username" />

    <label class="fieldset-label text-xs font-semibold text-base-content/60">密码</label>
    <input type="password" name="password" class="input input-bordered w-full" autocomplete="current-password" />

    <label class="flex items-center gap-2 cursor-pointer mt-1">
      <input type="checkbox" name="remember" value="1" checked class="checkbox checkbox-xs checkbox-primary" />
      <span class="text-xs text-base-content/50">记住此设备</span>
    </label>
  </fieldset>

  <button type="submit" class="btn btn-primary w-full mt-4">登录</button>
</form>

</div>
</div>

<script>
  const saved=localStorage.getItem('md-editor-theme')||'dark';
  document.documentElement.setAttribute('data-theme',saved);
</script>
</body>
</html>'''

HTML_TEMPLATE = r'''<!DOCTYPE html>
<html lang="zh" data-theme="dark">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>📝 修荷 Editor</title>
<!-- DaisyUI + Tailwind CSS (CDN) -->
<link href="https://cdn.jsdelivr.net/npm/daisyui@5/themes.css" rel="stylesheet" type="text/css" />
<link href="https://cdn.jsdelivr.net/npm/daisyui@5" rel="stylesheet" type="text/css" />
<script src="https://cdn.jsdelivr.net/npm/@tailwindcss/browser@4"></script>

<!-- CodeMirror CSS (after Tailwind to preserve specificity) -->
<link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/codemirror@5.65.19/lib/codemirror.min.css">
<link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/codemirror@5.65.19/theme/material-darker.css">
<link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/highlight.js@11.11.1/styles/github.min.css" id="hljs-light" disabled>
<link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/codemirror@5.65.19/addon/fold/foldgutter.css">
<link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/codemirror@5.65.19/addon/dialog/dialog.css">
<script src="https://cdn.jsdelivr.net/npm/codemirror@5.65.19/lib/codemirror.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/codemirror@5.65.19/mode/markdown/markdown.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/codemirror@5.65.19/addon/mode/overlay.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/codemirror@5.65.19/mode/gfm/gfm.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/codemirror@5.65.19/mode/javascript/javascript.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/codemirror@5.65.19/mode/htmlmixed/htmlmixed.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/codemirror@5.65.19/mode/xml/xml.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/codemirror@5.65.19/mode/css/css.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/codemirror@5.65.19/mode/python/python.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/codemirror@5.65.19/mode/yaml/yaml.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/codemirror@5.65.19/mode/shell/shell.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/codemirror@5.65.19/mode/sql/sql.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/codemirror@5.65.19/mode/toml/toml.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/codemirror@5.65.19/addon/edit/closebrackets.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/codemirror@5.65.19/addon/edit/matchbrackets.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/codemirror@5.65.19/addon/fold/foldcode.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/codemirror@5.65.19/addon/fold/foldgutter.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/codemirror@5.65.19/addon/fold/markdown-fold.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/codemirror@5.65.19/addon/selection/active-line.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/codemirror@5.65.19/addon/search/search.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/codemirror@5.65.19/addon/search/searchcursor.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/codemirror@5.65.19/addon/dialog/dialog.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/marked@15.0.7/marked.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/mermaid@11.4.1/dist/mermaid.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/highlight.js@11.11.1/highlight.min.js"></script>
<link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/highlight.js@11.11.1/styles/github-dark.min.css">
<style>
/* CodeMirror protection — prevent Tailwind reset from breaking editor */
.CodeMirror, .CodeMirror * { box-sizing: content-box !important; }
.CodeMirror pre { font-family: monospace; }
.CodeMirror textarea { appearance: auto; }
.CodeMirror-dialog { box-sizing: border-box; }
.CodeMirror-scroll { box-sizing: border-box !important; }

*{box-sizing:border-box;margin:0;padding:0}
:root{--bg:#0d1117;--bg2:#161b22;--border:#30363d;--text:#e6edf3;--dim:#7d8590;--accent:#ffa500;--accent2:#ff8c00;--green:#3fb950;--red:#f85149;--blue:#58a6ff;--purple:#8957e5}
body{font-family:-apple-system,system-ui,'Segoe UI',sans-serif;background:var(--bg);color:var(--text);height:100vh;display:flex;flex-direction:column;overflow:hidden}

/* Toolbar */
.toolbar{background:var(--bg2);padding:6px 16px;display:flex;align-items:center;gap:10px;border-bottom:1px solid var(--border);flex-shrink:0;height:40px}
.toolbar h1{font-size:14px;color:var(--accent)}
.toolbar .spacer{flex:1}
.toolbar button{background:transparent;color:var(--text);border:none;padding:6px 8px;border-radius:6px;font-size:13px;cursor:pointer;display:inline-flex;align-items:center;justify-content:center}
.toolbar button:hover{background:rgba(255,255,255,.1)}
.toolbar button svg{width:16px;height:16px}
[data-theme="light"] .toolbar button{color:#1f2328 !important}
[data-theme="light"] .toolbar button svg{stroke:#1f2328 !important}
[data-theme="light"] .toolbar{background:#f6f8fa !important;border-bottom-color:#d0d7de !important}
[data-theme="light"] .toolbar .ann-btn{color:#8957e5 !important}
[data-theme="light"] .toolbar .ann-btn svg{stroke:#8957e5 !important}
[data-theme="light"] .toolbar .admin-btn svg{stroke:#1f2328 !important}
[data-theme="light"] .toolbar .theme-btn svg{stroke:#1f2328 !important}
[data-theme="light"] .toolbar .logout-btn svg{stroke:#656d76 !important}
[data-theme="light"] .toolbar button:hover{background:rgba(0,0,0,.08)}
[data-theme="light"] .toolbar .sep{background:#d0d7de}
.toolbar button.ann-btn{background:transparent;color:var(--purple)}
.toolbar button.ann-btn:hover{background:rgba(137,87,229,.15)}
.toolbar button.ann-btn.active-mode{background:rgba(137,87,229,.2);color:#a371f7;box-shadow:0 0 0 2px rgba(137,87,229,.3)}
.toolbar button.ann-submit{background:var(--green)}
.toolbar button.ann-submit:hover{background:#46954a}
.toolbar .status{display:none}
.toolbar .filename{font-size:12px;color:var(--accent);font-family:monospace;max-width:400px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.toolbar .sep{width:1px;height:20px;background:var(--border)}
.toolbar .badge{background:var(--red);color:#fff;font-size:10px;padding:1px 6px;border-radius:8px;margin-left:-6px}

/* Layout */
.main{flex:1;display:flex;overflow:hidden}
.sidebar{width:260px;background:var(--bg2);border-right:1px solid var(--border);display:flex;flex-direction:column;flex-shrink:0}
.sidebar .search{padding:6px 8px;border-bottom:1px solid var(--border)}
.sidebar .search input{width:100%;background:var(--bg);color:var(--text);border:1px solid var(--border);border-radius:4px;padding:5px 8px;font-size:12px}
.sidebar .search input:focus{outline:none;border-color:var(--accent)}
.tree{flex:1;overflow-y:auto;-webkit-overflow-scrolling:touch;padding:2px 0;font-size:12px}
.tree-dir{} 
.tree-dir-label{padding:3px 8px;display:flex;align-items:center;gap:3px;color:var(--dim);cursor:pointer}
.tree-dir-label:hover{background:rgba(255,255,255,.04);color:var(--text)}
.tree-dir-label .arrow{display:inline-block;width:14px;font-size:9px;transition:transform .12s;text-align:center}
.tree-dir-label .arrow.open{transform:rotate(90deg)}
.tree-children{display:none;padding-left:12px}
.tree-children.open{display:block}
.tree-file{padding:3px 8px 3px 24px;cursor:pointer;color:var(--dim);white-space:nowrap;overflow:hidden;text-overflow:ellipsis;border-left:2px solid transparent;display:flex;align-items:center;gap:4px}
.tree-file svg{flex-shrink:0}
.tree-file:hover{background:rgba(255,255,255,.04);color:var(--text)}
.tree-file.active{color:var(--accent);border-left-color:var(--accent);background:rgba(255,165,0,.06)}
.file-count{padding:4px 8px;font-size:11px;color:#484f58;border-top:1px solid var(--border);flex-shrink:0}

/* Content */
.content{flex:1;display:flex;flex-direction:column;overflow:hidden}
.tabs{display:flex;background:var(--bg2);border-bottom:1px solid var(--border);flex-shrink:0}
.tabs button{background:none;border:none;color:var(--dim);padding:8px 12px;cursor:pointer;font-size:12px;border-bottom:2px solid transparent;display:inline-flex;align-items:center}
.tabs button svg{width:16px;height:16px}
.tabs button:hover{color:var(--text)}
.tabs button.active{color:var(--text);border-bottom-color:var(--accent)}
.panels{flex:1;display:flex;overflow:hidden}
.editor-pane,.preview-pane{flex:1;overflow:auto;-webkit-overflow-scrolling:touch;will-change:scroll-position}
.panels.split .editor-pane,.panels.split .preview-pane{display:block}
.panels.split .editor-pane{border-right:1px solid var(--border)}
.panels.edit-only .preview-pane{display:none}
.panels.preview-only .editor-pane{display:none}

/* CodeMirror */
.CodeMirror{height:100%!important;font-family:'JetBrains Mono',Menlo,'Courier New',monospace!important;font-size:13px!important;line-height:1.6!important}
.cm-s-material-darker .CodeMirror-activeline-background{background:rgba(255,165,0,.05)!important}
.cm-s-material-darker .CodeMirror-gutters{background:var(--bg2)!important;border-right:1px solid var(--border)!important}
.cm-s-material-darker .CodeMirror-linenumber{color:#484f58!important}
.cm-s-material-darker .CodeMirror-cursor{border-left-color:var(--accent)!important}
.cm-s-material-darker .cm-header{color:var(--accent)!important;font-weight:bold}
.cm-s-material-darker .cm-header-1{font-size:1.3em}
.cm-s-material-darker .cm-header-2{font-size:1.15em;color:var(--accent2)!important}
.cm-s-material-darker .cm-strong{color:#ffd700!important}
.cm-s-material-darker .cm-em{color:var(--blue)!important}
.cm-s-material-darker .cm-link{color:var(--blue)!important}
.cm-s-material-darker .cm-quote{color:#8b949e!important;font-style:italic}

/* Preview */
.preview-pane{padding:20px 28px;line-height:1.8;font-size:15px}
.preview-pane h1{font-size:1.6em;color:var(--accent);margin:20px 0 10px;border-bottom:1px solid var(--border);padding-bottom:8px}
.preview-pane h2{font-size:1.35em;color:var(--accent2);margin:18px 0 8px}
.preview-pane h3{font-size:1.15em;color:#e6a817;margin:14px 0 6px}
.preview-pane p{margin:8px 0}
.preview-pane strong{color:#ffd700}
.preview-pane em{color:var(--blue)}
.preview-pane a{color:var(--blue)}
.preview-pane code{background:var(--bg2);padding:2px 6px;border-radius:3px;font-size:.88em;font-family:monospace}
.preview-pane pre{background:var(--bg2);padding:14px;border-radius:6px;overflow-x:auto;margin:10px 0;border:1px solid var(--border)}
.preview-pane pre code{padding:0;background:none;font-size:13px}
.preview-pane blockquote{border-left:3px solid var(--accent);padding:4px 12px;color:#8b949e;margin:10px 0;background:rgba(255,165,0,.03)}
.preview-pane hr{border:none;border-top:1px solid var(--border);margin:20px 0}
.preview-pane ul,.preview-pane ol{padding-left:24px;margin:8px 0}
.preview-pane li{margin:4px 0}
.preview-pane table{border-collapse:collapse;margin:10px 0;width:100%;font-size:14px}
.preview-pane th,.preview-pane td{border:1px solid var(--border);padding:6px 12px}
.preview-pane th{background:var(--bg2)}
.preview-pane img{max-width:100%;border-radius:6px}

/* Mermaid */
.mermaid-wrap{background:var(--bg2);padding:16px;border-radius:6px;margin:10px 0;text-align:center;border:1px solid var(--border);position:relative;cursor:zoom-in}
.mermaid-wrap:hover{border-color:var(--accent)}
.mermaid-wrap .zoom-hint{position:absolute;top:6px;right:8px;font-size:11px;color:var(--dim);opacity:0;transition:opacity .2s}
.mermaid-wrap:hover .zoom-hint{opacity:1}

/* Mermaid light mode */
[data-theme="light"] .mermaid-wrap{background:#f6f8fa;border-color:#d0d7de}
[data-theme="light"] .mermaid-wrap svg rect,[data-theme="light"] .mermaid-wrap svg circle,[data-theme="light"] .mermaid-wrap svg polygon,[data-theme="light"] .mermaid-wrap svg ellipse{fill:#fff!important;stroke:#d0d7de!important}
[data-theme="light"] .mermaid-wrap svg text,[data-theme="light"] .mermaid-wrap svg tspan{fill:#1f2328!important}
[data-theme="light"] .mermaid-wrap svg foreignObject div,[data-theme="light"] .mermaid-wrap svg .nodeLabel{color:#1f2328!important}
[data-theme="light"] .mermaid-wrap svg .cluster rect{fill:#f0f0f0!important;stroke:#e67e22!important}
[data-theme="light"] .mermaid-wrap svg .edgePath path,[data-theme="light"] .mermaid-wrap svg .flowchart-link{stroke:#0969da!important}
[data-theme="light"] .mermaid-wrap svg marker path{fill:#0969da!important}
[data-theme="light"] .mermaid-wrap svg .edgeLabel rect{fill:#f6f8fa!important}

/* Mermaid nuclear: ALL shapes dark, ALL text white */
.mermaid-wrap svg rect,.mermaid-wrap svg circle,.mermaid-wrap svg polygon{fill:#2d333b!important;stroke:#444c56!important}
.mermaid-wrap svg foreignObject div,.mermaid-wrap svg foreignObject span,.mermaid-wrap svg .nodeLabel{color:#e6edf3!important}
.mermaid-wrap svg text,.mermaid-wrap svg tspan{fill:#e6edf3!important}
.mermaid-wrap svg .cluster rect{fill:#1c2128!important;stroke:#ffa500!important}
.mermaid-wrap svg .edgePath path{stroke:#58a6ff!important;fill:none!important}
.mermaid-wrap svg marker path{fill:#58a6ff!important;stroke:none!important}
/* Lightbox */
.mermaid-lightbox svg rect,.mermaid-lightbox svg circle,.mermaid-lightbox svg polygon{fill:#2d333b!important;stroke:#444c56!important}
.mermaid-lightbox svg foreignObject div,.mermaid-lightbox svg .nodeLabel{color:#e6edf3!important}
.mermaid-lightbox svg text,.mermaid-lightbox svg tspan{fill:#e6edf3!important}
.mermaid-lightbox svg .cluster rect{fill:#1c2128!important;stroke:#ffa500!important}

/* Mermaid Lightbox */
.mermaid-lightbox{position:fixed;inset:0;background:rgba(0,0,0,.9);z-index:300;display:flex;flex-direction:column}
.mermaid-lightbox .lb-viewport{flex:1;overflow:hidden;cursor:grab;display:flex;align-items:center;justify-content:center}
.mermaid-lightbox .lb-content{transform-origin:center center;transition:none;padding:32px}
.mermaid-lightbox .lb-content svg{max-width:none!important;max-height:none!important}
.mermaid-lightbox .lb-close{position:fixed;top:16px;right:24px;font-size:28px;color:var(--text);cursor:pointer;z-index:301;background:var(--bg2);border-radius:50%;width:40px;height:40px;display:flex;align-items:center;justify-content:center;border:1px solid var(--border)}
.mermaid-lightbox .lb-zoom-controls{position:fixed;bottom:24px;left:50%;transform:translateX(-50%);z-index:301;background:var(--bg2);border:1px solid var(--border);border-radius:10px;padding:6px 12px;display:flex;align-items:center;gap:8px}
.mermaid-lightbox .lb-zoom-controls button{background:none;border:1px solid var(--border);color:var(--text);width:32px;height:32px;border-radius:6px;font-size:18px;cursor:pointer;display:flex;align-items:center;justify-content:center}
.mermaid-lightbox .lb-zoom-controls button:hover{background:rgba(255,255,255,.1)}
.mermaid-lightbox .lb-zoom-controls span{color:var(--dim);font-size:13px;min-width:50px;text-align:center}

/* Annotation gutter */
.annotation-gutter{width:24px;cursor:pointer}
.ann-marker{color:var(--purple);font-size:16px;cursor:pointer;text-align:center;line-height:1.6}
.ann-marker:hover{color:#a371f7}
.ann-popup{position:fixed;z-index:100;background:var(--bg2);border:1px solid var(--border);border-radius:10px;padding:16px;width:380px;box-shadow:0 8px 30px rgba(0,0,0,.3);animation:modalIn .15s ease-out}
.ann-popup .ann-range{font-size:11px;color:var(--purple);margin-bottom:6px;font-weight:500}
.ann-popup .ann-preview{font-size:11px;color:var(--dim);background:var(--bg);padding:8px 10px;border-radius:6px;margin-bottom:10px;max-height:80px;overflow-y:auto;border-left:2px solid var(--purple);font-family:'JetBrains Mono',monospace;white-space:pre-wrap;line-height:1.5}
.ann-popup textarea{width:100%;background:var(--bg);color:var(--text);border:1px solid var(--border);border-radius:6px;padding:8px 10px;font-size:13px;resize:vertical;min-height:60px;font-family:inherit;transition:border-color .15s}
.ann-popup textarea:focus{outline:none;border-color:var(--purple)}
.ann-popup .ann-actions{display:flex;gap:6px;margin-top:10px;justify-content:flex-end}
.ann-popup .ann-actions button{padding:5px 14px;border-radius:6px;font-size:12px;cursor:pointer;border:none;font-weight:500;transition:opacity .15s}
.ann-popup .ann-actions button:hover{opacity:.85}
.ann-popup .ann-actions .save{background:var(--purple);color:#fff}
.ann-popup .ann-actions .delete{background:var(--red);color:#fff}
.ann-popup .ann-actions .cancel{background:transparent;color:var(--dim);border:1px solid var(--border)}
[data-theme="light"] .ann-popup{background:#fff;border-color:#d0d7de;box-shadow:0 8px 30px rgba(0,0,0,.1)}
.ann-line{background:rgba(137,87,229,.08)!important}
/* Selection annotation floating button */
.ann-sel-btn{position:fixed;z-index:90;background:var(--purple);color:#fff;border:none;padding:5px 12px;border-radius:6px;font-size:12px;cursor:pointer;font-weight:600;box-shadow:0 2px 8px rgba(0,0,0,.3);white-space:nowrap}
.ann-sel-btn:hover{background:#a371f7}
/* Annotation list panel */
#annListItems{max-height:100px;overflow-y:auto;margin-top:4px;border-top:1px solid var(--border);padding-top:4px}
.ann-list-item{display:flex;align-items:center;gap:6px;padding:3px 0;font-size:11px;border-bottom:1px solid rgba(255,255,255,.04)}
.ann-list-item .ann-range-tag{color:var(--purple);font-weight:600;white-space:nowrap;flex-shrink:0}
.ann-list-item .ann-comment{color:var(--text);flex:1;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.ann-list-item .ann-del{color:var(--red);cursor:pointer;flex-shrink:0;padding:0 4px}

/* Annotation bottom bar */
.ann-bar{position:fixed;bottom:0;left:0;right:0;background:var(--bg2);border-top:1px solid var(--border);padding:8px 16px;display:flex;flex-direction:column;gap:4px;font-size:13px;z-index:50;box-shadow:0 -4px 20px rgba(0,0,0,.15)}
.ann-bar .ann-header{display:flex;align-items:center;gap:12px;width:100%}
.ann-bar .ann-header .ann-badge{background:var(--purple);color:#fff;font-size:11px;font-weight:600;padding:2px 10px;border-radius:10px}
.ann-bar .ann-header .ann-hint{color:var(--dim);font-size:11px}
.ann-bar .ann-header .ann-copy-btn{margin-left:auto;padding:5px 14px;border-radius:6px;background:var(--green);color:#fff;border:none;cursor:pointer;font-size:12px;font-weight:500;display:flex;align-items:center;gap:4px}
.ann-bar .ann-header .ann-copy-btn:hover{opacity:.9}
.ann-list-item{display:flex;align-items:center;gap:8px;padding:6px 10px;border-radius:6px;transition:background .1s}
.ann-list-item:hover{background:rgba(137,87,229,.06)}
.ann-list-item .ann-range-tag{background:rgba(137,87,229,.12);color:var(--purple);padding:2px 8px;border-radius:4px;font-size:11px;font-weight:500;white-space:nowrap;cursor:pointer}
.ann-list-item .ann-comment{flex:1;font-size:12px;color:var(--text);overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.ann-list-item .ann-del{color:var(--dim);cursor:pointer;font-size:14px;padding:2px 4px;border-radius:4px;opacity:0;transition:opacity .15s}
.ann-list-item:hover .ann-del{opacity:1}
.ann-list-item .ann-del:hover{color:var(--red);background:rgba(248,81,73,.1)}
[data-theme="light"] .ann-bar{background:#fff;border-top-color:#d0d7de;box-shadow:0 -4px 20px rgba(0,0,0,.06)}
[data-theme="light"] .ann-list-item:hover{background:rgba(137,87,229,.04)}

/* Mobile Toggle Button */
.sidebar-toggle{display:none;position:fixed;bottom:20px;left:20px;z-index:200;background:var(--accent);color:var(--bg);border:none;width:48px;height:48px;border-radius:50%;font-size:22px;cursor:pointer;box-shadow:0 4px 12px rgba(0,0,0,.5);align-items:center;justify-content:center}
.sidebar-toggle:hover{background:var(--accent2)}

/* Mobile Responsive */
@media(max-width:768px){
  .sidebar{position:fixed;left:-280px;top:40px;bottom:0;width:280px;z-index:100;transition:left .25s ease;box-shadow:4px 0 20px rgba(0,0,0,.5)}
  .sidebar.mobile-open{left:0}
  .sidebar-toggle{display:flex}
  .sidebar-overlay{display:none;position:fixed;inset:0;background:rgba(0,0,0,.5);z-index:99}
  .sidebar-overlay.visible{display:block}
  /* Toolbar mobile: compact buttons */
  .toolbar{padding:4px 6px;gap:4px;height:auto;min-height:40px;flex-wrap:wrap}
  .toolbar h1{font-size:11px;white-space:nowrap;max-width:60px;overflow:hidden;text-overflow:ellipsis}
  .toolbar .filename{max-width:80px;font-size:11px}
  .toolbar .status{font-size:10px}
  .toolbar button{padding:4px 8px;font-size:11px}
  .toolbar .sep{display:none}
  .toolbar .spacer{flex:0 0 auto;width:4px}
  /* Hide button text on very small screens, show emoji only */
  .toolbar button 
  /* Preview pane mobile */
  .preview-pane{padding:12px 14px;font-size:14px;padding-bottom:80px}
  .preview-pane h1{font-size:1.3em}
  .preview-pane h2{font-size:1.15em}
  /* Editor pane bottom padding for keyboard/FABs */
  .editor-pane{padding-bottom:60px}
  .CodeMirror{font-size:13px}
  /* Dialogs full-width */
  .ann-popup{width:calc(100vw - 32px);left:16px!important;right:16px!important}
      /* View mode tabs */
  .view-tabs{font-size:12px}
  .view-tabs button{padding:4px 10px;font-size:12px}
  /* User info compact */
  .user-info{font-size:11px}
  .user-info .user-name{max-width:40px;overflow:hidden;text-overflow:ellipsis;display:inline-block;vertical-align:middle}
}

/* Drag & Drop */
.tree-file[draggable="true"]{cursor:grab}
.tree-file.dragging{opacity:.4}
.tree-dir-label.drag-over{background:rgba(255,165,0,.15)!important;border-radius:4px}
.tree-dir-label.drag-over::after{content:'↓';margin-left:auto;font-size:12px}
.drop-root{padding:4px 8px;font-size:11px;color:var(--dim);border-top:1px solid var(--border);text-align:center;min-height:24px}
.drop-root.drag-over{background:rgba(255,165,0,.15);color:var(--accent)}

/* Reference warning dialog */

.ref-dialog 
.ref-dialog .ref-file{margin:8px 0;padding:8px;background:var(--bg);border-radius:6px;border-left:3px solid var(--accent)}
.ref-dialog .ref-file .ref-path{color:var(--accent);font-family:monospace;font-size:12px;font-weight:600}
.ref-dialog .ref-file .ref-line{color:var(--dim);font-size:11px;margin-top:2px;font-family:monospace;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}

/* Context menu */
.ctx-menu{position:fixed;z-index:10000;background:var(--bg2);border:1px solid var(--border);border-radius:8px;padding:4px 0;min-width:200px;box-shadow:0 8px 30px rgba(0,0,0,.4);backdrop-filter:blur(12px);animation:ctxIn .1s ease-out}
@keyframes ctxIn{from{opacity:0;transform:scale(.96)}to{opacity:1;transform:scale(1)}}
.ctx-menu-item{padding:7px 14px;font-size:13px;cursor:pointer;color:var(--text);display:flex;align-items:center;gap:8px;transition:background .1s}
.ctx-menu-item:hover{background:rgba(255,255,255,.08)}
.ctx-menu-item svg{flex-shrink:0;opacity:.6}
.ctx-menu-item.danger{color:var(--red)}
.ctx-menu-item.danger:hover{background:rgba(248,81,73,.1)}
.ctx-menu-sep{border-top:1px solid var(--border);margin:4px 8px}
.ctx-menu-label{padding:4px 14px;font-size:10px;text-transform:uppercase;letter-spacing:.5px;color:var(--dim);pointer-events:none}
.ctx-menu-shortcut{margin-left:auto;font-size:11px;color:var(--dim);opacity:.5}
[data-theme="light"] .ctx-menu{box-shadow:0 8px 30px rgba(0,0,0,.15)}
[data-theme="light"] .ctx-menu-item:hover{background:rgba(0,0,0,.05)}
/* Multi-select */
.tree-file.selected{background:rgba(88,166,255,.15);border-radius:4px}
.tree-file.selected:hover{background:rgba(88,166,255,.22)}
.multi-select-bar{position:fixed;bottom:20px;left:50%;transform:translateX(-50%);background:var(--bg2);border:1px solid var(--border);border-radius:10px;padding:8px 16px;display:flex;align-items:center;gap:12px;z-index:500;box-shadow:0 8px 30px rgba(0,0,0,.3);font-size:13px;animation:ctxIn .15s ease-out}

/* Upload */
.drop-overlay{position:fixed;inset:0;z-index:9999;background:rgba(88,166,255,.1);border:3px dashed var(--blue);display:flex;align-items:center;justify-content:center;pointer-events:none}
.drop-overlay span{background:var(--bg2);padding:16px 32px;border-radius:12px;font-size:16px;font-weight:600;color:var(--blue);box-shadow:0 8px 30px rgba(0,0,0,.3)}
.multi-select-bar button{padding:5px 12px;border-radius:6px;border:none;cursor:pointer;font-size:12px;font-weight:600;display:flex;align-items:center;gap:4px}

/* Share dialog */
.share-link-box button{background:var(--blue);color:#000;border:none;padding:6px 14px;border-radius:4px;cursor:pointer;font-weight:600;font-size:12px;white-space:nowrap}
/* Read-only banner */
.readonly-banner{background:rgba(63,185,80,.1);border-bottom:2px solid var(--green);padding:8px 16px;font-size:13px;color:var(--green);text-align:center;flex-shrink:0}

/* Toast */
.toast{position:fixed;top:60px;right:20px;background:var(--green);color:#fff;padding:12px 20px;border-radius:8px;font-size:14px;font-weight:600;z-index:400;animation:fadeInOut 3s ease}
@keyframes fadeInOut{0%{opacity:0;transform:translateY(-10px)}10%{opacity:1;transform:translateY(0)}80%{opacity:1}100%{opacity:0}}
/* Theme: Light mode overrides */
[data-theme="light"] { --bg:#ffffff; --bg2:#f6f8fa; --border:#d0d7de; --text:#1f2328; --dim:#656d76; }
[data-theme="light"] .toolbar { background:#f6f8fa; border-bottom-color:#d0d7de; }
[data-theme="light"] .sidebar { background:#f6f8fa; border-right-color:#d0d7de; }
[data-theme="light"] .tabs { background:#f6f8fa; border-bottom-color:#d0d7de; }
[data-theme="light"] .tree-file.active { background:rgba(255,165,0,.08); }
[data-theme="light"] .preview-pane { color:#1f2328; }
[data-theme="light"] .preview-pane code { background:#f6f8fa; }
[data-theme="light"] .preview-pane pre { background:#f6f8fa; border-color:#d0d7de; }
[data-theme="light"] .preview-pane blockquote { background:rgba(255,165,0,.05); }
[data-theme="light"] .preview-pane th { background:#f6f8fa; }
[data-theme="light"] .preview-pane td, [data-theme="light"] .preview-pane th { border-color:#d0d7de; }
[data-theme="light"] .mermaid-wrap { background:#f6f8fa; border-color:#d0d7de; }
[data-theme="light"] .mermaid-wrap svg rect, [data-theme="light"] .mermaid-wrap svg circle, [data-theme="light"] .mermaid-wrap svg polygon { fill:#f0f3f6 !important; stroke:#d0d7de !important; }
[data-theme="light"] .mermaid-wrap svg text, [data-theme="light"] .mermaid-wrap svg tspan { fill:#1f2328 !important; }
[data-theme="light"] .mermaid-wrap svg foreignObject div, [data-theme="light"] .mermaid-wrap svg foreignObject span, [data-theme="light"] .mermaid-wrap svg .nodeLabel { color:#1f2328 !important; }
[data-theme="light"] .ann-bar { background:#f6f8fa; }
[data-theme="light"] .preview-pane pre code { color:#1f2328 !important; }
[data-theme="light"] .preview-pane pre code.hljs { background:#f6f8fa !important; color:#1f2328 !important; }
[data-theme="light"] .preview-pane .hljs { background:#f6f8fa !important; color:#1f2328 !important; }
[data-theme="light"] .preview-pane pre { background:#f6f8fa !important; }
[data-theme="light"] .hljs { background:#f6f8fa !important; }
[data-theme="light"] .preview-pane .mermaid-wrap { background:#f6f8fa !important; }
[data-theme="light"] .preview-pane img { background:transparent !important; }
[data-theme="light"] .preview-pane code { color:#cf222e; background:#f0f3f6; }
[data-theme="light"] .preview-pane strong { color:#1f2328; }
[data-theme="light"] .preview-pane em { color:#0550ae; }
[data-theme="light"] .preview-pane a { color:#0969da; }
[data-theme="light"] .preview-pane h1 { color:#d4760a; border-bottom-color:#d0d7de; }
[data-theme="light"] .preview-pane h2 { color:#bf6208; }
[data-theme="light"] .preview-pane h3 { color:#9a6700; }
[data-theme="light"] .preview-pane blockquote { color:#656d76; border-left-color:#d4760a; }
[data-theme="light"] .preview-pane hr { border-top-color:#d0d7de; }
[data-theme="light"] .preview-pane li { color:#1f2328; }
[data-theme="light"] .preview-pane p { color:#1f2328; }
[data-theme="light"] body { background:#ffffff; color:#1f2328; }
[data-theme="light"] .CodeMirror, [data-theme="light"] .CodeMirror-scroll { background:#ffffff !important; color:#1f2328 !important; }
[data-theme="light"] .CodeMirror pre, [data-theme="light"] .CodeMirror pre.CodeMirror-line, [data-theme="light"] .CodeMirror pre.CodeMirror-line-like { color:#1f2328 !important; }
[data-theme="light"] .CodeMirror span:not([class]) { color:#1f2328 !important; }
[data-theme="light"] .CodeMirror-gutters { background:#f6f8fa !important; border-right-color:#d0d7de !important; }
[data-theme="light"] .CodeMirror-linenumber { color:#8b949e !important; }
[data-theme="light"] .CodeMirror-cursor { border-left-color:#d4760a !important; }
[data-theme="light"] .CodeMirror-activeline-background { background:rgba(255,165,0,.06) !important; }
[data-theme="light"] .CodeMirror-selected { background:#b6d7ff !important; }
[data-theme="light"] .CodeMirror .cm-header { color:#d4760a !important; font-weight:bold !important; }
[data-theme="light"] .CodeMirror .cm-header-1 { color:#c45100 !important; }
[data-theme="light"] .CodeMirror .cm-header-2 { color:#d4760a !important; }
[data-theme="light"] .CodeMirror .cm-header-3 { color:#9a6700 !important; }
[data-theme="light"] .CodeMirror .cm-strong { color:#1f2328 !important; font-weight:bold !important; }
[data-theme="light"] .CodeMirror .cm-em { color:#0550ae !important; }
[data-theme="light"] .CodeMirror .cm-link { color:#0969da !important; }
[data-theme="light"] .CodeMirror .cm-url { color:#0969da !important; }
[data-theme="light"] .CodeMirror .cm-quote { color:#57606a !important; }
[data-theme="light"] .CodeMirror .cm-comment { color:#57606a !important; }
[data-theme="light"] .CodeMirror .cm-keyword { color:#cf222e !important; }
[data-theme="light"] .CodeMirror .cm-atom { color:#0550ae !important; }
[data-theme="light"] .CodeMirror .cm-number { color:#0550ae !important; }
[data-theme="light"] .CodeMirror .cm-string { color:#0a3069 !important; }
[data-theme="light"] .CodeMirror .cm-string-2 { color:#0a3069 !important; }
[data-theme="light"] .CodeMirror .cm-variable { color:#1f2328 !important; }
[data-theme="light"] .CodeMirror .cm-variable-2 { color:#1f2328 !important; }
[data-theme="light"] .CodeMirror .cm-variable-3 { color:#1f2328 !important; }
[data-theme="light"] .CodeMirror .cm-property { color:#1f2328 !important; }
[data-theme="light"] .CodeMirror .cm-tag { color:#116329 !important; }
[data-theme="light"] .CodeMirror .cm-attribute { color:#0550ae !important; }
[data-theme="light"] .CodeMirror .cm-builtin { color:#8250df !important; }
[data-theme="light"] .CodeMirror .cm-meta { color:#57606a !important; }
[data-theme="light"] .CodeMirror .cm-formatting { color:#57606a !important; }
[data-theme="light"] .CodeMirror .cm-hr { color:#d0d7de !important; }
[data-theme="light"] .CodeMirror .cm-image-marker { color:#0969da !important; }
[data-theme="light"] .CodeMirror .cm-image-alt-text { color:#0550ae !important; }
[data-theme="light"] .CodeMirror .cm-def { color:#8250df !important; }
[data-theme="light"] .CodeMirror .cm-operator { color:#cf222e !important; }
[data-theme="light"] .CodeMirror .cm-bracket { color:#1f2328 !important; }
[data-theme="light"] .CodeMirror .cm-punctuation { color:#1f2328 !important; }
[data-theme="light"] .CodeMirror .cm-type { color:#0550ae !important; }
[data-theme="light"] .CodeMirror .cm-qualifier { color:#8250df !important; }
[data-theme="light"] .CodeMirror .cm-error { color:#cf222e !important; }
[data-theme="light"] .CodeMirror .cm-invalidchar { color:#cf222e !important; }
[data-theme="light"] .CodeMirror .cm-matchhighlight { background:rgba(255,165,0,.15) !important; }
[data-theme="light"] .CodeMirror-foldgutter-open, [data-theme="light"] .CodeMirror-foldgutter-folded { color:#8b949e !important; }
[data-theme="light"] .CodeMirror-foldmarker { color:#0969da !important; }
[data-theme="light"] .CodeMirror .cm-formatting-code, [data-theme="light"] .CodeMirror .cm-formatting-code-block { color:#57606a !important; }
[data-theme="light"] .CodeMirror .cm-formatting-list { color:#cf222e !important; }
[data-theme="light"] .CodeMirror .cm-formatting-quote { color:#57606a !important; }
[data-theme="light"] .CodeMirror .cm-formatting-strong { color:#1f2328 !important; }
[data-theme="light"] .CodeMirror .cm-formatting-em { color:#0550ae !important; }
[data-theme="light"] .CodeMirror .cm-formatting-header { color:#d4760a !important; }
[data-theme="light"] .CodeMirror .cm-formatting-link { color:#0969da !important; }
[data-theme="light"] .CodeMirror .cm-formatting-image { color:#0969da !important; }
[data-theme="light"] .CodeMirror .cm-formatting-strikethrough { color:#57606a !important; }
[data-theme="light"] .cm-s-material-darker.CodeMirror { background:#ffffff !important; color:#1f2328 !important; }
[data-theme="light"] .cm-s-material-darker .CodeMirror-gutters { background:#f6f8fa !important; }
[data-theme="light"] .cm-s-material-darker .cm-header { color:#d4760a !important; }
[data-theme="light"] .cm-s-material-darker .cm-strong { color:#1f2328 !important; }
[data-theme="light"] .cm-s-material-darker .cm-em { color:#0550ae !important; }
[data-theme="light"] .cm-s-material-darker .cm-link { color:#0969da !important; }
[data-theme="light"] .cm-s-material-darker .cm-quote { color:#57606a !important; }
[data-theme="light"] .cm-s-material-darker .cm-comment { color:#57606a !important; }
[data-theme="light"] .cm-s-material-darker .cm-keyword { color:#cf222e !important; }
[data-theme="light"] .cm-s-material-darker .cm-atom { color:#0550ae !important; }
[data-theme="light"] .cm-s-material-darker .cm-number { color:#0550ae !important; }
[data-theme="light"] .cm-s-material-darker .cm-string { color:#0a3069 !important; }
[data-theme="light"] .cm-s-material-darker .cm-string-2 { color:#0a3069 !important; }
[data-theme="light"] .cm-s-material-darker .cm-variable { color:#1f2328 !important; }
[data-theme="light"] .cm-s-material-darker .cm-variable-2 { color:#1f2328 !important; }
[data-theme="light"] .cm-s-material-darker .cm-property { color:#1f2328 !important; }
[data-theme="light"] .cm-s-material-darker .cm-tag { color:#116329 !important; }
[data-theme="light"] .cm-s-material-darker .cm-attribute { color:#0550ae !important; }
[data-theme="light"] .cm-s-material-darker .cm-builtin { color:#8250df !important; }
[data-theme="light"] .cm-s-material-darker .cm-meta { color:#57606a !important; }
[data-theme="light"] .cm-s-material-darker .cm-def { color:#8250df !important; }
[data-theme="light"] .cm-s-material-darker .cm-operator { color:#cf222e !important; }
[data-theme="light"] .cm-s-material-darker .cm-bracket { color:#1f2328 !important; }
[data-theme="light"] .cm-s-material-darker .CodeMirror-cursor { border-left-color:#d4760a !important; }
[data-theme="light"] .cm-s-material-darker .CodeMirror-activeline-background { background:rgba(255,165,0,.06) !important; }
[data-theme="light"] .cm-s-material-darker .CodeMirror-selected { background:#b6d7ff !important; }
[data-theme="light"] .cm-s-material-darker .CodeMirror-linenumber { color:#8b949e !important; }
[data-theme="light"] .CodeMirror .cm-image-marker { color:#0969da !important; }
[data-theme="light"] .CodeMirror .CodeMirror-matchingbracket { color:#1f2328 !important; background:rgba(255,165,0,.2) !important; }
[data-theme="light"] .CodeMirror-dialog { background:#f6f8fa !important; color:#1f2328 !important; border-color:#d0d7de !important; }
[data-theme="light"] .CodeMirror-dialog input { color:#1f2328 !important; }
[data-theme="light"] .sidebar .search input { background:#fff; color:#1f2328; border-color:#d0d7de; }
[data-theme="light"] .tree-file { color:#656d76; }
[data-theme="light"] .tree-file:hover { color:#1f2328; background:rgba(0,0,0,.04); }
[data-theme="light"] .tree-file.active { color:#d4760a; }
[data-theme="light"] .tree-dir-label { color:#656d76; }
[data-theme="light"] .tree-dir-label:hover { color:#1f2328; background:rgba(0,0,0,.04); }
[data-theme="light"] .file-count { color:#8b949e; border-top-color:#d0d7de; }
[data-theme="light"] .toolbar button { color:#fff; }
[data-theme="light"] .toolbar .filename { color:#d4760a; }
[data-theme="light"] .toolbar h1 { color:#d4760a; }
[data-theme="light"] .toolbar .username { color:#1f2328; }
[data-theme="light"] .user-info .logout-btn { color:#656d76; }
[data-theme="light"] .user-info .logout-btn:hover { color:#1f2328; }
[data-theme="light"] .panels .editor-pane { border-right-color:#d0d7de; }
[data-theme="light"] .tabs button { color:#656d76; }
[data-theme="light"] .tabs button svg { stroke:#656d76; }
[data-theme="light"] .tabs button.active svg { stroke:#1f2328; }
[data-theme="light"] .tabs button.active { color:#1f2328; border-bottom-color:#d4760a; }
.theme-btn { background:transparent !important; padding:6px 8px !important; }
.theme-btn:hover { background:rgba(255,255,255,.1) !important; }

/* XH Modal System */
.xh-modal{background:var(--bg2);border:1px solid var(--border);border-radius:12px;padding:24px;box-shadow:0 20px 60px rgba(0,0,0,.4);animation:modalIn .15s ease-out}
@keyframes modalIn{from{opacity:0;transform:scale(.96) translateY(8px)}to{opacity:1;transform:scale(1) translateY(0)}}
.xh-modal h3{font-size:15px;font-weight:600;color:var(--text);margin-bottom:16px}
.xh-modal .xh-label{font-size:12px;color:var(--dim);margin-bottom:6px;font-weight:500;display:block}
.xh-modal .xh-input{width:100%;background:var(--bg);color:var(--text);border:1px solid var(--border);border-radius:8px;padding:8px 12px;font-size:13px;outline:none;transition:border-color .2s}
.xh-modal .xh-input:focus{border-color:var(--accent)}
.xh-modal .xh-input-mono{font-family:'JetBrains Mono',monospace;font-size:12px}
.xh-modal .xh-actions{display:flex;justify-content:flex-end;gap:8px;margin-top:20px}
.xh-modal .xh-btn{padding:7px 18px;border-radius:8px;border:none;cursor:pointer;font-size:13px;font-weight:500;transition:all .15s}
.xh-modal .xh-btn-ghost{background:transparent;color:var(--dim);border:1px solid var(--border)}
.xh-modal .xh-btn-ghost:hover{background:rgba(255,255,255,.04);color:var(--text)}
.xh-modal .xh-btn-primary{background:var(--accent);color:#fff}
.xh-modal .xh-btn-primary:hover{opacity:.9}
.xh-modal .xh-btn-danger{background:var(--red);color:#fff}
.xh-modal .xh-btn-success{background:var(--green);color:#fff}
.xh-modal .xh-card-select{display:flex;gap:8px}
.xh-modal .xh-card-select button{flex:1;display:flex;align-items:center;justify-content:center;gap:6px;padding:10px 12px;border-radius:8px;border:1.5px solid var(--border);background:transparent;color:var(--text);cursor:pointer;font-size:13px;transition:all .15s}
.xh-modal .xh-card-select button:hover{border-color:var(--dim)}
.xh-modal .xh-card-select button.sel{border-color:var(--accent);background:rgba(255,165,0,.08);color:var(--accent)}
.xh-modal .xh-link-box{display:flex;border-radius:8px;overflow:hidden;border:1px solid var(--border)}
.xh-modal .xh-link-box input{flex:1;padding:8px 12px;font-size:12px;font-family:monospace;background:var(--bg);color:var(--blue);border:none;outline:none;min-width:0}
.xh-modal .xh-link-box button{padding:8px 16px;background:var(--accent);color:#fff;border:none;cursor:pointer;font-size:12px;font-weight:600}
[data-theme="light"] .xh-modal{background:#fff;border-color:#d0d7de;box-shadow:0 20px 60px rgba(0,0,0,.12)}
[data-theme="light"] .xh-modal .xh-input{background:#f6f8fa;border-color:#d0d7de;color:#1f2328}
[data-theme="light"] .xh-modal .xh-btn-ghost{color:#656d76;border-color:#d0d7de}
[data-theme="light"] .xh-modal .xh-btn-ghost:hover{background:rgba(0,0,0,.04);color:#1f2328}

/* AI Assistant Floating Panel */
/* AI Assistant Floating Panel */
.ai-fab{position:fixed;bottom:24px;right:24px;width:56px;height:56px;border-radius:50%;background:linear-gradient(135deg,#7c3aed,#a855f7);color:#fff;border:none;cursor:pointer;font-size:28px;box-shadow:0 4px 20px rgba(124,58,237,.4);z-index:500;display:flex;align-items:center;justify-content:center;transition:transform .2s,box-shadow .2s}
.ai-fab:hover{transform:scale(1.1);box-shadow:0 6px 28px rgba(124,58,237,.6)}
.ai-panel{position:fixed;bottom:90px;right:24px;width:380px;max-height:500px;background:var(--surface);border:1px solid var(--border);border-radius:12px;box-shadow:0 8px 40px rgba(0,0,0,.5);z-index:500;display:none;flex-direction:column;overflow:hidden}
.ai-panel.open{display:flex}
.ai-panel-header{padding:12px 16px;border-bottom:1px solid var(--border);display:flex;align-items:center;justify-content:space-between}
.ai-panel-header h3{margin:0;font-size:14px;color:var(--fg)}
.ai-panel-header button{background:none;border:none;color:var(--dim);cursor:pointer;font-size:18px;padding:0 4px}
.ai-messages{flex:1;overflow-y:auto;padding:12px 16px;max-height:340px;display:flex;flex-direction:column;gap:10px}
.ai-msg{padding:8px 12px;border-radius:8px;font-size:13px;line-height:1.5;max-width:92%;word-wrap:break-word}
.ai-msg.user{background:rgba(124,58,237,.2);color:var(--fg);align-self:flex-end;border-bottom-right-radius:2px}
.ai-msg.ai{background:rgba(255,255,255,.06);color:var(--fg);align-self:flex-start;border-bottom-left-radius:2px}
.ai-msg.ai.loading{color:var(--dim);font-style:italic}
.ai-msg.ai.error{color:var(--red);background:rgba(255,0,0,.1)}
.ai-input-row{padding:10px 12px;border-top:1px solid var(--border);display:flex;gap:8px;align-items:center}
.ai-input-row input{flex:1;background:var(--bg);border:1px solid var(--border);border-radius:6px;padding:8px 12px;color:var(--fg);font-size:13px;outline:none}
.ai-input-row input:focus{border-color:var(--purple)}
.ai-input-row button{background:var(--purple);color:#fff;border:none;border-radius:6px;padding:8px 14px;cursor:pointer;font-size:13px;font-weight:600;white-space:nowrap}
.ai-input-row button:disabled{opacity:.5;cursor:not-allowed}
.ai-diff-actions{display:flex;gap:6px;margin-top:6px}
.ai-diff-actions button{font-size:11px;padding:4px 10px;border-radius:4px;border:none;cursor:pointer;font-weight:600}
.ai-diff-actions .accept{background:var(--green);color:#000}
.ai-diff-actions .reject{background:rgba(255,255,255,.1);color:var(--dim)}
/* Export dialog */
.export-row{margin-bottom:14px}
.export-row label{display:block;font-size:12px;color:var(--dim);margin-bottom:6px;font-weight:600}
.export-toggle{display:flex;gap:6px}
.export-toggle button{flex:1;padding:10px 8px;border-radius:6px;border:2px solid var(--border);background:var(--bg);color:var(--text);cursor:pointer;font-size:13px;transition:border-color .2s,background .2s}
.export-toggle button:hover{border-color:#e67e22}
.export-toggle button.sel{border-color:#e67e22;background:rgba(230,126,34,.12);color:#e67e22;font-weight:600}
.export-actions{display:flex;justify-content:flex-end;gap:8px;margin-top:18px}
.export-actions button{padding:8px 18px;border-radius:6px;border:none;cursor:pointer;font-size:13px;font-weight:600}
.export-spinner{display:inline-block;width:14px;height:14px;border:2px solid rgba(255,255,255,.3);border-top-color:#fff;border-radius:50%;animation:spin .6s linear infinite;margin-right:6px;vertical-align:middle}
@keyframes spin{to{transform:rotate(360deg)}}
/* Auth / Admin */
.toolbar .user-info{font-size:12px;color:var(--dim);display:flex;align-items:center;gap:4px}
.toolbar .user-info .username{color:var(--accent)}
.toolbar .logout-btn{background:none;border:1px solid var(--border);color:var(--dim);padding:4px 8px;border-radius:4px;font-size:12px;cursor:pointer}
.toolbar .logout-btn:hover{color:var(--text);border-color:var(--text)}
.toolbar .admin-btn{background:var(--purple);color:#fff;padding:5px 10px;border-radius:6px;font-size:12px;cursor:pointer;border:none;font-weight:600}
.toolbar .admin-btn:hover{background:#a371f7}
.admin-overlay{position:fixed;inset:0;background:rgba(0,0,0,.7);z-index:500;display:flex;align-items:center;justify-content:center}
.admin-panel{background:var(--bg2);border:1px solid var(--purple);border-radius:12px;width:700px;max-width:95vw;max-height:85vh;display:flex;flex-direction:column;overflow:hidden}
.admin-header{padding:16px 20px;border-bottom:1px solid var(--border);display:flex;align-items:center;justify-content:space-between}
.admin-header h2{font-size:16px;color:var(--purple);margin:0}
.admin-header button{background:none;border:none;color:var(--dim);font-size:20px;cursor:pointer}
.admin-body{flex:1;overflow-y:auto;padding:16px 20px}
.admin-table{width:100%;border-collapse:collapse;font-size:13px}
.admin-table th,.admin-table td{padding:10px 12px;border-bottom:1px solid var(--border);text-align:left}
.admin-table th{color:var(--dim);font-weight:600;font-size:11px;text-transform:uppercase}
.admin-table td{color:var(--text)}
.admin-table .role-badge{padding:2px 8px;border-radius:4px;font-size:11px;font-weight:600}
.admin-table .role-admin{background:rgba(137,87,229,.2);color:var(--purple)}
.admin-table .role-editor{background:rgba(88,166,255,.2);color:var(--blue)}
.admin-table .role-viewer{background:rgba(63,185,80,.2);color:var(--green)}
.admin-table .action-btn{background:none;border:1px solid var(--border);color:var(--dim);padding:4px 8px;border-radius:4px;font-size:11px;cursor:pointer;margin-right:4px}
.admin-table .action-btn:hover{color:var(--text);border-color:var(--text)}
.admin-table .action-btn.danger{border-color:var(--red);color:var(--red)}
.admin-table .action-btn.danger:hover{background:rgba(248,81,73,.1)}
.admin-form{background:var(--bg);border:1px solid var(--border);border-radius:8px;padding:16px;margin-top:16px}
.admin-form h3{font-size:14px;color:var(--text);margin-bottom:12px}
.admin-form label{display:block;font-size:12px;color:var(--dim);margin-bottom:4px;font-weight:600}
.admin-form input,.admin-form select{width:100%;background:var(--bg2);color:var(--text);border:1px solid var(--border);border-radius:4px;padding:8px;font-size:13px;margin-bottom:10px}
.admin-form input:focus,.admin-form select:focus{outline:none;border-color:var(--purple)}
.admin-form .form-row{display:flex;gap:12px}
.admin-form .form-row>div{flex:1}
.admin-form .form-actions{display:flex;gap:8px;justify-content:flex-end;margin-top:8px}
.admin-form .form-actions button{padding:6px 16px;border-radius:6px;border:none;cursor:pointer;font-size:13px;font-weight:600}

/* AI panel light mode */
[data-theme="light"] .ai-fab{box-shadow:0 4px 20px rgba(124,58,237,.25)}
[data-theme="light"] .ai-panel{background:var(--bg2);box-shadow:0 8px 40px rgba(0,0,0,.15)}
[data-theme="light"] .ai-msg.user{background:rgba(124,58,237,.12)}
[data-theme="light"] .ai-msg.ai{background:rgba(0,0,0,.04)}
[data-theme="light"] .ai-input-row input{background:var(--bg)}
/* Toast styling */
.toast-msg{position:fixed;top:16px;right:16px;z-index:9999;padding:10px 18px;border-radius:8px;font-size:13px;font-weight:500;color:#fff;backdrop-filter:blur(8px);animation:toastIn .2s ease-out;max-width:360px;box-shadow:0 4px 16px rgba(0,0,0,.2)}
.toast-msg.success{background:rgba(63,185,80,.9)}
.toast-msg.error{background:rgba(248,81,73,.9)}
.toast-msg.warn{background:rgba(210,153,34,.9)}
@keyframes toastIn{from{opacity:0;transform:translateY(-8px)}to{opacity:1;transform:translateY(0)}}


/* Recent files */
.recent-section{padding:6px 8px;border-bottom:1px solid var(--border)}
.recent-title{font-size:10px;text-transform:uppercase;letter-spacing:.5px;color:var(--dim);margin-bottom:4px;display:flex;align-items:center;gap:4px}
.recent-item{display:flex;align-items:center;gap:6px;padding:3px 6px;border-radius:4px;cursor:pointer;font-size:12px;color:var(--text);text-overflow:ellipsis;overflow:hidden;white-space:nowrap}
.recent-item:hover{background:rgba(255,255,255,.06)}
.recent-item svg{flex-shrink:0;opacity:.5}
[data-theme="light"] .recent-item:hover{background:rgba(0,0,0,.04)}
/* File stats in status area */
.file-stats{display:flex;gap:12px;font-size:11px;color:var(--dim);align-items:center}
.file-stats span{display:flex;align-items:center;gap:3px}
</style>
</head>
<body>
__READONLY_BANNER__
<div class="toolbar">
<h1>__EDITOR_TITLE__</h1>
<span class="sep"></span>
<span class="filename" id="fname">__INITIAL_FNAME__</span>
<span class="spacer"></span>
<span class="status" id="status" style="display:none"></span>
<div class="file-stats" id="fileStats" style="display:none"><span id="statWords"></span><span id="statChars"></span><span id="statRead"></span></div>
<button onclick="save()" id="saveBtn" __SAVE_DISPLAY__ title="保存"><svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M15.2 3a2 2 0 0 1 1.4.6l3.8 3.8a2 2 0 0 1 .6 1.4V19a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2V5a2 2 0 0 1 2-2z"/><path d="M17 21v-7a1 1 0 0 0-1-1H8a1 1 0 0 0-1 1v7"/><path d="M7 3v4a1 1 0 0 0 1 1h7"/></svg></button>
<span class="sep" __SHARE_BTN_DISPLAY__></span>
<button onclick="showShareDialog()" id="shareBtn" __SHARE_BTN_DISPLAY__ title="分享"><svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><circle cx="18" cy="5" r="3"/><circle cx="6" cy="12" r="3"/><circle cx="18" cy="19" r="3"/><line x1="8.59" y1="13.51" x2="15.42" y2="17.49"/><line x1="15.41" y1="6.51" x2="8.59" y2="10.49"/></svg></button>
<button onclick="showExportDialog()" id="exportBtn" title="导出"><svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg></button>
<span class="sep"></span>
<button class="ann-btn" onclick="toggleAnnotationMode()" id="annBtn" __ANN_BTN_DISPLAY__ title="标注"><svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M21 15a2 2 0 0 1-2 2H7l-4 4V5a2 2 0 0 1 2-2h14a2 2 0 0 1 2 2z"/><line x1="12" y1="8" x2="12" y2="14"/><line x1="9" y1="11" x2="15" y2="11"/></svg></button>
<span class="badge" id="annCount" style="display:none">0</span>
__THEME_BTN__
__ADMIN_BTN__
__LOGOUT_BTN__
</div>
<div class="sidebar-overlay" id="sidebarOverlay" onclick="toggleMobileSidebar()"></div>
<button class="sidebar-toggle" id="sidebarToggle" onclick="toggleMobileSidebar()"><svg width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="m6 14 1.5-2.9A2 2 0 0 1 9.24 10H20a2 2 0 0 1 1.94 2.5l-1.54 6a2 2 0 0 1-1.95 1.5H4a2 2 0 0 1-2-2V5a2 2 0 0 1 2-2h3.9a2 2 0 0 1 1.69.9l.81 1.2a2 2 0 0 0 1.67.9H18a2 2 0 0 1 2 2v2"/></svg></button>
<div class="main">
__SIDEBAR_HTML__
<div class="content">
<div class="tabs">
<button class="active" onclick="setView('split',this)" title="分屏"><svg width="13" height="13" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><rect width="18" height="18" x="3" y="3" rx="2"/><path d="M12 3v18"/></svg></button>
<button onclick="setView('edit-only',this)" title="编辑"><svg width="13" height="13" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M21.174 6.812a1 1 0 0 0-3.986-3.987L3.842 16.174a2 2 0 0 0-.5.83l-1.321 4.352a.5.5 0 0 0 .623.622l4.353-1.32a2 2 0 0 0 .83-.497z"/></svg></button>
<button onclick="setView('preview-only',this)" title="预览"><svg width="13" height="13" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M2.062 12.348a1 1 0 0 1 0-.696 10.75 10.75 0 0 1 19.876 0 1 1 0 0 1 0 .696 10.75 10.75 0 0 1-19.876 0"/><circle cx="12" cy="12" r="3"/></svg></button>
</div>
<div class="panels split" id="panels">
<div class="editor-pane" id="editorPane"></div>
<div class="preview-pane" id="preview"><div style="display:flex;align-items:center;justify-content:center;height:100%;color:#484f58">__EMPTY_MSG__</div></div>
</div>
</div>
</div>

<!-- Annotation bar (hidden by default) -->
<div class="ann-bar" id="annBar" style="display:none">
<div class="ann-header">
<span class="ann-badge" id="annTotal">0</span>
<span class="ann-hint">选中文字后添加标注</span>
<button class="ann-copy-btn" onclick="submitAnnotations()"><svg width="13" height="13" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><rect width="14" height="14" x="8" y="8" rx="2" ry="2"/><path d="M4 16c-1.1 0-2-.9-2-2V4c0-1.1.9-2 2-2h10c1.1 0 2 .9 2 2"/></svg> 复制全部</button>
</div>
<div id="annListItems"></div>
</div>

<script>
const BASE=(()=>{const p=location.pathname;return p.endsWith('/')?p:p+'/';})();
function api(ep){return BASE+ep;}
const FULL_MODE=__FULL_MODE__;
const SHARE_FILE=__SHARE_FILE__;
const IS_READONLY=__IS_READONLY__;
const AUTH_USER=__AUTH_USER__;

// Mobile sidebar toggle
function toggleMobileSidebar(){
  const sb=document.querySelector('.sidebar');
  const ov=document.getElementById('sidebarOverlay');
  if(!sb)return;
  sb.classList.toggle('mobile-open');
  ov.classList.toggle('visible');
}
// Close sidebar when a file is clicked on mobile
function closeMobileSidebar(){
  if(window.innerWidth<=768){
    const sb=document.querySelector('.sidebar');
    const ov=document.getElementById('sidebarOverlay');
    if(sb)sb.classList.remove('mobile-open');
    if(ov)ov.classList.remove('visible');
  }
}

// --- Theme ---
function initTheme(){
  const saved=localStorage.getItem('md-editor-theme')||'dark';
  document.documentElement.setAttribute('data-theme',saved);
  updateThemeIcon(saved);
  updateCMTheme(saved);
}
function toggleTheme(){
  const cur=document.documentElement.getAttribute('data-theme');
  const next=cur==='dark'?'light':'dark';
  document.documentElement.setAttribute('data-theme',next);
  localStorage.setItem('md-editor-theme',next);
  updateThemeIcon(next);
  updateCMTheme(next);
}
function updateThemeIcon(theme){
  const el=document.getElementById('themeIcon');
  if(!el)return;
  if(theme==='light'){
    el.innerHTML='<circle cx="12" cy="12" r="4"/><path d="M12 2v2"/><path d="M12 20v2"/><path d="m4.93 4.93 1.41 1.41"/><path d="m17.66 17.66 1.41 1.41"/><path d="M2 12h2"/><path d="M20 12h2"/><path d="m6.34 17.66-1.41 1.41"/><path d="m19.07 4.93-1.41 1.41"/>';
  } else {
    el.innerHTML='<path d="M12 3a6 6 0 0 0 9 9 9 9 0 1 1-9-9Z"/>';
  }
}
function updateCMTheme(theme){
  if(!window.cm)return;
  // Theme handled by CSS overrides - keep material-darker loaded
  // Toggle highlight.js theme
  const hljsDark=document.querySelector('link[href*="github-dark"]');
  const hljsLight=document.getElementById('hljs-light');
  if(hljsDark&&hljsLight){hljsDark.disabled=(theme==='light');hljsLight.disabled=(theme==='dark');}
}
// initTheme() moved after CM init

// --- Modal helper (DaisyUI native dialog) ---
function openModal(id,html,opts={}){
  let d=document.getElementById(id);
  if(d)d.remove();
  const w=opts.wide?'580px':opts.narrow?'340px':'440px';
  d=document.createElement('div');
  d.id=id;
  d.style.cssText='position:fixed;inset:0;background:rgba(0,0,0,.5);z-index:500;display:flex;align-items:center;justify-content:center;backdrop-filter:blur(2px)';
  d.innerHTML='<div class="xh-modal" style="width:'+w+';max-width:calc(100vw - 32px);max-height:80vh;overflow-y:auto">'+html+'</div>';
  d.addEventListener('click',e=>{if(e.target===d)closeModal(id);});
  document.addEventListener('keydown',function esc(e){if(e.key==='Escape'){closeModal(id);document.removeEventListener('keydown',esc);}});
  document.body.appendChild(d);
  return d;
}
function closeModal(id){const d=document.getElementById(id);if(d)d.remove();}

// --- File mode detection ---
function getCMMode(ext){
  const map={
    'md':'gfm','markdown':'gfm',
    'json':'application/json','jsonl':'application/json',
    'js':'javascript','mjs':'javascript','cjs':'javascript',
    'ts':'text/typescript','tsx':'text/typescript',
    'html':'htmlmixed','htm':'htmlmixed',
    'xml':'xml','svg':'xml',
    'css':'css','scss':'text/x-scss','less':'text/x-less',
    'py':'python',
    'yaml':'yaml','yml':'yaml',
    'sh':'shell','bash':'shell','zsh':'shell',
    'sql':'sql',
    'toml':'toml',
    'txt':'gfm','csv':'gfm','log':'gfm','env':'gfm'
  };
  return map[ext]||'gfm';
}

// --- Mermaid ---
mermaid.initialize({startOnLoad:false,theme:'base',
  themeCSS:`
    rect { fill: #2d333b !important; stroke: #444c56 !important; }
    circle { fill: #2d333b !important; stroke: #444c56 !important; }
    polygon { fill: #2d333b !important; stroke: #444c56 !important; }
    path.flowchart-link { stroke: #58a6ff !important; fill: none !important; }
    .edgePath .path { stroke: #58a6ff !important; }
    marker path { fill: #58a6ff !important; stroke: none !important; }
    text, tspan { fill: #e6edf3 !important; }
    .nodeLabel, .label div, .label span, foreignObject div { color: #e6edf3 !important; }
    .cluster rect { fill: #1c2128 !important; stroke: #ffa500 !important; }
    .cluster-label .nodeLabel { color: #ffa500 !important; }
    .edgeLabel { background: transparent !important; }
  `,
  themeVariables:{
    primaryColor:'#2d333b',primaryTextColor:'#e6edf3',primaryBorderColor:'#ffa500',
    secondaryColor:'#1c2128',secondaryTextColor:'#e6edf3',secondaryBorderColor:'#58a6ff',
    tertiaryColor:'#161b22',tertiaryTextColor:'#e6edf3',tertiaryBorderColor:'#8957e5',
    lineColor:'#58a6ff',textColor:'#e6edf3',
    mainBkg:'#2d333b',nodeBorder:'#ffa500',nodeTextColor:'#e6edf3',
    clusterBkg:'#1c2128',clusterBorder:'#30363d',titleColor:'#ffa500',
    edgeLabelBackground:'#161b22',
    labelTextColor:'#e6edf3',labelBoxBkgColor:'#2d333b',labelBoxBorderColor:'#30363d',
    noteBkgColor:'#2d333b',noteTextColor:'#e6edf3',noteBorderColor:'#ffa500',
    actorTextColor:'#e6edf3',actorBkg:'#2d333b',actorBorder:'#ffa500',actorLineColor:'#58a6ff',
    signalColor:'#e6edf3',signalTextColor:'#e6edf3',
    background:'#0d1117',
    fillType0:'#2d333b',fillType1:'#1c2128',fillType2:'#161b22',fillType3:'#2d333b',fillType4:'#1c2128',fillType5:'#161b22',fillType6:'#2d333b',fillType7:'#1c2128'
  }
});

// --- Marked ---
const _r=new marked.Renderer();
_r.code=function({text,lang}){
  if(lang==='mermaid')return'<div class="mermaid-wrap" onclick="zoomMermaid(this)"><div class="mermaid-src" style="display:none">'+text.replace(/</g,'&lt;')+'</div><div class="mermaid-render"></div><span class="zoom-hint"><svg width="12" height="12" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><circle cx="11" cy="11" r="8"/><line x1="21" y1="21" x2="16.65" y2="16.65"/><line x1="11" y1="8" x2="11" y2="14"/><line x1="8" y1="11" x2="14" y2="11"/></svg> 点击放大</span></div>';
  let h=text;
  try{if(lang&&hljs.getLanguage(lang))h=hljs.highlight(text,{language:lang}).value;else h=hljs.highlightAuto(text).value;}catch(e){}
  return'<pre><code class="hljs">'+h+'</code></pre>';
};
marked.use({renderer:_r,gfm:true,breaks:true});

// --- CodeMirror ---
const cm=CodeMirror(document.getElementById('editorPane'),{
  value:'',mode:'gfm',theme:'material-darker',lineNumbers:true,lineWrapping:true,
  styleActiveLine:true,matchBrackets:true,autoCloseBrackets:true,
  foldGutter:true,gutters:['CodeMirror-linenumbers','CodeMirror-foldgutter','annotation-gutter'],
  indentUnit:2,tabSize:2,indentWithTabs:false,
  extraKeys:{'Cmd-S':(e)=>{e&&e.preventDefault&&e.preventDefault();if(!IS_READONLY)save();},'Ctrl-S':(e)=>{e&&e.preventDefault&&e.preventDefault();if(!IS_READONLY)save();},'Tab':cm=>cm.replaceSelection('  ')}
});

// Apply saved theme now that CM is ready
initTheme();
updateCMTheme(localStorage.getItem('md-editor-theme')||'dark');

let dirty=false,previewTimer=null,currentFile=SHARE_FILE;
let _previewLocked=false;

// ==================== RECENT FILES ====================
const RECENT_KEY='md-editor-recent';
function getRecent(){try{return JSON.parse(localStorage.getItem(RECENT_KEY))||[];}catch(e){return[];}}
function addRecent(path){
  let list=getRecent().filter(p=>p!==path);
  list.unshift(path);
  if(list.length>10)list=list.slice(0,10);
  localStorage.setItem(RECENT_KEY,JSON.stringify(list));
  renderRecent();
}
function renderRecent(){
  const list=getRecent();
  const sec=document.getElementById('recentSection');
  const cont=document.getElementById('recentList');
  if(!sec||!cont)return;
  if(!list.length){sec.style.display='none';return;}
  sec.style.display='';
  cont.innerHTML=list.slice(0,5).map(p=>{
    const name=p.split('/').pop();
    return '<div class="recent-item" onclick="openFile(\''+p.replace(/'/g,"\\'")+'\')"><svg width="12" height="12" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M15 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V7Z"/><path d="M14 2v4a2 2 0 0 0 2 2h4"/></svg>'+name+'</div>';
  }).join('');
}
renderRecent();

// ==================== FILE STATS ====================
function updateFileStats(){
  const statsEl=document.getElementById('fileStats');
  if(!statsEl||!currentFile)return;
  const text=cm.getValue();
  if(!text){statsEl.style.display='none';return;}
  statsEl.style.display='flex';
  // Chinese + English word count
  const cn=(text.match(/[\u4e00-\u9fff\u3400-\u4dbf]/g)||[]).length;
  const en=(text.match(/[a-zA-Z]+/g)||[]).length;
  const words=cn+en;
  const chars=text.length;
  const mins=Math.max(1,Math.ceil(words/300));
  document.getElementById('statWords').textContent=words+' words';
  document.getElementById('statChars').textContent=chars+' chars';
  document.getElementById('statRead').textContent='~'+mins+' min';
}

let annotationMode=false;
// annotations: array of {id, startLine, endLine, comment}
let annotations=[];
let annIdCounter=0;

cm.on('change',()=>{
  if(_previewLocked)return;
  dirty=true;
  clearTimeout(window._statsTimer);window._statsTimer=setTimeout(updateFileStats,1000);
  document.getElementById('saveBtn').style.color='var(--red)';
  document.getElementById('status').style.color='var(--red)';
  clearTimeout(previewTimer);
  previewTimer=setTimeout(renderPreview,300);
});

// Show "add annotation" button when user selects text in annotation mode
cm.on('cursorActivity',()=>{
  if(!annotationMode)return;
  removeSelBtn();
  const sel=cm.listSelections()[0];
  if(!sel)return;
  const from=sel.anchor,to=sel.head;
  const startLine=Math.min(from.line,to.line);
  const endLine=Math.max(from.line,to.line);
  if(startLine===endLine&&from.ch===to.ch)return; // no selection
  // Show floating button
  const coords=cm.charCoords(to,'page');
  const btn=document.createElement('button');btn.className='ann-sel-btn';btn.id='annSelBtn';
  btn.innerHTML='<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M21 15a2 2 0 0 1-2 2H7l-4 4V5a2 2 0 0 1 2-2h14a2 2 0 0 1 2 2z"/><line x1="12" y1="8" x2="12" y2="14"/><line x1="9" y1="11" x2="15" y2="11"/></svg> '+((endLine-startLine)+1)+'行';
  btn.style.left=Math.min(Math.max(10,coords.left),window.innerWidth-200)+'px';
  btn.style.top=Math.min(coords.bottom+4,window.innerHeight-50)+'px';
  btn.onclick=()=>{removeSelBtn();showRangePopup(startLine,endLine);};
  document.body.appendChild(btn);
});

function removeSelBtn(){const b=document.getElementById('annSelBtn');if(b)b.remove();}

// Click gutter marker to edit existing annotation
cm.on('gutterClick',(cm,line,gutter)=>{
  if(gutter!=='annotation-gutter'||!annotationMode)return;
  const ann=annotations.find(a=>line>=a.startLine&&line<=a.endLine);
  if(ann)showRangePopup(ann.startLine,ann.endLine,ann);
  else showRangePopup(line,line);
});

// ==================== ANNOTATION SYSTEM ====================

function toggleAnnotationMode(){
  annotationMode=!annotationMode;
  const btn=document.getElementById('annBtn');
  btn.classList.toggle('active-mode',annotationMode);
  btn.innerHTML=annotationMode?'<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M21 15a2 2 0 0 1-2 2H7l-4 4V5a2 2 0 0 1 2-2h14a2 2 0 0 1 2 2z"/><line x1="12" y1="8" x2="12" y2="14"/><line x1="9" y1="11" x2="15" y2="11"/></svg> 标注中...':'<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M21 15a2 2 0 0 1-2 2H7l-4 4V5a2 2 0 0 1 2-2h14a2 2 0 0 1 2 2z"/><line x1="12" y1="8" x2="12" y2="14"/><line x1="9" y1="11" x2="15" y2="11"/></svg> 标注';
  document.getElementById('annBar').style.display=annotationMode?'flex':'none';
  if(!annotationMode)removeSelBtn();
  refreshAnnotationDisplay();
}

function refreshAnnotationDisplay(){
  cm.clearGutter('annotation-gutter');
  // Clear all ann-line backgrounds
  for(let i=0;i<cm.lineCount();i++)cm.removeLineClass(i,'background','ann-line');
  
  annotations.forEach(ann=>{
    // Gutter marker on start line
    const m=document.createElement('div');m.className='ann-marker';m.textContent='●';
    m.title=ann.comment;
    cm.setGutterMarker(ann.startLine,'annotation-gutter',m);
    // Highlight range
    for(let l=ann.startLine;l<=ann.endLine;l++){
      cm.addLineClass(l,'background','ann-line');
    }
  });
  
  updateAnnotationCount();
  renderAnnotationList();
}

function updateAnnotationCount(){
  const count=annotations.length;
  document.getElementById('annCount').textContent=count;
  document.getElementById('annCount').style.display=count?'':'none';
  document.getElementById('annTotal').textContent=count;
}

function renderAnnotationList(){
  const container=document.getElementById('annListItems');
  if(!container)return;
  container.innerHTML='';
  if(!annotations.length){container.innerHTML='';return;}
  annotations.forEach((ann,i)=>{
    const range=ann.startLine===ann.endLine?'第'+(ann.startLine+1)+'行':'第'+(ann.startLine+1)+'-'+(ann.endLine+1)+'行';
    const d=document.createElement('div');d.className='ann-list-item';
    d.innerHTML='<span class="ann-range-tag">'+range+'</span><span class="ann-comment">'+escHtml(ann.comment)+'</span><span class="ann-del" onclick="deleteAnnotation('+ann.id+')" title="删除">✕</span>';
    d.style.cursor='pointer';
    d.querySelector('.ann-range-tag').onclick=()=>{cm.scrollIntoView({line:ann.startLine,ch:0},100);cm.setSelection({line:ann.startLine,ch:0},{line:ann.endLine,ch:cm.getLine(ann.endLine).length});};
    container.appendChild(d);
  });
}

function escHtml(s){return s.replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');}

function showRangePopup(startLine,endLine,existingAnn){
  document.querySelectorAll('.ann-popup').forEach(p=>p.remove());
  removeSelBtn();
  const popup=document.createElement('div');popup.className='ann-popup';
  
  // Preview selected text
  let previewLines=[];
  for(let l=startLine;l<=endLine&&l<startLine+5;l++){
    previewLines.push((l+1)+': '+(cm.getLine(l)||''));
  }
  if(endLine-startLine>=5)previewLines.push('...(共'+(endLine-startLine+1)+'行)');
  const previewText=previewLines.join('\n');
  
  const rangeLabel=startLine===endLine?'第 '+(startLine+1)+' 行':'第 '+(startLine+1)+' - '+(endLine+1)+' 行 ('+(endLine-startLine+1)+'行)';
  const existing=existingAnn?escHtml(existingAnn.comment):'';
  
  popup.innerHTML=
    '<div class="ann-range">'+rangeLabel+'</div>'+
    '<div class="ann-preview">'+escHtml(previewText)+'</div>'+
    '<textarea placeholder="写你的修改意见（如：这段全删、改成XX、语气太硬...）">'+existing+'</textarea>'+
    '<div class="ann-actions">'+
    (existingAnn?'<button class="delete" onclick="deleteAnnotation('+existingAnn.id+');closePopup()">删除</button>':'')+
    '<button class="cancel" onclick="closePopup()">取消</button>'+
    '<button class="save" id="annSaveBtn">保存</button>'+
    '</div>';
  
  // Position popup: near selection start, but always visible
  const coords=cm.charCoords({line:startLine,ch:0},'page');
  let popLeft=Math.min(coords.left+60,window.innerWidth-400);
  let popTop=coords.top-20;
  // If selection is large or popup would be off-screen, center it
  if(popTop<10||popTop>window.innerHeight-280||(endLine-startLine)>20){
    popTop=Math.max(60,window.innerHeight/2-150);
    popLeft=Math.max(20,window.innerWidth/2-190);
  }
  popup.style.left=Math.max(10,popLeft)+'px';
  popup.style.top=Math.max(10,popTop)+'px';
  document.body.appendChild(popup);
  
  const ta=popup.querySelector('textarea');ta.focus();
  
  popup.querySelector('#annSaveBtn').onclick=()=>{
    const text=ta.value.trim();
    if(!text){closePopup();return;}
    if(existingAnn){
      existingAnn.comment=text;
    }else{
      annotations.push({id:annIdCounter++,startLine,endLine,comment:text});
    }
    closePopup();
    refreshAnnotationDisplay();
  };
  
  ta.addEventListener('keydown',e=>{
    if((e.ctrlKey||e.metaKey)&&e.key==='Enter'){popup.querySelector('#annSaveBtn').click();e.preventDefault();}
    if(e.key==='Escape')closePopup();
  });
}

window.closePopup=function(){document.querySelectorAll('.ann-popup').forEach(p=>p.remove());};

window.deleteAnnotation=function(id){
  annotations=annotations.filter(a=>a.id!==id);
  refreshAnnotationDisplay();
};

// Submit annotations to server
function submitAnnotations(){
  if(!annotations.length){showToast('⚠️ 还没有标注','var(--accent)');return;}
  
  const fileName=currentFile||'未知文件';
  const ts=new Date().toISOString();
  let text='📌 标注反馈 — '+fileName+'\n时间: '+ts+'\n';
  
  annotations.forEach((ann,i)=>{
    let lines=[];
    for(let l=ann.startLine;l<=ann.endLine;l++)lines.push(cm.getLine(l)||'');
    const selectedText=lines.join('\n');
    const preview=lines.slice(0,3).join('\n');
    const ellipsis=lines.length>3?'...':'';
    
    text+='\n---\n【标注 '+(i+1)+'】第 '+(ann.startLine+1)+'-'+(ann.endLine+1)+' 行\n';
    text+='> '+preview.split('\n').join('\n> ')+ellipsis+'\n';
    text+='💬 '+ann.comment+'\n';
  });
  
  const count=annotations.length;
  
  function onSuccess(){
    showToast('📋 已复制 '+count+' 条标注到剪贴板');
    annotations=[];
    refreshAnnotationDisplay();
    toggleAnnotationMode();
  }
  
  function fallbackCopy(str){
    const ta=document.createElement('textarea');
    ta.value=str;
    ta.style.cssText='position:fixed;left:-9999px;top:-9999px;opacity:0';
    document.body.appendChild(ta);
    ta.select();
    try{document.execCommand('copy');onSuccess();}
    catch(e){showToast('❌ 复制失败，请手动复制','var(--red)');}
    finally{ta.remove();}
  }
  
  if(navigator.clipboard&&navigator.clipboard.writeText){
    navigator.clipboard.writeText(text).then(onSuccess).catch(()=>fallbackCopy(text));
  }else{
    fallbackCopy(text);
  }
}

function showToast(msg,bg){
  document.querySelectorAll('.toast-msg').forEach(t=>t.remove());
  const cls=bg&&bg.includes('red')?'error':bg&&bg.includes('accent')?'warn':'success';
  const t=document.createElement('div');
  t.className='toast-msg '+cls;
  t.textContent=msg.replace(/^[\u2705\u274c\u26a0\ufe0f\ud83d\udccb\ud83d\uddd1\ufe0f\u2728]+ ?/,'');
  document.body.appendChild(t);
  setTimeout(()=>{t.style.opacity='0';t.style.transition='opacity .3s';setTimeout(()=>t.remove(),300);},2500);
}

// ==================== MERMAID ZOOM ====================

let mmId=0;
window.zoomMermaid=function(wrap){
  const svg=wrap.querySelector('.mermaid-render svg');
  if(!svg)return;
  const lb=document.createElement('div');lb.className='mermaid-lightbox';
  lb.innerHTML=`
    <div class="lb-close" onclick="this.parentElement.remove()">✕</div>
    <div class="lb-zoom-controls">
      <button onclick="lbZoom(0.3)">＋</button>
      <span id="lbZoomLevel">100%</span>
      <button onclick="lbZoom(-0.3)">－</button>
      <button onclick="lbResetZoom()">↺</button>
    </div>
    <div class="lb-viewport" id="lbViewport">
      <div class="lb-content" id="lbContent">`+svg.outerHTML+`</div>
    </div>`;
  lb.addEventListener('click',e=>{if(e.target===lb)lb.remove();});
  document.addEventListener('keydown',function esc(e){if(e.key==='Escape'){lb.remove();document.removeEventListener('keydown',esc);}});
  document.body.appendChild(lb);
  
  const lbSvg=lb.querySelector('svg');
  if(lbSvg){lbSvg.style.width='100%';lbSvg.style.height='auto';lbSvg.removeAttribute('height');lbSvg.removeAttribute('width');}
  forceDarkMermaid(lb);
  
  // Zoom state
  let scale=1,posX=0,posY=0,isDragging=false,startX=0,startY=0;
  const content=document.getElementById('lbContent');
  const viewport=document.getElementById('lbViewport');
  
  function applyTransform(){
    content.style.transform='translate('+posX+'px,'+posY+'px) scale('+scale+')';
    document.getElementById('lbZoomLevel').textContent=Math.round(scale*100)+'%';
  }
  
  // Scroll wheel zoom
  viewport.addEventListener('wheel',e=>{
    e.preventDefault();
    const delta=e.deltaY>0?-0.15:0.15;
    scale=Math.max(0.2,Math.min(10,scale+delta));
    applyTransform();
  },{passive:false});
  
  // Drag to pan
  viewport.addEventListener('mousedown',e=>{
    if(e.target.closest('.lb-zoom-controls'))return;
    isDragging=true;startX=e.clientX-posX;startY=e.clientY-posY;
    viewport.style.cursor='grabbing';
  });
  viewport.addEventListener('mousemove',e=>{
    if(!isDragging)return;
    posX=e.clientX-startX;posY=e.clientY-startY;
    applyTransform();
  });
  viewport.addEventListener('mouseup',()=>{isDragging=false;viewport.style.cursor='grab';});
  viewport.addEventListener('mouseleave',()=>{isDragging=false;viewport.style.cursor='grab';});
  
  // Touch: pinch zoom + drag
  let lastTouchDist=0,lastTouchX=0,lastTouchY=0;
  viewport.addEventListener('touchstart',e=>{
    if(e.touches.length===2){
      lastTouchDist=Math.hypot(e.touches[0].clientX-e.touches[1].clientX,e.touches[0].clientY-e.touches[1].clientY);
    }else if(e.touches.length===1){
      isDragging=true;lastTouchX=e.touches[0].clientX-posX;lastTouchY=e.touches[0].clientY-posY;
    }
  },{passive:true});
  viewport.addEventListener('touchmove',e=>{
    e.preventDefault();
    if(e.touches.length===2){
      const dist=Math.hypot(e.touches[0].clientX-e.touches[1].clientX,e.touches[0].clientY-e.touches[1].clientY);
      scale=Math.max(0.2,Math.min(10,scale*(dist/lastTouchDist)));
      lastTouchDist=dist;applyTransform();
    }else if(e.touches.length===1&&isDragging){
      posX=e.touches[0].clientX-lastTouchX;posY=e.touches[0].clientY-lastTouchY;
      applyTransform();
    }
  },{passive:false});
  viewport.addEventListener('touchend',()=>{isDragging=false;lastTouchDist=0;});
  
  // Global zoom functions
  window.lbZoom=function(d){scale=Math.max(0.2,Math.min(10,scale+d));applyTransform();};
  window.lbResetZoom=function(){scale=1;posX=0;posY=0;applyTransform();};
  
  // Start at 150% for better visibility
  scale=1.5;applyTransform();
};

function forceDarkMermaid(container){
  const isLight=document.documentElement.getAttribute('data-theme')==='light';
  const palette=isLight?[
    {fill:'#dbeafe',stroke:'#2563eb'},  // blue
    {fill:'#ede9fe',stroke:'#7c3aed'},  // purple
    {fill:'#dcfce7',stroke:'#16a34a'},  // green
    {fill:'#fef3c7',stroke:'#d97706'},  // amber
    {fill:'#fce7f3',stroke:'#db2777'},  // pink
    {fill:'#ccfbf1',stroke:'#0d9488'},  // teal
    {fill:'#fee2e2',stroke:'#dc2626'},  // red
    {fill:'#f3f4f6',stroke:'#6b7280'},  // grey
  ]:[
    {fill:'#1e3a5f',stroke:'#58a6ff'},  // deep blue
    {fill:'#2d1f4e',stroke:'#a371f7'},  // deep purple
    {fill:'#1a3c2e',stroke:'#3fb950'},  // deep green
    {fill:'#3c2a1a',stroke:'#ffa500'},  // deep amber
    {fill:'#3a1a2d',stroke:'#f778ba'},  // deep pink
    {fill:'#1a3636',stroke:'#56d4dd'},  // deep teal
    {fill:'#3c1a1a',stroke:'#f85149'},  // deep red
    {fill:'#2d333b',stroke:'#768390'},  // neutral grey
  ];
  let nodeIdx=0;
  
  // Assign colors to nodes
  container.querySelectorAll('.node, .flowchart-node').forEach(node=>{
    const c=palette[nodeIdx%palette.length];
    node.querySelectorAll('rect,circle,polygon,ellipse,path').forEach(shape=>{
      if(shape.closest('marker'))return;
      shape.style.setProperty('fill',c.fill,'important');
      shape.style.setProperty('stroke',c.stroke,'important');
    });
    nodeIdx++;
  });
  
  // Any remaining shapes not inside .node (like standalone rects)
  container.querySelectorAll('rect,circle,polygon,ellipse').forEach(el=>{
    if(el.closest('.node')||el.closest('.flowchart-node')||el.closest('marker'))return;
    if(el.closest('.cluster')){
      el.style.setProperty('fill',isLight?'#f0f0f0':'#161b22','important');
      el.style.setProperty('stroke',isLight?'#d97706':'#ffa500','important');
      return;
    }
    // Edge label backgrounds etc
    if(el.closest('.edgeLabel')){
      el.style.setProperty('fill',isLight?'#f6f8fa':'#0d1117','important');
      el.style.setProperty('stroke','none','important');
      return;
    }
  });
  const textColor=isLight?'#1f2328':'#e6edf3';
  container.querySelectorAll('text,tspan').forEach(el=>{
    el.style.setProperty('fill',textColor,'important');
  });
  container.querySelectorAll('foreignObject div, foreignObject span, .nodeLabel').forEach(el=>{
    el.style.setProperty('color',textColor,'important');
  });
  const clusterLabel=isLight?'#d97706':'#ffa500';
  container.querySelectorAll('.cluster-label .nodeLabel, .cluster-label span').forEach(el=>{
    el.style.setProperty('color',clusterLabel,'important');
  });
  const edgeColor=isLight?'#0969da':'#58a6ff';
  container.querySelectorAll('.edgePath path, .flowchart-link').forEach(el=>{
    el.style.setProperty('stroke',edgeColor,'important');
    el.style.setProperty('fill','none','important');
  });
  container.querySelectorAll('marker path').forEach(el=>{
    el.style.setProperty('fill',edgeColor,'important');
  });
}

function renderPreview(){
  const ext=(currentFile||'').split('.').pop().toLowerCase();
  // HTML files: render in iframe in preview pane
  if(ext==='html'||ext==='htm'){
    const blob=new Blob([cm.getValue()],{type:'text/html'});
    const url=URL.createObjectURL(blob);
    document.getElementById('preview').innerHTML='<iframe src="'+url+'" style="width:100%;height:100%;border:none;background:#fff"></iframe>';
    return;
  }
  // JSON files: formatted preview
  if(ext==='json'||ext==='jsonl'){
    try{
      const obj=JSON.parse(cm.getValue());
      document.getElementById('preview').innerHTML='<pre style="padding:20px;font-size:13px;line-height:1.6;white-space:pre-wrap;color:var(--text)">'+JSON.stringify(obj,null,2).replace(/</g,'&lt;')+'</pre>';
    }catch(e){
      document.getElementById('preview').innerHTML='<pre style="padding:20px;color:var(--red)">JSON 解析错误: '+e.message+'</pre>';
    }
    return;
  }
  // Non-markdown code files: syntax highlighted preview
  if(['js','ts','py','css','scss','sh','bash','yaml','yml','sql','toml','xml','svg','env','log','txt','csv'].includes(ext)){
    const escaped=cm.getValue().replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');
    const langMap={'js':'javascript','ts':'typescript','py':'python','sh':'bash','yml':'yaml','bash':'bash'};
    const lang=langMap[ext]||ext;
    let highlighted=escaped;
    if(window.hljs){try{highlighted=hljs.highlight(cm.getValue(),{language:lang}).value;}catch(e){}}
    document.getElementById('preview').innerHTML='<pre style="padding:20px;font-size:13px;line-height:1.6"><code class="hljs">'+highlighted+'</code></pre>';
    return;
  }
  let html=marked.parse(cm.getValue());
  // Proxy local image paths through /api/img endpoint
  html=html.replace(/<img\s([^>]*)src="([^"]+)"([^>]*)>/g,(m,pre,src,post)=>{
    if(src.startsWith('http')||src.startsWith('data:')||src.startsWith('/api/'))return m;
    // Encode path for URL
    const encoded=encodeURIComponent(src);
    return '<img '+pre+'src="'+api('api/img?p='+encoded)+'"'+post+'>';
  });
  document.getElementById('preview').innerHTML=html;
  document.querySelectorAll('#preview .mermaid-wrap').forEach(wrap=>{
    const src=wrap.querySelector('.mermaid-src');
    const render=wrap.querySelector('.mermaid-render');
    if(src&&render){
      try{
        const id='mm-'+(mmId++);
        const code=src.textContent;
        mermaid.render(id,code).then(r=>{
          render.innerHTML=r.svg;
          forceDarkMermaid(render);
        }).catch(e=>{render.innerHTML='<span style="color:var(--red)">Mermaid error: '+e.message+'</span>';});
      }catch(e){render.innerHTML='<span style="color:var(--red)">'+e.message+'</span>';}
    }
  });
}

// ==================== FILE TREE ====================

let files=[],treeData={};

if(FULL_MODE){
  fetch(api('api/files')).then(r=>r.json()).then(d=>{
    files=d.files;
    buildTree(files);
    renderTree(treeData);
    document.getElementById('fileCount').textContent=files.length+' 个文件';
    const urlFile=new URLSearchParams(location.search).get('file')||
      (location.hash.length>2?decodeURIComponent(location.hash.slice(1)):'');
    if(urlFile){
      // Expand parent dirs then open
      const parts=urlFile.split('/');
      for(let i=1;i<parts.length;i++){
        const dirPath=parts.slice(0,i).join('/');
        const dirEl=document.querySelector('.tree-dir[data-path="'+dirPath+'"]');
        if(dirEl&&!dirEl.classList.contains('open')){dirEl.classList.add('open');}
      }
      openFile(urlFile);
    }
    // Mobile: default to preview-only mode
    if(window.innerWidth<=768){
      const previewBtn=document.querySelector('.tabs button:last-child');
      if(previewBtn)setView('preview-only',previewBtn);
    }
    // Folders start collapsed — click to expand
  }).catch(e=>{
    document.getElementById('tree').innerHTML='<div style="padding:12px;color:var(--red)">加载失败: '+e.message+'</div>';
  });
}else{
  fetch(api('api/load')).then(r=>r.json()).then(d=>{
    cm.setValue(d.content);cm.clearHistory();cm.setOption('mode',getCMMode(ext));
    document.getElementById('fname').textContent=d.filename;
    renderPreview();
  });
}

function buildTree(list){
  treeData={__files:[]};
  list.forEach(f=>{
    const p=f.path.split('/');let n=treeData;
    for(let i=0;i<p.length-1;i++){if(!n[p[i]])n[p[i]]={__files:[]};n=n[p[i]];}
    if(!n.__files)n.__files=[];
    n.__files.push({name:p[p.length-1],path:f.path});
  });
}

function renderTree(node,container,pathPrefix){
  if(!container)container=document.getElementById('tree');
  if(!pathPrefix)pathPrefix='';
  container.innerHTML='';
  const dirs=Object.keys(node).filter(k=>k!=='__files').sort();
  const fls=(node.__files||[]).sort((a,b)=>a.name.localeCompare(b.name));
  
  dirs.forEach(dir=>{
    const dirPath=pathPrefix?pathPrefix+'/'+dir:dir;
    const c=countFiles(node[dir]);
    const d=document.createElement('div');d.className='tree-dir';
    d.setAttribute('data-path',dirPath);
    const dirLabels={'.xiao-a':'📋 业务资料','cognitive-archive':'💬 认知归档','golden-quotes':'✨ 金句库','clients':'👤 客户档案','conversations':'💬 对话','articles':'📝 文章','logs':'📋 日志','3-internal':'📂 公司内部','legal':'⚖️ 法律','finance':'💰 财务','hq':'🏢 总部','credentials':'🔑 凭证','assets':'🏷️ 资产','registration':'📋 注册','contracts':'📄 合同','ledger':'📒 账本','daily-sync':'📅 每日讨论','memory':'🧠 记忆'};
    const displayName=dirLabels[dir]||('📁 '+dir);
    d.innerHTML='<div class="tree-dir-label"><span class="arrow">▶</span> '+displayName+' <span style="color:#484f58;margin-left:auto;font-size:10px">'+c+'</span></div><div class="tree-children"></div>';
    const label=d.querySelector('.tree-dir-label');
    label.onclick=function(){
      this.querySelector('.arrow').classList.toggle('open');
      this.nextElementSibling.classList.toggle('open');
      activeDir=dirPath; // track for smart new-file
    };
    // Drop target
    if(!IS_READONLY&&AUTH_USER&&AUTH_USER.role!=='viewer')setupDropTarget(label,dirPath);
    container.appendChild(d);
    renderTree(node[dir],d.querySelector('.tree-children'),dirPath);
  });
  
  fls.forEach(f=>{
    const d=document.createElement('div');
    d.className='tree-file'+(currentFile===f.path?' active':'');
    const ext=(f.name.split('.').pop()||'').toLowerCase();
    const icons={
'pdf':'<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="#e5484d" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M15 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V7z"/><polyline points="14 2 14 8 20 8"/></svg>',
'docx':'<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="#3b82f6" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M15 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V7z"/><polyline points="14 2 14 8 20 8"/><line x1="16" y1="13" x2="8" y2="13"/><line x1="16" y1="17" x2="8" y2="17"/></svg>',
'doc':'<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="#3b82f6" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M15 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V7z"/><polyline points="14 2 14 8 20 8"/><line x1="16" y1="13" x2="8" y2="13"/><line x1="16" y1="17" x2="8" y2="17"/></svg>',
'xlsx':'<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="#22c55e" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><rect width="18" height="18" x="3" y="3" rx="2"/><path d="M3 9h18"/><path d="M3 15h18"/><path d="M9 3v18"/></svg>',
'xls':'<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="#22c55e" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><rect width="18" height="18" x="3" y="3" rx="2"/><path d="M3 9h18"/><path d="M3 15h18"/><path d="M9 3v18"/></svg>',
'csv':'<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="#22c55e" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><rect width="18" height="18" x="3" y="3" rx="2"/><path d="M3 9h18"/><path d="M3 15h18"/><path d="M9 3v18"/></svg>',
'jpg':'<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="#a855f7" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><rect width="18" height="18" x="3" y="3" rx="2" ry="2"/><circle cx="9" cy="9" r="2"/><path d="m21 15-3.086-3.086a2 2 0 0 0-2.828 0L6 21"/></svg>',
'jpeg':'<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="#a855f7" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><rect width="18" height="18" x="3" y="3" rx="2" ry="2"/><circle cx="9" cy="9" r="2"/><path d="m21 15-3.086-3.086a2 2 0 0 0-2.828 0L6 21"/></svg>',
'png':'<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="#a855f7" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><rect width="18" height="18" x="3" y="3" rx="2" ry="2"/><circle cx="9" cy="9" r="2"/><path d="m21 15-3.086-3.086a2 2 0 0 0-2.828 0L6 21"/></svg>',
'gif':'<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="#a855f7" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><rect width="18" height="18" x="3" y="3" rx="2" ry="2"/><circle cx="9" cy="9" r="2"/><path d="m21 15-3.086-3.086a2 2 0 0 0-2.828 0L6 21"/></svg>',
'webp':'<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="#a855f7" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><rect width="18" height="18" x="3" y="3" rx="2" ry="2"/><circle cx="9" cy="9" r="2"/><path d="m21 15-3.086-3.086a2 2 0 0 0-2.828 0L6 21"/></svg>',
'svg':'<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="#f59e0b" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><rect width="18" height="18" x="3" y="3" rx="2" ry="2"/><circle cx="9" cy="9" r="2"/><path d="m21 15-3.086-3.086a2 2 0 0 0-2.828 0L6 21"/></svg>',
'json':'<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="#f59e0b" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M15 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V7z"/><polyline points="14 2 14 8 20 8"/><path d="M10 12a1 1 0 0 0-1 1v1a1 1 0 0 1-1 1 1 1 0 0 1 1 1v1a1 1 0 0 0 1 1"/><path d="M14 18a1 1 0 0 0 1-1v-1a1 1 0 0 1 1-1 1 1 0 0 1-1-1v-1a1 1 0 0 0-1-1"/></svg>',
'js':'<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="#eab308" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M15 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V7z"/><polyline points="14 2 14 8 20 8"/><path d="m10 15-1 3"/><path d="m14 15 1 3"/></svg>',
'ts':'<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="#3b82f6" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M15 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V7z"/><polyline points="14 2 14 8 20 8"/><path d="m10 15-1 3"/><path d="m14 15 1 3"/></svg>',
'py':'<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="#3b82f6" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M15 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V7z"/><polyline points="14 2 14 8 20 8"/><path d="m10 13 2 2 4-4"/></svg>',
'html':'<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="#f97316" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="m18 16 4-4-4-4"/><path d="m6 8-4 4 4 4"/><path d="m14.5 4-5 16"/></svg>',
'htm':'<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="#f97316" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="m18 16 4-4-4-4"/><path d="m6 8-4 4 4 4"/><path d="m14.5 4-5 16"/></svg>',
'css':'<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="#06b6d4" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="m18 16 4-4-4-4"/><path d="m6 8-4 4 4 4"/><path d="m14.5 4-5 16"/></svg>',
'scss':'<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="#ec4899" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="m18 16 4-4-4-4"/><path d="m6 8-4 4 4 4"/><path d="m14.5 4-5 16"/></svg>',
'sh':'<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="#22c55e" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><polyline points="4 17 10 11 4 5"/><line x1="12" y1="19" x2="20" y2="19"/></svg>',
'bash':'<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="#22c55e" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><polyline points="4 17 10 11 4 5"/><line x1="12" y1="19" x2="20" y2="19"/></svg>',
'zsh':'<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="#22c55e" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><polyline points="4 17 10 11 4 5"/><line x1="12" y1="19" x2="20" y2="19"/></svg>',
'yaml':'<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="#ef4444" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M15 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V7z"/><polyline points="14 2 14 8 20 8"/></svg>',
'yml':'<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="#ef4444" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M15 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V7z"/><polyline points="14 2 14 8 20 8"/></svg>',
'toml':'<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="#8b5cf6" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M15 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V7z"/><polyline points="14 2 14 8 20 8"/></svg>',
'xml':'<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="#f97316" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="m18 16 4-4-4-4"/><path d="m6 8-4 4 4 4"/><path d="m14.5 4-5 16"/></svg>',
'txt':'<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M15 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V7z"/><polyline points="14 2 14 8 20 8"/><line x1="16" y1="13" x2="8" y2="13"/><line x1="16" y1="17" x2="8" y2="17"/></svg>',
'md':'<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="#58a6ff" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M15 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V7z"/><polyline points="14 2 14 8 20 8"/></svg>',
'env':'<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="#f59e0b" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><rect width="18" height="11" x="3" y="11" rx="2" ry="2"/><path d="M7 11V7a5 5 0 0 1 10 0v4"/></svg>',
'lock':'<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="#f59e0b" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><rect width="18" height="11" x="3" y="11" rx="2" ry="2"/><path d="M7 11V7a5 5 0 0 1 10 0v4"/></svg>',
'log':'<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="#6b7280" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><polyline points="4 17 10 11 4 5"/><line x1="12" y1="19" x2="20" y2="19"/></svg>',
'sql':'<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="#06b6d4" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><ellipse cx="12" cy="5" rx="9" ry="3"/><path d="M3 5v14a9 3 0 0 0 18 0V5"/><path d="M3 12a9 3 0 0 0 18 0"/></svg>',
'db':'<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="#06b6d4" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><ellipse cx="12" cy="5" rx="9" ry="3"/><path d="M3 5v14a9 3 0 0 0 18 0V5"/><path d="M3 12a9 3 0 0 0 18 0"/></svg>',
'zip':'<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="#8b5cf6" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M15 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V7z"/><polyline points="14 2 14 8 20 8"/><rect x="10" y="12" width="4" height="6" rx="1"/></svg>',
'gz':'<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="#8b5cf6" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M15 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V7z"/><polyline points="14 2 14 8 20 8"/><rect x="10" y="12" width="4" height="6" rx="1"/></svg>',
'tar':'<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="#8b5cf6" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M15 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V7z"/><polyline points="14 2 14 8 20 8"/><rect x="10" y="12" width="4" height="6" rx="1"/></svg>'
};
    const icon=icons[ext]||'';
    d.innerHTML=(icon||'<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M15 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V7z"/><polyline points="14 2 14 8 20 8"/></svg>')+'<span style="overflow:hidden;text-overflow:ellipsis">'+f.name+'</span>';d.title=f.path;
    d.setAttribute('data-path',f.path);
    d.onclick=()=>{openFile(f.path);closeMobileSidebar();};
    // Draggable (only for editors/admins)
    if(!IS_READONLY&&AUTH_USER&&AUTH_USER.role!=='viewer')setupDrag(d,f.path);
    container.appendChild(d);
  });
}

function countFiles(n){
  let c=(n.__files||[]).length;
  Object.keys(n).filter(k=>k!=='__files').forEach(k=>{c+=countFiles(n[k]);});
  return c;
}

// Expand tree to show a specific file path and scroll to it
function revealFileInTree(filePath){
  const parts=filePath.split('/');
  // Expand each directory level
  let pathSoFar='';
  for(let i=0;i<parts.length-1;i++){
    pathSoFar=pathSoFar?pathSoFar+'/'+parts[i]:parts[i];
    const dirEl=document.querySelector('.tree-dir[data-path="'+pathSoFar+'"]');
    if(dirEl){
      const arrow=dirEl.querySelector('.tree-dir-label .arrow');
      const children=dirEl.querySelector('.tree-children');
      if(arrow&&!arrow.classList.contains('open'))arrow.classList.add('open');
      if(children&&!children.classList.contains('open'))children.classList.add('open');
    }
  }
  // Highlight and scroll to file
  document.querySelectorAll('.tree-file').forEach(f=>{
    f.classList.toggle('active',f.getAttribute('data-path')===filePath);
  });
  const activeEl=document.querySelector('.tree-file.active');
  if(activeEl)activeEl.scrollIntoView({block:'center',behavior:'smooth'});
}

let foldersExpanded=false;
function toggleAllFolders(){
  foldersExpanded=!foldersExpanded;
  document.querySelectorAll('.tree-children').forEach(c=>{
    if(foldersExpanded)c.classList.add('open');else c.classList.remove('open');
  });
  document.querySelectorAll('.tree-dir-label .arrow').forEach(a=>{
    if(foldersExpanded)a.classList.add('open');else a.classList.remove('open');
  });
  document.getElementById('toggleFoldersBtn').innerHTML=foldersExpanded?'<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M20 20a2 2 0 0 0 2-2V8a2 2 0 0 0-2-2h-7.9a2 2 0 0 1-1.69-.9L9.6 3.9A2 2 0 0 0 7.93 3H4a2 2 0 0 0-2 2v13a2 2 0 0 0 2 2z"/></svg>':'<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="m6 14 1.5-2.9A2 2 0 0 1 9.24 10H20a2 2 0 0 1 1.94 2.5l-1.54 6a2 2 0 0 1-1.95 1.5H4a2 2 0 0 1-2-2V5a2 2 0 0 1 2-2h3.9a2 2 0 0 1 1.69.9l.81 1.2a2 2 0 0 0 1.67.9H18a2 2 0 0 1 2 2v2"/></svg>';
}

function filterFiles(){
  const q=document.getElementById('searchInput').value.toLowerCase();
  if(!q){buildTree(files);renderTree(treeData);if(currentFile)revealFileInTree(currentFile);return;}
  const filtered=files.filter(f=>f.path.toLowerCase().includes(q));
  buildTree(filtered);renderTree(treeData);
  document.querySelectorAll('.tree-children').forEach(c=>c.classList.add('open'));
  document.querySelectorAll('.arrow').forEach(a=>a.classList.add('open'));
}

function openFile(path){
  if(dirty&&!confirm('未保存，确定切换？'))return;
  history.replaceState(null,'','#'+encodeURIComponent(path));
  annotations=[];annIdCounter=0;refreshAnnotationDisplay();
  const ext=path.split('.').pop().toLowerCase();
  const imgTypes=['jpg','jpeg','png','gif','webp','svg'];
  if(imgTypes.includes(ext)){
    // Image: render directly, no API call needed
    _previewLocked=true;
    currentFile=path;
    cm.setValue('');cm.clearHistory();
    document.getElementById('fname').textContent=path;
    document.getElementById('saveBtn').style.display='none';
    dirty=false;
    revealFileInTree(path);addRecent(path);
    const previewBtn=document.querySelector('.tabs button:last-child');
    if(previewBtn)setView('preview-only',previewBtn);
    const imgUrl=api('api/raw')+'?path='+encodeURIComponent(path);
    document.getElementById('preview').innerHTML='<div style="text-align:center;padding:20px"><img src="'+imgUrl+'" style="max-width:100%;border-radius:6px;border:1px solid var(--border)" alt="'+path.split('/').pop()+'" onclick="window.open(this.src)"></div>';
    return;
  }
  const nonMdTypes=['pdf','docx','xlsx','xls'];
  if(nonMdTypes.includes(ext)){
    if(ext==='html'||ext==='htm'){
      _previewLocked=true;
      currentFile=path;
      cm.setValue('');cm.clearHistory();
      document.getElementById('fname').textContent=path;
      document.getElementById('saveBtn').style.display='none';
      // preview mode - no status text needed
      document.getElementById('status').style.color='var(--blue)';
      dirty=false;
      revealFileInTree(path);
      const previewBtn=document.querySelector('.tabs button:last-child');
      if(previewBtn)setView('preview-only',previewBtn);
      const iframeSrc=api('api/raw')+'?path='+encodeURIComponent(path);
      document.getElementById('preview').innerHTML='<iframe src="'+iframeSrc+'" style="width:100%;min-height:80vh;border:1px solid #30363d;border-radius:6px;background:#0D0D0F" onload="try{this.style.height=this.contentDocument.body.scrollHeight+40+\'px\'}catch(e){}"></iframe>';
      return;
    }
    _previewLocked=true;
    fetch(api('api/preview')+'?path='+encodeURIComponent(path)).then(r=>r.json()).then(d=>{
      currentFile=path;
      cm.setValue('');cm.clearHistory();
      document.getElementById('fname').textContent=path;
      document.getElementById('saveBtn').style.display='none';
      // preview mode - no status text needed
      document.getElementById('status').style.color='var(--blue)';
      dirty=false;
      revealFileInTree(path);
      // Force preview-only mode
      const previewBtn=document.querySelector('.tabs button:last-child');
      if(previewBtn)setView('preview-only',previewBtn);
      document.getElementById('preview').innerHTML=d.html;
    }).catch(e=>{
      document.getElementById('preview').innerHTML='<div style="color:#f85149;padding:20px">加载失败: '+e.message+'</div>';
    });
    return;
  }
  _previewLocked=false;
  fetch(api('api/load')+'?path='+encodeURIComponent(path)).then(r=>r.json()).then(d=>{
    currentFile=path;cm.operation(()=>{cm.setOption('mode',getCMMode(ext));cm.setValue(d.content);cm.clearHistory();});cm.focus();
    addRecent(path);updateFileStats();
    document.getElementById('fname').textContent=path;
    document.getElementById('saveBtn').style.display='';
    document.getElementById('saveBtn').style.color='';
    document.getElementById('status').style.color='var(--dim)';
    dirty=false;
    revealFileInTree(path);
    // Defer preview render for faster file switching
    clearTimeout(previewTimer);
    previewTimer=setTimeout(renderPreview,50);
  });
}

function save(){
  if(IS_READONLY){console.warn('Read-only mode, save blocked');return;}
  if(!currentFile)return;
  const btn=document.getElementById('saveBtn');btn.innerHTML='⏳';
  fetch(api('api/save'),{
    method:'POST',
    headers:{'Content-Type':'application/json'},
    body:JSON.stringify(FULL_MODE?{path:currentFile,content:cm.getValue()}:{content:cm.getValue()})
  }).then(r=>r.json()).then(d=>{
    if(d.ok){document.getElementById('saveBtn').style.color='var(--green)';dirty=false;setTimeout(()=>{if(!dirty)document.getElementById('saveBtn').style.color='';},2000);}
    btn.innerHTML='<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M15.2 3a2 2 0 0 1 1.4.6l3.8 3.8a2 2 0 0 1 .6 1.4V19a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2V5a2 2 0 0 1 2-2z"/><path d="M17 21v-7a1 1 0 0 0-1-1H8a1 1 0 0 0-1 1v7"/><path d="M7 3v4a1 1 0 0 0 1 1h7"/></svg>';
  });
}

function setView(mode,btn){
  document.getElementById('panels').className='panels '+mode;
  document.querySelectorAll('.tabs button').forEach(b=>b.classList.remove('active'));
  btn.classList.add('active');cm.refresh();if(mode!=='edit-only')renderPreview();
}

// ==================== NEW FILE (SMART) ====================
let activeDir=''; // track current folder context

function showNewFileDialog(presetDir){
  let defaultDir=presetDir||activeDir||(currentFile?currentFile.replace(/\/[^\/]+$/,''):'');
  if(defaultDir&&!defaultDir.endsWith('/'))defaultDir+='/';

  const d=openModal('newFileModal',`
    <h3 class="font-bold text-lg text-primary mb-4">新建文档</h3>
    <label class="text-xs text-base-content/50 mb-1 block">文件名（.md 后缀可省略）</label>
    <div class="flex items-center gap-2 mb-3">
      <span class="text-xs font-mono text-base-content/40 whitespace-nowrap" id="newFilePrefix">${defaultDir||'(根目录) '}</span>
      <input type="text" id="newFileName" placeholder="my-new-doc" class="input input-bordered input-sm flex-1 font-mono" />
    </div>
    <label class="text-xs text-base-content/40 mb-1 block">或输入完整路径</label>
    <input type="text" id="newFileFull" placeholder="Creations/subfolder/filename.md" value="${defaultDir}" class="input input-bordered input-sm w-full font-mono text-base-content/50 mb-2" />
    <div id="newFileErr" class="text-error text-xs min-h-[18px]"></div>
    <div class="modal-action">
      <button class="btn btn-sm btn-ghost" onclick="closeModal('newFileModal')">取消</button>
      <button class="btn btn-sm btn-primary" onclick="createNewFile()">创建</button>
    </div>
  `);

  const nameInp=document.getElementById('newFileName');
  const fullInp=document.getElementById('newFileFull');
  nameInp.focus();
  nameInp.addEventListener('input',()=>{fullInp.value=defaultDir+nameInp.value;fullInp.classList.remove('text-base-content/50');});
  fullInp.addEventListener('focus',()=>{fullInp.classList.remove('text-base-content/50');});
  const submit=e=>{if(e.key==='Enter')createNewFile();if(e.key==='Escape')closeModal('newFileModal');};
  nameInp.addEventListener('keydown',submit);
  fullInp.addEventListener('keydown',submit);
}

function createNewFile(){
  const nameInp=document.getElementById('newFileName');
  const fullInp=document.getElementById('newFileFull');
  const err=document.getElementById('newFileErr');
  
  // Prefer full path if name is empty, else build from prefix+name
  let path=nameInp.value.trim()?fullInp.value.trim():fullInp.value.trim();
  if(!path||path.endsWith('/')){err.textContent='请输入文件名';return;}
  if(!path.endsWith('.md'))path+='.md';
  
  fetch(api('api/new'),{
    method:'POST',
    headers:{'Content-Type':'application/json'},
    body:JSON.stringify({path:path})
  }).then(r=>r.json()).then(d=>{
    if(d.ok){
      closeModal('newFileModal');
      showToast('✅ 已创建: '+path);
      refreshTreeAndOpen(path);
    }else{
      err.textContent=d.error||'创建失败';
    }
  }).catch(e=>{err.textContent='请求失败: '+e.message;});
}

function refreshTreeAndOpen(path){
  fetch(api('api/files')).then(r=>r.json()).then(fd=>{
    files=fd.files;buildTree(files);renderTree(treeData);
    document.getElementById('fileCount').textContent=files.length+' 个文件';
    if(path)openFile(path);
  });
}

// ==================== DRAG & DROP ====================
let dragSrcPath=null;

function setupDrag(el,filePath){
  el.draggable=true;
  el.addEventListener('dragstart',e=>{
    dragSrcPath=filePath;
    e.dataTransfer.setData('text/plain',filePath);
    e.dataTransfer.effectAllowed='move';
    el.classList.add('dragging');
  });
  el.addEventListener('dragend',()=>{
    el.classList.remove('dragging');
    dragSrcPath=null;
    // Clean up all drag-over highlights
    document.querySelectorAll('.drag-over').forEach(d=>d.classList.remove('drag-over'));
  });
}

function setupDropTarget(el,dirPath){
  el.addEventListener('dragover',e=>{
    e.preventDefault();e.dataTransfer.dropEffect='move';
    el.classList.add('drag-over');
  });
  el.addEventListener('dragleave',()=>{el.classList.remove('drag-over');});
  el.addEventListener('drop',e=>{
    e.preventDefault();el.classList.remove('drag-over');
    const src=e.dataTransfer.getData('text/plain');
    if(!src)return;
    // Don't move to same directory
    const srcDir=src.replace(/\/[^\/]+$/,'')||'';
    if(srcDir===dirPath){showToast('⚠️ 已经在这个文件夹里','var(--accent)');return;}
    moveFileWithRefCheck(src,dirPath);
  });
}

async function moveFileWithRefCheck(srcPath,destDir){
  // First check references
  try{
    const refRes=await fetch(api('api/references')+'?path='+encodeURIComponent(srcPath));
    const refData=await refRes.json();
    if(refData.count>0){
      // Show warning
      const proceed=await showRefWarning(srcPath,destDir,refData.refs);
      if(!proceed)return;
    }
  }catch(e){
    console.warn('Reference check failed:',e);
  }
  
  // Do the move
  try{
    const res=await fetch(api('api/move'),{
      method:'POST',
      headers:{'Content-Type':'application/json'},
      body:JSON.stringify({from:srcPath,toDir:destDir})
    });
    const d=await res.json();
    if(d.ok){
      showToast('✅ 已移动到 '+(destDir||'根目录'));
      // If moved current file, update
      if(currentFile===srcPath)currentFile=d.newPath;
      refreshTreeAndOpen(currentFile);
    }else{
      showToast('❌ '+d.error,'var(--red)');
    }
  }catch(e){
    showToast('❌ 移动失败: '+e.message,'var(--red)');
  }
}

function showRefWarning(srcPath,destDir,refs){
  return new Promise(resolve=>{
    const overlay=document.createElement('div');overlay.className='ref-dialog';
    let refsHtml=refs.map(r=>{
      const lines=r.matches.map(m=>'<div class="ref-line">L'+m.line+': '+escHtml(m.text)+'</div>').join('');
      return '<div class="ref-file"><div class="ref-path">📄 '+r.file+'</div>'+lines+'</div>';
    }).join('');
    
    overlay.innerHTML=`
      <div class="ref-box">
        <h3 style="color:var(--accent);margin-bottom:8px">⚠️ 这个文件被 ${refs.length} 个文档引用</h3>
        <div style="font-size:13px;color:var(--dim);margin-bottom:12px">
          移动 <code style="color:var(--accent)">${srcPath}</code> → <code style="color:var(--green)">${destDir||'(根目录)'}/</code>
          <br>以下文档中的链接可能会失效：
        </div>
        <div style="max-height:300px;overflow-y:auto">${refsHtml}</div>
        <div style="display:flex;gap:8px;justify-content:flex-end;margin-top:16px">
          <button onclick="this.closest('.ref-dialog').remove()" style="background:var(--border);color:var(--text);border:none;padding:6px 16px;border-radius:6px;cursor:pointer;font-size:13px" id="refCancel">取消</button>
          <button style="background:var(--accent);color:#000;border:none;padding:6px 16px;border-radius:6px;cursor:pointer;font-size:13px;font-weight:600" id="refProceed">仍然移动</button>
        </div>
      </div>`;
    document.body.appendChild(overlay);
    overlay.querySelector('#refCancel').onclick=()=>{overlay.remove();resolve(false);};
    overlay.querySelector('#refProceed').onclick=()=>{overlay.remove();resolve(true);};
    overlay.addEventListener('click',e=>{if(e.target===overlay){overlay.remove();resolve(false);}});
  });
}

// ==================== CONTEXT MENU ====================
// ==================== CONTEXT MENU (Enhanced) ====================
let _selectedFiles=new Set();

document.addEventListener('contextmenu',e=>{
  const fileEl=e.target.closest('.tree-file');
  const dirLabel=e.target.closest('.tree-dir-label');
  if(!fileEl && !dirLabel)return;
  e.preventDefault();
  closeContextMenu();

  const menu=document.createElement('div');menu.className='ctx-menu';menu.id='ctxMenu';
  const isReadonly=typeof IS_READONLY!=='undefined'&&IS_READONLY;

  if(fileEl){
    const filePath=fileEl.getAttribute('data-path');
    const fileName=filePath.split('/').pop();
    const ext=fileName.split('.').pop().toLowerCase();
    const isMulti=_selectedFiles.size>1&&_selectedFiles.has(filePath);

    if(isMulti){
      // Multi-select context menu
      const count=_selectedFiles.size;
      menu.innerHTML=`
        <div class="ctx-menu-label">${count} files selected</div>
        <div class="ctx-menu-sep"></div>
        <div class="ctx-menu-item" onclick="batchShareSelected();closeContextMenu()"><svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><circle cx="18" cy="5" r="3"/><circle cx="6" cy="12" r="3"/><circle cx="18" cy="19" r="3"/><line x1="8.59" y1="13.51" x2="15.42" y2="17.49"/><line x1="15.41" y1="6.51" x2="8.59" y2="10.49"/></svg>Share selected</div>
        <div class="ctx-menu-item" onclick="copySelectedPaths();closeContextMenu()"><svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><rect width="14" height="14" x="8" y="8" rx="2" ry="2"/><path d="M4 16c-1.1 0-2-.9-2-2V4c0-1.1.9-2 2-2h10c1.1 0 2 .9 2 2"/></svg>Copy paths</div>
        <div class="ctx-menu-sep"></div>
        <div class="ctx-menu-item danger" onclick="batchDeleteSelected();closeContextMenu()"><svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M3 6h18"/><path d="M19 6v14c0 1-1 2-2 2H7c-1 0-2-1-2-2V6"/><path d="M8 6V4c0-1 1-2 2-2h4c1 0 2 1 2 2v2"/></svg>Delete selected</div>
        <div class="ctx-menu-sep"></div>
        <div class="ctx-menu-item" onclick="clearSelection();closeContextMenu()">Clear selection</div>
      `;
    } else {
      menu.innerHTML=`
        <div class="ctx-menu-label">${fileName}</div>
        <div class="ctx-menu-sep"></div>
        <div class="ctx-menu-item" onclick="openFile('${filePath}');closeContextMenu()"><svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M15 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V7Z"/><path d="M14 2v4a2 2 0 0 0 2 2h4"/></svg>Open</div>
        ${!isReadonly?'<div class="ctx-menu-item" onclick="renameFile(\''+filePath+'\');closeContextMenu()"><svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M21.174 6.812a1 1 0 0 0-3.986-3.987L3.842 16.174a2 2 0 0 0-.5.83l-1.321 4.352a.5.5 0 0 0 .623.622l4.353-1.32a2 2 0 0 0 .83-.497z"/></svg>Rename</div>':''}
        <div class="ctx-menu-sep"></div>
        <div class="ctx-menu-item" onclick="copyRelPath(\''+filePath+'\');closeContextMenu()"><svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><rect width="14" height="14" x="8" y="8" rx="2" ry="2"/><path d="M4 16c-1.1 0-2-.9-2-2V4c0-1.1.9-2 2-2h10c1.1 0 2 .9 2 2"/></svg>Copy path</div>
        <div class="ctx-menu-item" onclick="copyFileLink(\''+filePath+'\');closeContextMenu()"><svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M10 13a5 5 0 0 0 7.54.54l3-3a5 5 0 0 0-7.07-7.07l-1.72 1.71"/><path d="M14 11a5 5 0 0 0-7.54-.54l-3 3a5 5 0 0 0 7.07 7.07l1.71-1.71"/></svg>Copy link</div>
        <div class="ctx-menu-item" onclick="downloadFile(\''+filePath+'\');closeContextMenu()"><svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg>Download</div>
        ${!isReadonly?'<div class="ctx-menu-sep"></div><div class="ctx-menu-item" onclick="toggleSelectFile(\''+filePath+'\',null);closeContextMenu()"><svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><rect width="18" height="18" x="3" y="3" rx="2"/><path d="m9 12 2 2 4-4"/></svg>Select</div>':''}
        ${!isReadonly?'<div class="ctx-menu-sep"></div><div class="ctx-menu-item danger" onclick="deleteItem(\''+filePath+'\');closeContextMenu()"><svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M3 6h18"/><path d="M19 6v14c0 1-1 2-2 2H7c-1 0-2-1-2-2V6"/><path d="M8 6V4c0-1 1-2 2-2h4c1 0 2 1 2 2v2"/></svg>Delete</div>':''}
      `;
    }
  } else {
    const dirEl=dirLabel.closest('.tree-dir');
    const dirPath=dirEl?dirEl.getAttribute('data-path'):'';
    const dirName=dirPath.split('/').pop()||'Root';
    menu.innerHTML=`
      <div class="ctx-menu-label">${dirName}/</div>
      <div class="ctx-menu-sep"></div>
      ${!isReadonly?'<div class="ctx-menu-item" onclick="triggerUpload(\''+dirPath+'\');closeContextMenu()"><svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="17 8 12 3 7 8"/><line x1="12" y1="3" x2="12" y2="15"/></svg>Upload here</div>':''}
      ${!isReadonly?'<div class="ctx-menu-item" onclick="showNewFileDialog(\''+dirPath+'\');closeContextMenu()"><svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M15 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V7z"/><polyline points="14 2 14 8 20 8"/><line x1="12" y1="18" x2="12" y2="12"/><line x1="9" y1="15" x2="15" y2="15"/></svg>New file</div>':''}
      <div class="ctx-menu-item" onclick="copyRelPath(\''+dirPath+'\');closeContextMenu()"><svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><rect width="14" height="14" x="8" y="8" rx="2" ry="2"/><path d="M4 16c-1.1 0-2-.9-2-2V4c0-1.1.9-2 2-2h10c1.1 0 2 .9 2 2"/></svg>Copy path</div>
      <div class="ctx-menu-item" onclick="showFolderShareDialog(\''+dirPath+'\');closeContextMenu()"><svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M20 20a2 2 0 0 0 2-2V8a2 2 0 0 0-2-2h-7.9a2 2 0 0 1-1.69-.9L9.6 3.9A2 2 0 0 0 7.93 3H4a2 2 0 0 0-2 2v13a2 2 0 0 0 2 2z"/></svg>Share folder</div>
      <div class="ctx-menu-item" onclick="selectAllInDir(\''+dirPath+'\');closeContextMenu()"><svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><rect width="18" height="18" x="3" y="3" rx="2"/><path d="m9 12 2 2 4-4"/></svg>Select all files</div>
      ${!isReadonly?'<div class="ctx-menu-sep"></div><div class="ctx-menu-item danger" onclick="deleteItem(\''+dirPath+'\');closeContextMenu()"><svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M3 6h18"/><path d="M19 6v14c0 1-1 2-2 2H7c-1 0-2-1-2-2V6"/><path d="M8 6V4c0-1 1-2 2-2h4c1 0 2 1 2 2v2"/></svg>Delete folder</div>':''}
    `;
  }

  menu.style.left=e.clientX+'px';menu.style.top=e.clientY+'px';
  menu.addEventListener('mousedown',ev=>ev.stopPropagation());
  document.body.appendChild(menu);
  const rect=menu.getBoundingClientRect();
  if(rect.right>window.innerWidth)menu.style.left=(window.innerWidth-rect.width-8)+'px';
  if(rect.bottom>window.innerHeight)menu.style.top=(window.innerHeight-rect.height-8)+'px';
});

function closeContextMenu(){const m=document.getElementById('ctxMenu');if(m)m.remove();}
document.addEventListener('mousedown',e=>{if(!e.target.closest('.ctx-menu'))closeContextMenu();});
document.addEventListener('keydown',e=>{if(e.key==='Escape'){closeContextMenu();if(_selectedFiles.size)clearSelection();}});

// Clipboard helper (works on HTTP too)
function copyText(text,msg){
  msg=msg||'Copied';
  if(navigator.clipboard&&window.isSecureContext){
    navigator.clipboard.writeText(text).then(()=>showToast(msg)).catch(()=>_fallbackCopy(text,msg));
  }else{_fallbackCopy(text,msg);}
}
function _fallbackCopy(text,msg){
  const ta=document.createElement('textarea');
  ta.value=text;ta.style.cssText='position:fixed;left:-9999px;top:-9999px;opacity:0';
  document.body.appendChild(ta);ta.select();
  try{document.execCommand('copy');showToast(msg||'Copied');}
  catch(e){showToast('Copy failed','var(--red)');}
  finally{ta.remove();}
}

// Copy relative path
function copyRelPath(p){copyText(p,'Copied: '+p);}

// Copy shareable link
function copyFileLink(p){
  const url=location.origin+BASE+'#'+encodeURIComponent(p);
  copyText(url,'Link copied');
}

// Download file
function downloadFile(p){
  const a=document.createElement('a');
  a.href=api('api/raw')+'?path='+encodeURIComponent(p)+'&dl=1';
  a.download=p.split('/').pop();
  document.body.appendChild(a);a.click();a.remove();
}

// Rename file
function renameFile(oldPath){
  const oldName=oldPath.split('/').pop();
  const dir=oldPath.substring(0,oldPath.length-oldName.length);
  openModal('renameModal',`
    <h3 style="margin:0 0 12px;font-size:15px">Rename</h3>
    <input class="xh-input" id="renameInput" value="${oldName}" style="width:100%" autofocus>
    <div style="display:flex;gap:8px;margin-top:16px;justify-content:flex-end">
      <button class="xh-btn xh-btn-ghost" onclick="closeModal('renameModal')">Cancel</button>
      <button class="xh-btn xh-btn-primary" onclick="doRename('${oldPath}')">Rename</button>
    </div>
  `,{narrow:true});
  setTimeout(()=>{
    const inp=document.getElementById('renameInput');
    if(inp){inp.focus();const dot=oldName.lastIndexOf('.');if(dot>0)inp.setSelectionRange(0,dot);else inp.select();}
  },50);
}
async function doRename(oldPath){
  const newName=document.getElementById('renameInput').value.trim();
  if(!newName){showToast('Name cannot be empty','var(--red)');return;}
  const oldName=oldPath.split('/').pop();
  if(newName===oldName){closeModal('renameModal');return;}
  try{
    const r=await fetch(api('api/rename'),{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({path:oldPath,name:newName})});
    const d=await r.json();
    if(d.ok){
      showToast('Renamed to '+newName);
      closeModal('renameModal');
      if(currentFile===oldPath)currentFile=d.newPath;
      refreshTreeAndOpen(currentFile);
    }else{showToast(d.error||'Rename failed','var(--red)');}
  }catch(e){showToast('Rename failed','var(--red)');}
}

// Multi-select
function toggleSelectFile(path,el){
  if(_selectedFiles.has(path)){
    _selectedFiles.delete(path);
  }else{
    _selectedFiles.add(path);
  }
  updateSelectionUI();
}

function selectAllInDir(dirPath){
  files.filter(f=>f.path.startsWith(dirPath+'/')).forEach(f=>_selectedFiles.add(f.path));
  updateSelectionUI();
}

function clearSelection(){
  _selectedFiles.clear();
  updateSelectionUI();
}

function updateSelectionUI(){
  document.querySelectorAll('.tree-file').forEach(el=>{
    const p=el.getAttribute('data-path');
    el.classList.toggle('selected',_selectedFiles.has(p));
  });
  // Multi-select floating bar
  let bar=document.getElementById('multiSelectBar');
  if(_selectedFiles.size>0){
    if(!bar){
      bar=document.createElement('div');bar.className='multi-select-bar';bar.id='multiSelectBar';
      document.body.appendChild(bar);
    }
    bar.innerHTML='<span style="color:var(--blue);font-weight:600">'+_selectedFiles.size+' files</span>'
      +'<button onclick="batchShareSelected()" style="background:var(--blue);color:#000"><svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><circle cx="18" cy="5" r="3"/><circle cx="6" cy="12" r="3"/><circle cx="18" cy="19" r="3"/><line x1="8.59" y1="13.51" x2="15.42" y2="17.49"/><line x1="15.41" y1="6.51" x2="8.59" y2="10.49"/></svg> Share</button>'
      +'<button onclick="copySelectedPaths()" style="background:var(--border);color:var(--text)"><svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><rect width="14" height="14" x="8" y="8" rx="2" ry="2"/><path d="M4 16c-1.1 0-2-.9-2-2V4c0-1.1.9-2 2-2h10c1.1 0 2 .9 2 2"/></svg> Copy</button>'
      +'<button onclick="batchDeleteSelected()" style="background:var(--red);color:#fff"><svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M3 6h18"/><path d="M19 6v14c0 1-1 2-2 2H7c-1 0-2-1-2-2V6"/><path d="M8 6V4c0-1 1-2 2-2h4c1 0 2 1 2 2v2"/></svg> Delete</button>'
      +'<button onclick="clearSelection()" style="background:none;color:var(--dim);border:1px solid var(--border)">\u2715</button>';
  }else{
    if(bar)bar.remove();
  }
}


// Cmd/Ctrl+Click for multi-select in file tree (not for viewers)
if(!IS_READONLY&&AUTH_USER&&AUTH_USER.role!=='viewer'){
  document.querySelector('.tree')?.addEventListener('click',function(e){
    const fileEl=e.target.closest('.tree-file');
    if(!fileEl)return;
    if(e.metaKey||e.ctrlKey){
      e.preventDefault();e.stopPropagation();
      const path=fileEl.getAttribute('data-path');
      if(path)toggleSelectFile(path,fileEl);
    }
  },true);
}

function copySelectedPaths(){
  const paths=Array.from(_selectedFiles).join('\n');
  copyText(paths,'Copied '+_selectedFiles.size+' paths');
}

async function batchDeleteSelected(){
  if(!confirm('Delete '+_selectedFiles.size+' files?'))return;
  let ok=0,fail=0;
  for(const p of _selectedFiles){
    try{
      const r=await fetch(api('/api/delete'),{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({path:p})});
      const d=await r.json();
      if(d.ok)ok++;else fail++;
    }catch(e){fail++;}
  }
  showToast('Deleted '+ok+(fail?' ('+fail+' failed)':''));
  clearSelection();
  refreshTreeAndOpen(currentFile);
}

function batchShareSelected(){
  const paths=Array.from(_selectedFiles);
  if(!paths.length)return;
  // Group by directory
  const dirMap={};
  paths.forEach(p=>{
    const dir=p.substring(0,p.lastIndexOf('/'))||'.';
    if(!dirMap[dir])dirMap[dir]=[];
    dirMap[dir].push(p.split('/').pop());
  });
  const dirs=Object.keys(dirMap);
  if(dirs.length>1){
    showToast('Selected files span multiple folders, sharing first folder only','var(--accent)');
  }
  const dir=dirs[0];
  const fileNames=dirMap[dir];
  // Directly create share link
  const mode='readonly';
  openModal('batchShareModal',`
    <h3 style="margin:0 0 12px;font-size:15px">Share ${fileNames.length} files</h3>
    <div style="font-size:13px;color:var(--dim);margin-bottom:12px">${dir}/</div>
    <div style="max-height:150px;overflow-y:auto;background:var(--bg);border:1px solid var(--border);border-radius:6px;padding:8px;margin-bottom:16px;font-size:12px">${fileNames.map(f=>'<div style="padding:2px 0">'+f+'</div>').join('')}</div>
    <div style="display:flex;gap:8px;margin-bottom:16px">
      <button id="bsReadonly" class="xh-card-select sel" onclick="document.getElementById('bsReadonly').classList.add('sel');document.getElementById('bsEditable').classList.remove('sel');window._bsMode='readonly'" style="flex:1;padding:10px;text-align:center">Read-only</button>
      <button id="bsEditable" class="xh-card-select" onclick="document.getElementById('bsEditable').classList.add('sel');document.getElementById('bsReadonly').classList.remove('sel');window._bsMode='editable'" style="flex:1;padding:10px;text-align:center">Editable</button>
    </div>
    <div id="bsLinkArea" style="display:none;margin-bottom:12px">
      <div class="xh-link-box" style="display:flex;align-items:center;gap:8px;background:var(--bg);border:1px solid var(--border);border-radius:6px;padding:8px">
        <input type="text" id="bsLinkInput" readonly style="flex:1;background:none;border:none;color:var(--blue);font-family:monospace;font-size:12px;outline:none">
        <button class="xh-btn xh-btn-primary" onclick="copyText(document.getElementById('bsLinkInput').value,'Link copied')" style="padding:4px 10px;font-size:12px">Copy</button>
      </div>
    </div>
    <div style="display:flex;gap:8px;justify-content:flex-end">
      <button class="xh-btn xh-btn-ghost" onclick="closeModal('batchShareModal')">Close</button>
      <button class="xh-btn xh-btn-primary" id="bsCreateBtn" onclick="createBatchShare('${dir==='.'?'':dir}')">Create Link</button>
    </div>
  `);
  window._bsMode='readonly';
  window._bsFiles=fileNames;
}

async function createBatchShare(dir){
  try{
    const r=await fetch(api('api/share'),{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({file:dir||'.',mode:window._bsMode,type:'folder',files:window._bsFiles})});
    const d=await r.json();
    if(d.ok){
      const url=location.origin+BASE+'s/'+d.token;
      document.getElementById('bsLinkInput').value=url;
      document.getElementById('bsLinkArea').style.display='';
      document.getElementById('bsCreateBtn').textContent='Created!';
      document.getElementById('bsCreateBtn').disabled=true;
    }else{showToast(d.error||'Failed','var(--red)');}
  }catch(e){showToast('Failed: '+e.message,'var(--red)');}
}

// ==================== TRASH SYSTEM ====================
async function deleteItem(path){
  const name=path.split('/').pop();
  if(!confirm('确定删除 "'+name+'" 吗？\n\n文件会移到回收站，可以恢复。'))return;
  try{
    const r=await fetch(api('/api/delete'),{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({path})});
    const d=await r.json();
    if(d.ok){
      showToast('🗑️ 已移到回收站','var(--green)');
      // If deleted file is current, clear editor
      if(currentFile===path){currentFile=null;cm.setValue('');document.getElementById('fname').textContent='';document.getElementById('previewPane').innerHTML='';}
      loadFiles();
    }else{showToast('❌ '+d.error,'var(--red)');}
  }catch(e){showToast('❌ 删除失败','var(--red)');}
}

async function showTrashPanel(){
  try{
    const r=await fetch(api('/api/trash'));
    const d=await r.json();
    if(!d.ok)return;
    const items=d.items||[];
    let rows='';
    if(items.length===0){
      rows='<div class="text-center text-base-content/40 py-8">回收站是空的</div>';
    }else{
      items.slice().reverse().forEach(it=>{
        rows+=`<div class="flex items-center gap-3 px-4 py-3 border-b border-base-300">
          <div class="flex-1 min-w-0">
            <div class="text-sm truncate">${it.name}</div>
            <div class="text-xs text-base-content/40">${it.original}</div>
            <div class="text-xs text-base-content/30">${it.deleted_at}</div>
          </div>
          <button onclick="restoreItem('${it.id}',this)" class="btn btn-xs btn-success">恢复</button>
        </div>`;
      });
    }
    openModal('trashModal',`
      <div class="flex items-center gap-2 mb-4">
        <h3 class="font-bold text-lg flex-1">回收站</h3>
        ${items.length>0?'<button onclick="emptyTrash(this)" class="btn btn-xs btn-error">清空</button>':''}
      </div>
      <div class="overflow-y-auto max-h-[50vh] -mx-2" id="trashList">${rows}</div>
    `,{wide:false});
  }catch(e){showToast('加载回收站失败','var(--red)');}
}

async function restoreItem(id,btn){
  try{
    btn.disabled=true;btn.textContent='恢复中...';
    const r=await fetch(api('/api/restore'),{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({id})});
    const d=await r.json();
    if(d.ok){
      showToast('✅ 已恢复: '+d.restored,'var(--green)');
      btn.closest('.flex').remove();
      loadFiles();
      // Check if trash is now empty
      const remaining=document.querySelectorAll('#trashList > .flex');
      if(remaining.length===0){
        document.getElementById('trashList').innerHTML='<div class="text-center text-base-content/40 py-8">回收站是空的</div>';
      }
    }else{showToast('❌ '+d.error,'var(--red)');btn.disabled=false;btn.textContent='恢复';}
  }catch(e){showToast('❌ 恢复失败','var(--red)');btn.disabled=false;btn.textContent='恢复';}
}

async function emptyTrash(btn){
  if(!confirm('确定永久清空回收站吗？此操作不可恢复！'))return;
  try{
    btn.disabled=true;btn.textContent='清空中...';
    const r=await fetch(api('/api/trash-empty'),{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({})});
    const d=await r.json();
    if(d.ok){
      showToast('🗑️ 回收站已清空','var(--green)');
      document.getElementById('trashList').innerHTML='<div class="text-center text-base-content/40 py-8">回收站是空的</div>';
      btn.remove();
    }
  }catch(e){showToast('❌ 清空失败','var(--red)');btn.disabled=false;btn.textContent='清空';}
}

// ==================== SHARE SYSTEM ====================

function showShareDialog(){
  if(!currentFile){showToast('先打开一个文件','var(--accent)');return;}
  openModal('shareModal',`
    <h3>分享 · ${currentFile.split('/').pop()}</h3>
    <div class="xh-card-select" style="margin-bottom:16px">
      <button id="shareReadonly" onclick="selectShareMode('readonly')">
        <svg width="15" height="15" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M2.062 12.348a1 1 0 0 1 0-.696 10.75 10.75 0 0 1 19.876 0 1 1 0 0 1 0 .696 10.75 10.75 0 0 1-19.876 0"/><circle cx="12" cy="12" r="3"/></svg>
        只读
      </button>
      <button id="shareEditable" onclick="selectShareMode('editable')">
        <svg width="15" height="15" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M21.174 6.812a1 1 0 0 0-3.986-3.987L3.842 16.174a2 2 0 0 0-.5.83l-1.321 4.352a.5.5 0 0 0 .623.622l4.353-1.32a2 2 0 0 0 .83-.497z"/></svg>
        可编辑
      </button>
    </div>
    <div id="shareLinkArea" style="display:none;margin-bottom:16px">
      <div class="xh-link-box">
        <input type="text" id="shareLinkInput" readonly>
        <button onclick="copyShareLink()">复制</button>
      </div>
    </div>
    <div id="shareExisting"></div>
    <div class="xh-actions">
      <button class="xh-btn xh-btn-ghost" onclick="closeModal('shareModal')">关闭</button>
    </div>
  `);
  loadExistingShares();
}

function selectShareMode(mode){
  document.getElementById('shareReadonly').classList.toggle('sel',mode==='readonly');
  document.getElementById('shareEditable').classList.toggle('sel',mode==='editable');
  // Create share
  fetch(api('api/share'),{
    method:'POST',
    headers:{'Content-Type':'application/json'},
    body:JSON.stringify({file:currentFile,mode:mode})
  }).then(r=>r.json()).then(d=>{
    if(d.ok){
      const url=location.origin+BASE+'s/'+d.token;
      document.getElementById('shareLinkInput').value=url;
      document.getElementById('shareLinkArea').style.display='';
      loadExistingShares();
    }
  });
}

function copyShareLink(){
  const inp=document.getElementById('shareLinkInput');
  inp.select();
  copyText(inp.value,'Link copied');(()=>{
    showToast('✅ 链接已复制');
  }).catch(()=>{
    document.execCommand('copy');
    showToast('✅ 链接已复制');
  });
}

function loadExistingShares(){
  fetch(api('api/shares')+'?file='+encodeURIComponent(currentFile)).then(r=>r.json()).then(d=>{
    const container=document.getElementById('shareExisting');
    if(!container)return;
    if(!d.shares||!d.shares.length){container.innerHTML='';return;}
    let html='<div class="space-y-2"><div class="text-xs text-base-content/40 mb-2">已有分享链接</div>';
    d.shares.forEach(s=>{
      const url=location.origin+BASE+'s/'+s.token;
      const modeTag=s.mode==='readonly'?'<span class="share-mode-tag readonly">只读</span>':'<span class="share-mode-tag editable">可编辑</span>';
      html+='<div class="share-item">'+modeTag+
        '<span style="flex:1;font-family:monospace;color:var(--blue);cursor:pointer" onclick="navigator.clipboard.writeText(\''+url+'\');showToast(\'✅ 已复制\')">'+s.token+'</span>'+
        '<span style="color:var(--dim)">'+s.created+'</span>'+
        '<span class="share-del" onclick="deleteShare(\''+s.token+'\')" title="删除">✕</span></div>';
    });
    html+='</div>';
    container.innerHTML=html;
  });
}

function deleteShare(token){
  fetch(api('api/share'),{
    method:'DELETE',
    headers:{'Content-Type':'application/json'},
    body:JSON.stringify({token:token})
  }).then(r=>r.json()).then(d=>{
    if(d.ok){showToast('✅ 分享已撤销');loadExistingShares();}
  });
}

// ==================== FOLDER SHARE ====================

function showFolderShareDialog(dirPath){
  closeModal('folderShareModal');
  // Get files in this directory from the tree
  const dirFiles=files.filter(f=>f.path.startsWith(dirPath+'/')&&!f.path.slice(dirPath.length+1).includes('/'));
  if(!dirFiles.length){showToast('⚠️ 此文件夹没有 .md 文件','var(--accent)');return;}
  

  let fileListHtml=dirFiles.map((f,i)=>{
    const name=f.path.split('/').pop();
    return '<label style="display:flex;align-items:center;gap:8px;padding:6px 8px;cursor:pointer;border-radius:4px;font-size:13px" onmouseover="this.style.background=\'rgba(255,255,255,.04)\'" onmouseout="this.style.background=\'none\'"><input type="checkbox" class="folder-share-cb" value="'+name+'" checked style="accent-color:var(--blue)"> '+name+'</label>';
  }).join('');
  
  const fsHtml=`
      <h3 style="margin:0 0 12px;font-size:15px;display:flex;align-items:center;gap:8px"><svg width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M20 20a2 2 0 0 0 2-2V8a2 2 0 0 0-2-2h-7.9a2 2 0 0 1-1.69-.9L9.6 3.9A2 2 0 0 0 7.93 3H4a2 2 0 0 0-2 2v13a2 2 0 0 0 2 2z"/></svg> Share "${dirPath.split('/').pop()}"</h3>
      <div style="margin-bottom:12px">
        <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:8px">
          <span style="font-size:13px;color:var(--dim)">${dirFiles.length} 个文件，勾选要分享的：</span>
          <span>
            <button onclick="document.querySelectorAll('.folder-share-cb').forEach(c=>c.checked=true)" style="background:none;border:1px solid var(--border);color:var(--dim);padding:2px 8px;border-radius:4px;cursor:pointer;font-size:11px">全选</button>
            <button onclick="document.querySelectorAll('.folder-share-cb').forEach(c=>c.checked=false)" style="background:none;border:1px solid var(--border);color:var(--dim);padding:2px 8px;border-radius:4px;cursor:pointer;font-size:11px">全不选</button>
          </span>
        </div>
        <div style="background:var(--bg);border:1px solid var(--border);border-radius:6px;max-height:300px;overflow-y:auto;padding:4px">
          ${fileListHtml}
        </div>
      </div>
      <div class="share-mode-btn">
        <button id="folderShareReadonly" onclick="selectFolderShareMode('readonly')" class="selected">
          <div class="mode-icon"><svg width="24" height="24" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M2.062 12.348a1 1 0 0 1 0-.696 10.75 10.75 0 0 1 19.876 0 1 1 0 0 1 0 .696 10.75 10.75 0 0 1-19.876 0"/><circle cx="12" cy="12" r="3"/></svg></div>
          <div class="mode-title">只读分享</div>
          <div class="mode-desc">对方只能查看</div>
        </button>
        <button id="folderShareEditable" onclick="selectFolderShareMode('editable')">
          <div class="mode-icon"><svg width="24" height="24" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M21.174 6.812a1 1 0 0 0-3.986-3.987L3.842 16.174a2 2 0 0 0-.5.83l-1.321 4.352a.5.5 0 0 0 .623.622l4.353-1.32a2 2 0 0 0 .83-.497z"/></svg></div>
          <div class="mode-title">可编辑分享</div>
          <div class="mode-desc">对方可以修改</div>
        </button>
      </div>
      <div id="folderShareLinkArea" style="display:none">
        <div class="share-link-box">
          <input type="text" id="folderShareLinkInput" readonly>
          <button onclick="copyFolderShareLink()"><svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><rect width="14" height="14" x="8" y="8" rx="2" ry="2"/><path d="M4 16c-1.1 0-2-.9-2-2V4c0-1.1.9-2 2-2h10c1.1 0 2 .9 2 2"/></svg> 复制</button>
        </div>
      </div>
      <div style="display:flex;gap:8px;justify-content:flex-end;margin-top:12px">
        <button class="xh-btn xh-btn-ghost" onclick="closeModal('folderShareModal')">Close</button>
        <button class="xh-btn xh-btn-primary" onclick="createFolderShare('${dirPath}')" id="folderShareCreateBtn">Create Link</button>
      </div>
  `;
  openModal('folderShareModal',fsHtml,{wide:true});
}

let _folderShareMode='readonly';
function selectFolderShareMode(mode){
  _folderShareMode=mode;
  document.getElementById('folderShareReadonly').classList.toggle('selected',mode==='readonly');
  document.getElementById('folderShareEditable').classList.toggle('selected',mode==='editable');
}

function createFolderShare(dirPath){
  const checked=Array.from(document.querySelectorAll('.folder-share-cb:checked')).map(c=>c.value);
  if(!checked.length){showToast('⚠️ 请至少选择一个文件','var(--accent)');return;}
  
  fetch(api('api/share'),{
    method:'POST',
    headers:{'Content-Type':'application/json'},
    body:JSON.stringify({file:dirPath,mode:_folderShareMode,type:'folder',files:checked})
  }).then(r=>r.json()).then(d=>{
    if(d.ok){
      const url=location.origin+BASE+'s/'+d.token;
      document.getElementById('folderShareLinkInput').value=url;
      document.getElementById('folderShareLinkArea').style.display='';
      document.getElementById('folderShareCreateBtn').textContent='✅ 已生成 ('+d.fileCount+' 个文件)';
      document.getElementById('folderShareCreateBtn').disabled=true;
    }
  });
}

function copyFolderShareLink(){
  const inp=document.getElementById('folderShareLinkInput');
  inp.select();
  copyText(inp.value,'Link copied');(()=>{showToast('✅ 链接已复制');}).catch(()=>{document.execCommand('copy');showToast('✅ 链接已复制');});
}

// ==================== EXPORT ====================
let _expFmt='pdf',_expTheme='light',_expSig='none',_expLayout='a4',_expWordStyle='modern',_expWordMargin='normal',_expWordPage='a4';
let _customSigName='',_customSigAvatar=null,_customSigSelectedId=null;
let _cachedExportConfig=null,_cachedCustomSigs=[];

function _buildToggleButtons(items, type){
  let html='';
  for(const[key,val] of Object.entries(items)){
    const sel=val.default?'sel':'';
    html+=`<button class="${sel}" onclick="selExp('${type}','${key}',this)">${val.label||key}</button>`;
  }
  return html;
}
function _findDefault(items){
  for(const[key,val] of Object.entries(items)){if(val.default)return key;}
  return Object.keys(items)[0]||'';
}

async function showExportDialog(){
  const content=cm.getValue();
  if(!content||!content.trim()){showToast('⚠️ 没有内容可导出','var(--accent)');return;}
  // Fetch config
  try{const r=await fetch(api('api/config'));_cachedExportConfig=await r.json();}
  catch(e){_cachedExportConfig={signatures:{none:null},pdfThemes:{light:{label:'☀️ 浅色',default:true}},pdfLayouts:{a4:{label:'A4',default:true}},wordStyles:{modern:{label:'现代',default:true}},wordMargins:{normal:{label:'标准',default:true}},wordPages:{a4:{label:'A4',default:true}}};}
  const cfg=_cachedExportConfig;

  // Fetch saved custom signatures
  _cachedCustomSigs=[];
  if(AUTH_USER){
    try{const sr=await fetch(api('api/custom-signatures'));const sd=await sr.json();
      if(sd.ok)_cachedCustomSigs=sd.signatures||[];
    }catch(e){}
  }

  // Build signature buttons: presets + saved custom sigs + "+" add new
  let sigHtml='';
  for(const[key,val] of Object.entries(cfg.signatures||{})){
    const sel=key==='none'?'sel':'';
    const label=val===null?'无署名':(val.name||key);
    sigHtml+=`<button class="${sel}" onclick="selExp('sig','${key}',this)">${label}</button>`;
  }
  for(const cs of _cachedCustomSigs){
    const av=cs.avatarUrl?`<img src="${api(cs.avatarUrl)}" style="width:16px;height:16px;border-radius:50%;object-fit:cover;vertical-align:middle;margin-right:3px">`:'';
    sigHtml+=`<button onclick="selectCustomSig('${cs.id}',this)" data-csid="${cs.id}">${av}${cs.name}</button>`;
  }
  if(AUTH_USER) sigHtml+=`<button onclick="selExp('sig','new',this)" style="opacity:.7">+ 新建</button>`;

  const _expModal=openModal('exportModal',`
      <h3><svg width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg> 导出文档</h3>
      <div class="export-row"><label>格式</label>
        <div class="export-toggle" id="expFmt">
          <button class="sel" onclick="selExp('fmt','pdf',this)"><svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M15 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V7z"/><polyline points="14 2 14 8 20 8"/><line x1="16" y1="13" x2="8" y2="13"/><line x1="16" y1="17" x2="8" y2="17"/><line x1="10" y1="9" x2="8" y2="9"/></svg> PDF</button>
          <button onclick="selExp('fmt','word',this)"><svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M15 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V7z"/><polyline points="14 2 14 8 20 8"/><path d="M10 13h4"/><path d="M12 17V13"/></svg> Word</button>
        </div>
      </div>
      <div class="pdf-opts" id="pdfOpts">
        <div class="export-row"><label>页面</label>
          <div class="export-toggle" id="expLayout">${_buildToggleButtons(cfg.pdfLayouts||{},'layout')}</div>
        </div>
        <div class="export-row"><label>主题</label>
          <div class="export-toggle" id="expTheme">${_buildToggleButtons(cfg.pdfThemes||{},'theme')}</div>
        </div>
      </div>
      <div class="word-opts" id="wordOpts" style="display:none">
        <div class="export-row"><label>页面</label>
          <div class="export-toggle" id="expWordPage">${_buildToggleButtons(cfg.wordPages||{},'wordPage')}</div>
        </div>
        <div class="export-row"><label>风格</label>
          <div class="export-toggle" id="expWordStyle">${_buildToggleButtons(cfg.wordStyles||{},'wordStyle')}</div>
        </div>
        <div class="export-row"><label>页边距</label>
          <div class="export-toggle" id="expWordMargin">${_buildToggleButtons(cfg.wordMargins||{},'wordMargin')}</div>
        </div>
      </div>
      <div class="export-row"><label>署名</label>
        <div class="export-toggle" id="expSig" style="flex-wrap:wrap">${sigHtml}</div>
      </div>
      <div id="customSigArea" style="display:none;margin:8px 0 4px;padding:10px 14px;border-radius:8px;background:var(--sidebar-bg);border:1px solid var(--border)">
        <div style="display:flex;align-items:center;gap:10px">
          <label style="width:42px;height:42px;border-radius:50%;background:var(--border);display:flex;align-items:center;justify-content:center;cursor:pointer;overflow:hidden;flex-shrink:0;border:2px dashed var(--text-muted,#888);position:relative" title="上传头像">
            <span id="customAvatarPlaceholder" style="opacity:.5"><svg width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M14.5 4h-5L7 7H4a2 2 0 0 0-2 2v9a2 2 0 0 0 2 2h16a2 2 0 0 0 2-2V9a2 2 0 0 0-2-2h-3l-2.5-3z"/><circle cx="12" cy="13" r="3"/></svg></span>
            <img id="customAvatarPreview" style="display:none;width:100%;height:100%;object-fit:cover;position:absolute;inset:0" alt="">
            <input type="file" accept="image/*" style="display:none" onchange="handleCustomAvatar(this)">
          </label>
          <input type="text" id="customSigNameInput" placeholder="输入署名名称" style="flex:1;padding:6px 10px;border-radius:6px;border:1px solid var(--border);background:var(--bg);color:var(--text);font-size:14px;outline:none" oninput="_customSigName=this.value">
          <button onclick="saveNewCustomSig()" style="padding:4px 12px;border-radius:6px;background:var(--accent);color:#fff;border:none;cursor:pointer;font-size:13px;white-space:nowrap">保存</button>
        </div>
      </div>
      <div class="export-actions">
        <button onclick="closeModal('exportModal')" class="xh-btn xh-btn-ghost">取消</button>
        ${AUTH_USER&&_cachedCustomSigs.length?'<button onclick="showSigManager()" style="background:transparent;color:var(--text-muted,#888);font-size:12px;border:1px solid var(--border)">管理署名</button>':''}
        <button id="expDoBtn" onclick="doExport()" class="xh-btn xh-btn-primary">导出</button>
      </div>
  `,{wide:false});
  _expFmt='pdf';_expTheme=_findDefault(cfg.pdfThemes||{});_expLayout=_findDefault(cfg.pdfLayouts||{});
  _expSig='none';_customSigName='';_customSigAvatar=null;_customSigSelectedId=null;
  _expWordStyle=_findDefault(cfg.wordStyles||{});_expWordMargin=_findDefault(cfg.wordMargins||{});_expWordPage=_findDefault(cfg.wordPages||{});
}

function selectCustomSig(id,btn){
  document.getElementById('expSig').querySelectorAll('button').forEach(b=>b.classList.remove('sel'));
  btn.classList.add('sel');
  _expSig='custom';_customSigSelectedId=id;
  const cs=_cachedCustomSigs.find(s=>s.id===id);
  if(cs){_customSigName=cs.name;}
  document.getElementById('customSigArea').style.display='none';
}

function handleCustomAvatar(input){
  const file=input.files&&input.files[0];
  if(!file)return;
  if(file.size>2*1024*1024){showToast('头像不能超过 2MB','var(--red)');return;}
  const reader=new FileReader();
  reader.onload=function(e){
    _customSigAvatar=e.target.result;
    const img=document.getElementById('customAvatarPreview');
    img.src=_customSigAvatar;img.style.display='block';
    document.getElementById('customAvatarPlaceholder').style.display='none';
  };
  reader.readAsDataURL(file);
}

async function saveNewCustomSig(){
  const name=_customSigName.trim();
  if(!name){showToast('请输入署名名称','var(--red)');return;}
  try{
    const saveBody={action:'create',name:name,avatar:_customSigAvatar||null};
    const r=await fetch(api('api/custom-signatures'),{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(saveBody)});
    const d=await r.json();
    if(d.ok){showToast('✅ 署名已保存');
      // Refresh export dialog to show new sig
      document.querySelectorAll('.share-dialog').forEach(d=>d.remove());
      showExportDialog();
    }else{showToast('❌ '+d.error,'var(--red)');}
  }catch(e){showToast('❌ 保存失败','var(--red)');}
}

async function showSigManager(){
  let listHtml='';
  for(const cs of _cachedCustomSigs){
    const av=cs.avatarUrl?`<img src="${api(cs.avatarUrl)}" style="width:32px;height:32px;border-radius:50%;object-fit:cover">`
      :'<div style="width:32px;height:32px;border-radius:50%;background:var(--border);display:flex;align-items:center;justify-content:center;font-size:14px;opacity:.5">👤</div>';
    listHtml+=`<div style="display:flex;align-items:center;gap:12px;padding:10px 0;border-bottom:1px solid var(--border)" data-sigid="${cs.id}">
      ${av}
      <span style="flex:1;font-size:14px">${cs.name}</span>
      <button onclick="deleteCustomSig('${cs.id}')" style="padding:3px 10px;border-radius:4px;background:var(--red,#e74c3c);color:#fff;border:none;cursor:pointer;font-size:12px">删除</button>
    </div>`;
  }
  if(!listHtml)listHtml='<p style="text-align:center;opacity:.5;padding:20px 0">暂无自定义署名</p>';
  const mgr=document.createElement('div');
  mgr.className='sig-manager-overlay';
  mgr.style.cssText='position:fixed;inset:0;background:rgba(0,0,0,.5);display:flex;align-items:center;justify-content:center;z-index:10001';
  mgr.innerHTML=`<div style="background:var(--bg);border:1px solid var(--border);border-radius:12px;padding:20px;width:360px;max-width:90vw;max-height:70vh;overflow-y:auto">
    <h3 style="margin:0 0 14px;font-size:16px">管理自定义署名</h3>
    ${listHtml}
    <div style="margin-top:14px;text-align:right">
      <button onclick="this.closest('.sig-manager-overlay').remove()" style="padding:6px 16px;border-radius:6px;background:var(--border);color:var(--text);border:none;cursor:pointer">关闭</button>
    </div>
  </div>`;
  document.body.appendChild(mgr);
  mgr.addEventListener('click',e=>{if(e.target===mgr)mgr.remove();});
}

async function deleteCustomSig(id){
  if(!confirm('确定删除此署名？'))return;
  try{
    const r=await fetch(api('api/custom-signatures'),{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({action:'delete',id:id})});
    const d=await r.json();
    if(d.ok){
      _cachedCustomSigs=_cachedCustomSigs.filter(s=>s.id!==id);
      // Remove from manager UI
      document.querySelectorAll(`[data-sigid="${id}"]`).forEach(el=>el.remove());
      // Also remove from export dialog sig buttons
      document.querySelectorAll(`#expSig [data-csid="${id}"]`).forEach(el=>el.remove());
      if(_customSigSelectedId===id){_expSig='none';_customSigSelectedId=null;}
      showToast('✅ 已删除');
    }
  }catch(e){showToast('❌ 删除失败','var(--red)');}
}

function selExp(type,val,btn){
  btn.parentElement.querySelectorAll('button').forEach(b=>b.classList.remove('sel'));
  btn.classList.add('sel');
  if(type==='fmt'){
    _expFmt=val;
    document.getElementById('pdfOpts').style.display=val==='pdf'?'':'none';
    document.getElementById('wordOpts').style.display=val==='word'?'':'none';
  }
  else if(type==='theme')_expTheme=val;
  else if(type==='layout')_expLayout=val;
  else if(type==='wordStyle')_expWordStyle=val;
  else if(type==='wordMargin')_expWordMargin=val;
  else if(type==='wordPage')_expWordPage=val;
  else{
    _expSig=val;_customSigSelectedId=null;
    const area=document.getElementById('customSigArea');
    if(area)area.style.display=val==='new'?'':'none';
    if(val==='new'){_customSigName='';_customSigAvatar=null;
      const inp=document.getElementById('customSigNameInput');if(inp)inp.value='';
      const prev=document.getElementById('customAvatarPreview');if(prev){prev.style.display='none';}
      const ph=document.getElementById('customAvatarPlaceholder');if(ph)ph.style.display='';
    }
  }
}

async function doExport(){
  const btn=document.getElementById('expDoBtn');
  btn.innerHTML='<span class="export-spinner"></span>导出中...';btn.disabled=true;
  try{
    const body={content:cm.getValue(),format:_expFmt,theme:_expTheme,signature:_expSig,
      layout:_expLayout,wordStyle:_expWordStyle,wordMargin:_expWordMargin,wordPage:_expWordPage,
      title:document.getElementById('fname').textContent||'document'};
    if(_expSig==='custom'){
      // Using a saved custom signature
      const cs=_cachedCustomSigs.find(s=>s.id===_customSigSelectedId);
      if(!cs){showToast('请选择一个署名','var(--red)');btn.innerHTML='导出';btn.disabled=false;return;}
      body.customName=cs.name;
      if(cs.avatarUrl){
        try{const ar=await fetch(api(cs.avatarUrl));const ab=await ar.blob();
          body.customAvatar=await new Promise(r=>{const rd=new FileReader();rd.onload=()=>r(rd.result);rd.readAsDataURL(ab);});
        }catch(e){body.customAvatar=null;}
      }else{body.customAvatar=null;}
    }else if(_expSig==='new'){
      // Inline new custom sig (not yet saved)
      body.signature='custom';
      body.customName=_customSigName||'';
      body.customAvatar=_customSigAvatar||null;
      if(!body.customName.trim()){showToast('请输入署名名称','var(--red)');btn.innerHTML='导出';btn.disabled=false;return;}
    }
    const resp=await fetch(api('api/export'),{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(body)});
    const data=await resp.json();
    if(data.ok){
      const a=document.createElement('a');a.href=api('api/download/'+data.downloadId);a.download=data.filename;
      document.body.appendChild(a);a.click();a.remove();
      showToast('✅ 导出成功');
      document.querySelectorAll('.share-dialog').forEach(d=>d.remove());
    }else{showToast('❌ '+data.error,'var(--red)');}
  }catch(e){showToast('❌ 导出失败: '+e.message,'var(--red)');}
  btn.innerHTML='导出';btn.disabled=false;
}

// ==================== SHARE MODE (read-only) ====================

const DEFAULT_VIEW='__DEFAULT_VIEW__';
if(IS_READONLY){
  cm.setOption('readOnly',true);
  document.querySelector('.CodeMirror').style.display='none';
  document.querySelector('.editor-pane').style.display='none';
  // Force preview-only, hide view toggle buttons that could reveal editor
  document.querySelectorAll('.toolbar button').forEach(b=>{
    if(b.textContent.includes('编辑')||b.textContent.includes('分栏'))b.style.display='none';
  });
}
// Apply default view on load
if(DEFAULT_VIEW!=='split'){
  setTimeout(()=>{
    const btns=document.querySelectorAll('.tabs button');
    const viewMap={'split':0,'edit-only':1,'preview-only':2};
    const idx=viewMap[DEFAULT_VIEW]||0;
    if(btns[idx])setView(DEFAULT_VIEW,btns[idx]);
  },100);
}

window.addEventListener('beforeunload',e=>{if(dirty){e.preventDefault();e.returnValue='';}});

// ── AI Assistant ──
if(!IS_READONLY){
  const aiHTML=`
  <button class="ai-fab" onclick="toggleAI()" title="AI 助手"><svg width="24" height="24" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M12 8V4H8"/><rect width="16" height="12" x="4" y="8" rx="2"/><path d="M2 14h2"/><path d="M20 14h2"/><path d="M15 13v2"/><path d="M9 13v2"/></svg></button>
  <div class="ai-panel" id="aiPanel">
    <div class="ai-panel-header">
      <h3><svg width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M12 8V4H8"/><rect width="16" height="12" x="4" y="8" rx="2"/><path d="M2 14h2"/><path d="M20 14h2"/><path d="M15 13v2"/><path d="M9 13v2"/></svg> 小A · AI 编辑助手</h3>
      <button onclick="toggleAI()">✕</button>
    </div>
    <div class="ai-messages" id="aiMessages">
      <div class="ai-msg ai">告诉我你想怎么改这篇文章。比如「把第三段改口语化」「加一个总结段落」「帮我润色」</div>
    </div>
    <div class="ai-input-row">
      <input type="text" id="aiInput" placeholder="输入修改指令…" onkeydown="if(event.key==='Enter'&&!event.shiftKey)sendAI()">
      <button onclick="sendAI()" id="aiSendBtn">发送</button>
    </div>
  </div>`;
  document.body.insertAdjacentHTML('beforeend',aiHTML);
}

function toggleAI(){
  const p=document.getElementById('aiPanel');
  if(p)p.classList.toggle('open');
}

let _aiPending=null;
async function sendAI(){
  const inp=document.getElementById('aiInput');
  const msg=inp.value.trim();
  if(!msg)return;
  inp.value='';
  const msgs=document.getElementById('aiMessages');
  msgs.insertAdjacentHTML('beforeend',`<div class="ai-msg user">${msg.replace(/</g,'&lt;')}</div>`);
  msgs.insertAdjacentHTML('beforeend',`<div class="ai-msg ai loading" id="aiLoading">思考中…</div>`);
  msgs.scrollTop=msgs.scrollHeight;
  document.getElementById('aiSendBtn').disabled=true;

  try{
    const content=cm.getValue();
    // Check if there's a selection
    const sel=cm.getSelection();
    const body={instruction:msg,content:content};
    if(sel)body.selection=sel;
    if(currentFile)body.path=currentFile;

    const res=await fetch(api('api/ai-edit'),{
      method:'POST',
      headers:{'Content-Type':'application/json'},
      body:JSON.stringify(body)
    });
    const data=await res.json();
    document.getElementById('aiLoading').remove();
    if(data.ok){
      _aiPending=data.content;
      const preview=data.summary||'修改完成';
      msgs.insertAdjacentHTML('beforeend',`<div class="ai-msg ai">${preview.replace(/</g,'&lt;').replace(/\\n/g,'<br>')}<div class="ai-diff-actions"><button class="accept" onclick="acceptAI()">✅ 应用</button><button class="reject" onclick="rejectAI()">❌ 放弃</button></div></div>`);
    }else{
      msgs.insertAdjacentHTML('beforeend',`<div class="ai-msg ai error">❌ ${data.error||'请求失败'}</div>`);
    }
  }catch(e){
    document.getElementById('aiLoading')?.remove();
    msgs.insertAdjacentHTML('beforeend',`<div class="ai-msg ai error">❌ 网络错误: ${e.message}</div>`);
  }
  document.getElementById('aiSendBtn').disabled=false;
  msgs.scrollTop=msgs.scrollHeight;
}

function acceptAI(){
  if(_aiPending!=null){
    cm.setValue(_aiPending);
    _aiPending=null;
    dirty=true;
    renderPreview();
    const msgs=document.getElementById('aiMessages');
    msgs.insertAdjacentHTML('beforeend',`<div class="ai-msg ai" style="color:var(--green)">✅ 已应用到编辑器。记得保存！</div>`);
    msgs.scrollTop=msgs.scrollHeight;
  }
}
function rejectAI(){
  _aiPending=null;
  const msgs=document.getElementById('aiMessages');
  msgs.insertAdjacentHTML('beforeend',`<div class="ai-msg ai" style="color:var(--dim)">已放弃修改。</div>`);
  msgs.scrollTop=msgs.scrollHeight;
}

// ==================== ADMIN PANEL ====================



// ==================== FILE UPLOAD ====================
function getActiveDir(){
  // Get directory from current file or active folder context
  if(activeDir)return activeDir;
  if(currentFile){
    const idx=currentFile.lastIndexOf('/');
    return idx>0?currentFile.substring(0,idx):'';
  }
  return '';
}

function triggerUpload(destDir){
  if(destDir===undefined||destDir===null||destDir==='')destDir=getActiveDir();
  const inp=document.createElement('input');
  inp.type='file';inp.multiple=true;
  inp.onchange=()=>{
    if(inp.files.length)uploadFiles(inp.files,destDir);
  };
  inp.click();
}

async function uploadFiles(fileList,destDir){
  const fd=new FormData();
  fd.append('dir',destDir||'');
  for(const f of fileList)fd.append('file',f);
  showToast('Uploading '+fileList.length+' file(s)...');
  try{
    const r=await fetch(api('api/upload'),{method:'POST',body:fd});
    const d=await r.json();
    if(d.ok){
      const msg=d.uploaded.length+' uploaded'+(d.failed.length?' ('+d.failed.length+' failed)':'');
      showToast(msg);
      refreshTreeAndOpen(currentFile);
    }else{showToast(d.error||'Upload failed','var(--red)');}
  }catch(e){showToast('Upload failed: '+e.message,'var(--red)');}
}

// Drag & drop upload
let _dropOverlay=null;
document.addEventListener('dragover',e=>{
  if(!e.dataTransfer.types.includes('Files'))return;
  e.preventDefault();
  if(!_dropOverlay){
    _dropOverlay=document.createElement('div');
    _dropOverlay.className='drop-overlay';
    _dropOverlay.innerHTML='<span>Drop files to upload</span>';
    document.body.appendChild(_dropOverlay);
  }
});
document.addEventListener('dragleave',e=>{
  if(e.relatedTarget)return;
  if(_dropOverlay){_dropOverlay.remove();_dropOverlay=null;}
});
document.addEventListener('drop',e=>{
  e.preventDefault();
  if(_dropOverlay){_dropOverlay.remove();_dropOverlay=null;}
  if(e.dataTransfer.files.length){
    uploadFiles(e.dataTransfer.files,getActiveDir());
  }
});

// ==================== ADMIN TABS + ACTIVITY LOG ====================
let _activityOffset=0;
function switchAdminTab(tab){
  document.getElementById('adminTabUsersContent').style.display=tab==='users'?'':'none';
  document.getElementById('adminTabLogContent').style.display=tab==='log'?'':'none';
  document.getElementById('adminTabUsers').style.borderBottomColor=tab==='users'?'var(--purple)':'transparent';
  document.getElementById('adminTabUsers').style.color=tab==='users'?'var(--text)':'var(--dim)';
  document.getElementById('adminTabLog').style.borderBottomColor=tab==='log'?'var(--purple)':'transparent';
  document.getElementById('adminTabLog').style.color=tab==='log'?'var(--text)':'var(--dim)';
  if(tab==='log'){_activityOffset=0;loadActivity();}
}
async function loadActivity(){
  try{
    const r=await fetch(api('api/activity')+'?limit=50&offset='+_activityOffset);
    const d=await r.json();
    if(!d.ok)return;
    const actions={'save':'saved','delete':'deleted','move':'moved','restore':'restored','empty_trash':'emptied trash','share':'shared','user_create':'created user','user_edit':'edited user','user_delete':'deleted user'};
    const icons={'save':'\u1f4be','delete':'\u1f5d1','move':'\u27a1','restore':'\u267b','empty_trash':'\u1f5d1','share':'\u1f517','user_create':'\u2795','user_edit':'\u270f','user_delete':'\u274c'};
    let html='';
    if(!d.items.length) html='<div style="color:var(--dim);text-align:center;padding:20px">No activity yet</div>';
    d.items.forEach(it=>{
      const date=new Date(it.ts*1000);
      const time=date.toLocaleDateString('zh-CN',{month:'short',day:'numeric'})+' '+date.toLocaleTimeString('zh-CN',{hour:'2-digit',minute:'2-digit'});
      const act=actions[it.action]||it.action;
      const detail=it.detail?' \u2192 '+it.detail:'';
      const target=it.target?'<span style="color:var(--blue)">'+it.target+'</span>':'';
      html+='<div style="display:flex;gap:8px;padding:6px 0;border-bottom:1px solid var(--border);align-items:baseline">'
        +'<span style="color:var(--dim);white-space:nowrap;min-width:100px">'+time+'</span>'
        +'<span style="color:var(--accent);font-weight:600;min-width:60px">'+it.user+'</span>'
        +'<span>'+act+' '+target+detail+'</span>'
        +'</div>';
    });
    document.getElementById('activityLog').innerHTML=html;
    // Pager
    const total=d.total;
    const pager=document.getElementById('activityPager');
    let pg='';
    if(_activityOffset>0) pg+='<button onclick="_activityOffset-=50;loadActivity()" style="padding:4px 12px;border-radius:4px;border:1px solid var(--border);background:none;color:var(--text);cursor:pointer;font-size:12px">\u2190 Newer</button>';
    if(_activityOffset+50<total) pg+='<button onclick="_activityOffset+=50;loadActivity()" style="padding:4px 12px;border-radius:4px;border:1px solid var(--border);background:none;color:var(--text);cursor:pointer;font-size:12px">Older \u2192</button>';
    pager.innerHTML=pg;
  }catch(e){document.getElementById('activityLog').innerHTML='<div style="color:var(--red)">Load failed: '+e.message+'</div>';}
}

async function showAdminPanel(){
  document.querySelectorAll('.admin-overlay').forEach(d=>d.remove());
  try{
    const r=await fetch(api('api/users'));
    const d=await r.json();
    if(!d.ok){showToast('❌ '+d.error,'var(--red)');return;}
    const users=d.users;
    let rows='';
    for(const[uname,u] of Object.entries(users)){
      const roleClass={'admin':'role-admin','editor':'role-editor','viewer':'role-viewer'}[u.role]||'role-viewer';
      const paths=u.paths.join(', ');
      const canDel=uname!==AUTH_USER.user;
      rows+=`<tr>
        <td><strong>${uname}</strong></td>
        <td>${u.name||''}</td>
        <td><span class="role-badge ${roleClass}">${u.role}</span></td>
        <td style="font-size:11px;color:var(--dim);max-width:200px;overflow:hidden;text-overflow:ellipsis" title="${paths}">${paths}</td>
        <td style="font-size:11px;color:var(--dim)">${u.created||''}</td>
        <td>
          <button class="action-btn" onclick="editUser('${uname}')" title="编辑"><svg width="13" height="13" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M21.174 6.812a1 1 0 0 0-3.986-3.987L3.842 16.174a2 2 0 0 0-.5.83l-1.321 4.352a.5.5 0 0 0 .623.622l4.353-1.32a2 2 0 0 0 .83-.497z"/></svg></button>
          ${canDel?'<button class="action-btn danger" onclick="deleteUser(\''+uname+'\')" title="删除"><svg width="13" height="13" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M3 6h18"/><path d="M19 6v14c0 1-1 2-2 2H7c-1 0-2-1-2-2V6"/><path d="M8 6V4c0-1 1-2 2-2h4c1 0 2 1 2 2v2"/></svg></button>':''}
        </td>
      </tr>`;
    }
    const overlay=document.createElement('div');overlay.className='admin-overlay';
    overlay.innerHTML=`<div class="admin-panel">
      <div class="admin-header"><h2><svg width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M16 21v-2a4 4 0 0 0-4-4H6a4 4 0 0 0-4 4v2"/><circle cx="9" cy="7" r="4"/><path d="M22 21v-2a4 4 0 0 0-3-3.87"/><path d="M16 3.13a4 4 0 0 1 0 7.75"/></svg> 管理</h2><button onclick="this.closest('.admin-overlay').remove()">\u2715</button></div>
      <div style="display:flex;gap:0;border-bottom:1px solid var(--border);margin-bottom:12px">
        <button id="adminTabUsers" onclick="switchAdminTab('users')" style="flex:1;padding:8px;background:none;border:none;border-bottom:2px solid var(--purple);color:var(--text);cursor:pointer;font-size:13px;font-weight:600">用户</button>
        <button id="adminTabLog" onclick="switchAdminTab('log')" style="flex:1;padding:8px;background:none;border:none;border-bottom:2px solid transparent;color:var(--dim);cursor:pointer;font-size:13px">活动日志</button>
      </div>
      <div id="adminTabUsersContent"><div class="admin-body">
        <table class="admin-table">
          <thead><tr><th>用户名</th><th>名称</th><th>角色</th><th>路径权限</th><th>创建日期</th><th>操作</th></tr></thead>
          <tbody id="adminUserRows">${rows}</tbody>
        </table>
        <div class="admin-form" id="adminForm">
          <h3 id="adminFormTitle">➕ 添加用户</h3>
          <input type="hidden" id="adminEditMode" value="create">
          <input type="hidden" id="adminEditOrigUser" value="">
          <div class="form-row">
            <div><label>用户名</label><input type="text" id="adminUsername" placeholder="username"></div>
            <div><label>显示名称</label><input type="text" id="adminName" placeholder="显示名称"></div>
          </div>
          <div class="form-row">
            <div><label>密码 <span style="color:var(--dim);font-weight:normal">(编辑时留空不改)</span></label><input type="password" id="adminPassword" placeholder="密码"></div>
            <div><label>角色</label><select id="adminRole"><option value="admin">admin</option><option value="editor">editor</option><option value="viewer" selected>viewer</option></select></div>
          </div>
          <div><label>路径权限 <span style="color:var(--dim);font-weight:normal">(逗号分隔，* = 全部)</span></label><input type="text" id="adminPaths" placeholder="clients/C-0001/, output/" value="*"></div>
          <div class="form-actions">
            <button onclick="resetAdminForm()" style="background:var(--border);color:var(--text)">重置</button>
            <button onclick="submitAdminForm()" style="background:var(--purple);color:#fff" id="adminSubmitBtn">创建</button>
          </div>
        </div>
      </div></div>
      <div id="adminTabLogContent" style="display:none;padding:16px">
        <div id="activityLog" style="font-size:12px;max-height:60vh;overflow-y:auto"><div style="color:var(--dim);text-align:center;padding:20px">Loading...</div></div>
        <div style="display:flex;justify-content:center;gap:8px;margin-top:12px" id="activityPager"></div>
      </div>
    </div>`;
    overlay.addEventListener('click',e=>{if(e.target===overlay)overlay.remove();});
    document.body.appendChild(overlay);
  }catch(e){showToast('❌ 加载失败: '+e.message,'var(--red)');}
}

function resetAdminForm(){
  document.getElementById('adminEditMode').value='create';
  document.getElementById('adminEditOrigUser').value='';
  document.getElementById('adminFormTitle').textContent='添加用户';
  document.getElementById('adminSubmitBtn').textContent='创建';
  document.getElementById('adminUsername').value='';
  document.getElementById('adminUsername').disabled=false;
  document.getElementById('adminName').value='';
  document.getElementById('adminPassword').value='';
  document.getElementById('adminRole').value='viewer';
  document.getElementById('adminPaths').value='*';
}

async function editUser(uname){
  try{
    const r=await fetch(api('api/users'));
    const d=await r.json();
    if(!d.ok)return;
    const u=d.users[uname];
    if(!u)return;
    document.getElementById('adminEditMode').value='edit';
    document.getElementById('adminEditOrigUser').value=uname;
    document.getElementById('adminFormTitle').textContent='编辑: '+uname;
    document.getElementById('adminSubmitBtn').textContent='保存';
    document.getElementById('adminUsername').value=uname;
    document.getElementById('adminUsername').disabled=true;
    document.getElementById('adminName').value=u.name||'';
    document.getElementById('adminPassword').value='';
    document.getElementById('adminRole').value=u.role;
    document.getElementById('adminPaths').value=(u.paths||[]).join(', ');
  }catch(e){}
}

async function submitAdminForm(){
  const mode=document.getElementById('adminEditMode').value;
  const username=document.getElementById('adminUsername').value.trim();
  const name=document.getElementById('adminName').value.trim();
  const password=document.getElementById('adminPassword').value;
  const role=document.getElementById('adminRole').value;
  const paths=document.getElementById('adminPaths').value.split(',').map(s=>s.trim()).filter(Boolean);
  if(!username){showToast('⚠️ 用户名不能为空','var(--accent)');return;}
  if(mode==='create'&&!password){showToast('⚠️ 新用户必须设置密码','var(--accent)');return;}
  try{
    const body={username,name,role,paths};
    if(password)body.password=password;
    let r;
    if(mode==='create'){
      r=await fetch(api('api/users'),{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(body)});
    }else{
      r=await fetch(api('api/users/'+encodeURIComponent(username)),{method:'PUT',headers:{'Content-Type':'application/json'},body:JSON.stringify(body)});
    }
    const d=await r.json();
    if(d.ok){
      showToast('✅ '+(mode==='create'?'用户已创建':'用户已更新'));
      document.querySelectorAll('.admin-overlay').forEach(o=>o.remove());
      showAdminPanel();
    }else{showToast('❌ '+d.error,'var(--red)');}
  }catch(e){showToast('❌ 请求失败','var(--red)');}
}

async function deleteUser(uname){
  if(!confirm('确定删除用户 "'+uname+'" 吗？'))return;
  try{
    const r=await fetch(api('api/users/'+encodeURIComponent(uname)),{method:'DELETE'});
    const d=await r.json();
    if(d.ok){
      showToast('✅ 用户已删除');
      document.querySelectorAll('.admin-overlay').forEach(o=>o.remove());
      showAdminPanel();
    }else{showToast('❌ '+d.error,'var(--red)');}
  }catch(e){showToast('❌ 删除失败','var(--red)');}
}

function showPinDialog(){
  openModal('pinModal',`
    <h3 style="margin:0 0 4px;font-size:15px;display:flex;align-items:center;gap:8px"><svg width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><rect width="18" height="11" x="3" y="11" rx="2" ry="2"/><path d="M7 11V7a5 5 0 0 1 10 0v4"/></svg> PIN</h3>
    <p style="font-size:12px;color:var(--dim);margin:8px 0 16px">4 digit PIN for quick login. Leave empty for auto-login.</p>
    <input id="pinInput" class="xh-input" type="text" inputmode="numeric" maxlength="4" placeholder="4 digits" style="text-align:center;letter-spacing:8px;font-size:18px">
    <div style="display:flex;gap:8px;margin-top:16px;justify-content:flex-end">
      <button class="xh-btn xh-btn-ghost" onclick="closeModal('pinModal')">Cancel</button>
      <button class="xh-btn xh-btn-primary" onclick="savePin()">Save</button>
    </div>
  `,{narrow:true});
  setTimeout(()=>document.getElementById('pinInput')?.focus(),50);
}
async function savePin(){
  const pin=document.getElementById('pinInput').value.trim();
  if(pin&&(!/^\d{4}$/.test(pin))){showToast('⚠️ PIN 必须是 4 位数字','var(--red)');return;}
  const r=await fetch(api('api/pin'),{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({pin})});
  const d=await r.json();
  if(d.ok){showToast(pin?'✅ PIN 已设置':'✅ PIN 已清除（免密登录）','var(--green)');closeModal('pinModal');}
  else{showToast('❌ '+d.error,'var(--red)');}
}
async function doLogout(){
  try{
    await fetch(api('api/logout'),{method:'POST',headers:{'Content-Type':'application/json'},body:'{}'});
  }catch(e){}
  location.href=BASE+'login';
}
</script>
</body>
</html>'''

SIDEBAR_FULL = '''<div class="sidebar">
<div class="search" style="display:flex;gap:4px;align-items:center">
<input type="text" id="searchInput" placeholder="搜索文件..." oninput="filterFiles()" style="flex:1">
<button onclick="toggleAllFolders()" id="toggleFoldersBtn" style="background:none;color:var(--dim);border:1px solid var(--border);padding:4px 6px;border-radius:4px;font-size:11px;cursor:pointer;white-space:nowrap" title="展开/收起所有文件夹" title="Toggle folders"><svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="m6 14 1.5-2.9A2 2 0 0 1 9.24 10H20a2 2 0 0 1 1.94 2.5l-1.54 6a2 2 0 0 1-1.95 1.5H4a2 2 0 0 1-2-2V5a2 2 0 0 1 2-2h3.9a2 2 0 0 1 1.69.9l.81 1.2a2 2 0 0 0 1.67.9H18a2 2 0 0 1 2 2v2"/></svg></button>
<button onclick="triggerUpload()" style="background:none;color:var(--dim);border:1px solid var(--border);padding:4px 6px;border-radius:4px;font-size:11px;cursor:pointer;white-space:nowrap" title="Upload"><svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="17 8 12 3 7 8"/><line x1="12" y1="3" x2="12" y2="15"/></svg></button>
<button onclick="showNewFileDialog()" style="background:var(--green);color:#fff;border:none;padding:5px 8px;border-radius:4px;font-size:12px;cursor:pointer;white-space:nowrap" title="新建文件"><svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M5 12h14"/><path d="M12 5v14"/></svg></button>
</div>
<div class="recent-section" id="recentSection" style="display:none"><div class="recent-title"><svg width="11" height="11" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><circle cx="12" cy="12" r="10"/><polyline points="12 6 12 12 16 14"/></svg> Recent</div><div id="recentList"></div></div>
<div class="tree" id="tree"></div>
<div style="border-top:1px solid var(--border);padding:6px 8px;display:flex;align-items:center;gap:6px">
<button onclick="showTrashPanel()" style="background:none;border:1px solid var(--border);color:var(--dim);padding:4px 10px;border-radius:4px;font-size:12px;cursor:pointer;flex:1;text-align:left" title="回收站" title="Trash"><svg width="13" height="13" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M3 6h18"/><path d="M19 6v14c0 1-1 2-2 2H7c-1 0-2-1-2-2V6"/><path d="M8 6V4c0-1 1-2 2-2h4c1 0 2 1 2 2v2"/></svg> Trash</button>
</div>
<div class="file-count" id="fileCount"></div>
</div>'''

# ==================== ACTIVITY LOG ====================
_ACTIVITY_LOG = os.path.join(_EDITOR_DIR, 'activity.jsonl')
_activity_lock = threading.Lock()

def log_activity(user, action, target, detail=''):
    entry = {'ts': int(time.time()), 'user': user, 'action': action, 'target': target}
    if detail:
        entry['detail'] = detail
    with _activity_lock:
        with open(_ACTIVITY_LOG, 'a') as f:
            f.write(json.dumps(entry, ensure_ascii=False) + chr(10))

def get_activity(limit=100, offset=0):
    if not os.path.exists(_ACTIVITY_LOG):
        return [], 0
    with _activity_lock:
        with open(_ACTIVITY_LOG, 'r') as f:
            lines = f.readlines()
    total = len(lines)
    lines.reverse()
    selected = lines[offset:offset+limit]
    result = []
    for line in selected:
        line = line.strip()
        if line:
            try:
                result.append(json.loads(line))
            except:
                pass
    return result, total


# ==================== API TOKEN AUTH ====================
_TOKENS_PATH = os.path.join(_EDITOR_DIR, 'api_tokens.json')

def _load_api_tokens():
    if not os.path.exists(_TOKENS_PATH):
        return []
    try:
        with open(_TOKENS_PATH) as f:
            return json.load(f)
    except (json.JSONDecodeError, IOError):
        return []

def _check_api_token(auth_header):
    if not auth_header or not auth_header.startswith('Bearer '):
        return None
    token = auth_header[7:].strip()
    if not token.startswith('asf_'):
        return None
    token_hash = hashlib.sha256(token.encode()).hexdigest()
    tokens = _load_api_tokens()
    for t in tokens:
        if t.get('token_hash') == token_hash:
            t['last_used'] = time.strftime('%Y-%m-%dT%H:%M:%SZ')
            try:
                with open(_TOKENS_PATH, 'w') as f:
                    json.dump(tokens, f, indent=2)
            except:
                pass
            return t
    return None

def _has_permission(token_entry, perm):
    return perm in token_entry.get('permissions', [])


class Handler(SimpleHTTPRequestHandler):
    def _json_resp(self, code, data):
        self.send_response(code)
        self.send_header('Content-Type', 'application/json')
        self.end_headers()
        self.wfile.write(json.dumps(data).encode())

    def _save_custom_avatar(self, username, sig_id, avatar_data):
        """Save avatar from base64 data URL. Returns relative path or None."""
        if not avatar_data or ',' not in str(avatar_data):
            return None
        try:
            header, b64 = avatar_data.split(',', 1)
            ext = 'jpg'
            if 'png' in header: ext = 'png'
            elif 'gif' in header: ext = 'gif'
            elif 'webp' in header: ext = 'webp'
            avatar_bytes = base64.b64decode(b64)
            os.makedirs(os.path.join(_EDITOR_DIR, 'avatars'), exist_ok=True)
            avatar_filename = f'custom-{username}-{sig_id}.{ext}'
            avatar_path = os.path.join(_EDITOR_DIR, 'avatars', avatar_filename)
            with open(avatar_path, 'wb') as f:
                f.write(avatar_bytes)
            return f'avatars/{avatar_filename}'
        except Exception:
            return None

    def _check_auth(self):
        """Check auth and return session dict or None. Also cleans up expired sessions."""
        cleanup_sessions()
        token = _get_auth_cookie(self)
        return get_session(token)

    def _check_device_token(self):
        """Check for device_token cookie, return (username, user_info) or (None, None)."""
        cookie_str = self.headers.get('Cookie', '')
        if not cookie_str:
            return None, None
        cookies = SimpleCookie()
        try:
            cookies.load(cookie_str)
        except:
            return None, None
        dev_cookie = cookies.get('device_token')
        if not dev_cookie:
            return None, None
        return verify_device_token(dev_cookie.value)

    def _serve_pin_page(self, username, redirect_to='', error=''):
        """Show PIN entry page or auto-login if no PIN is set."""
        users_data = get_users()
        user_info = users_data.get('users', {}).get(username)
        if not user_info:
            self._serve_login()
            return
        # No PIN set → auto-login immediately
        if not user_info.get('pin'):
            token = create_session(username, user_info)
            expire_hours = users_data.get('settings', {}).get('sessionExpireHours', 72)
            max_age = int(expire_hours * 3600)
            self.send_response(302)
            self.send_header('Set-Cookie', f'auth_token={token}; Path=/; HttpOnly; Max-Age={max_age}')
            self.send_header('Location', _safe_redirect(redirect_to) or './')
            self.end_headers()
            return
        # Has PIN → show PIN page
        pin_html = f'''<!DOCTYPE html><html><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>🔐 PIN</title>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{background:#0d1117;color:#e6edf3;font-family:-apple-system,BlinkMacSystemFont,sans-serif;display:flex;align-items:center;justify-content:center;min-height:100vh}}
.pin-box{{background:#161b22;border:1px solid #30363d;border-radius:12px;padding:40px;width:320px;max-width:90vw;text-align:center}}
.pin-box h2{{color:#ffa500;margin-bottom:8px;font-size:18px}}
.pin-box .subtitle{{color:#7d8590;font-size:13px;margin-bottom:24px}}
.pin-box .error{{background:rgba(248,81,73,.1);border:1px solid rgba(248,81,73,.3);color:#f85149;padding:8px;border-radius:6px;font-size:13px;margin-bottom:16px;display:{'block' if error else 'none'}}}
.pin-input{{display:flex;gap:8px;justify-content:center;margin-bottom:20px}}
.pin-input input{{width:48px;height:56px;text-align:center;font-size:24px;background:#0d1117;color:#e6edf3;border:2px solid #30363d;border-radius:8px;outline:none}}
.pin-input input:focus{{border-color:#ffa500}}
.pin-box button{{width:100%;background:#238636;color:#fff;border:none;padding:12px;border-radius:6px;font-size:14px;font-weight:600;cursor:pointer}}
.pin-box button:hover{{background:#2ea043}}
.pin-box .alt{{margin-top:16px;font-size:12px;color:#7d8590}}
.pin-box .alt a{{color:#ffa500;text-decoration:none;cursor:pointer}}
</style></head><body>
<div class="pin-box">
<h2>👋 {user_info.get('name', username)}</h2>
<div class="subtitle">请输入 PIN 码</div>
<div class="error">{error}</div>
<form method="POST" action="login">
<input type="hidden" name="username" value="{username}">
<input type="hidden" name="pin_login" value="1">
<input type="hidden" name="redirect" value="{redirect_to}">
<div class="pin-input">
<input type="password" name="p1" maxlength="1" inputmode="numeric" autofocus
  oninput="if(this.value)this.nextElementSibling?.focus()">
<input type="password" name="p2" maxlength="1" inputmode="numeric"
  oninput="if(this.value)this.nextElementSibling?.focus()">
<input type="password" name="p3" maxlength="1" inputmode="numeric"
  oninput="if(this.value)this.nextElementSibling?.focus()">
<input type="password" name="p4" maxlength="1" inputmode="numeric"
  oninput="if(this.value)document.getElementById('pinSubmit').click()">
</div>
<input type="hidden" name="pin" id="pinVal" value="">
<button type="submit" id="pinSubmit" onclick="document.getElementById('pinVal').value=
  (document.querySelector('[name=p1]').value||'')+
  (document.querySelector('[name=p2]').value||'')+
  (document.querySelector('[name=p3]').value||'')+
  (document.querySelector('[name=p4]').value||'')">确认</button>
</form>
<div class="alt">不是你？<a href="login" onclick="document.cookie='device_token=;Path=/;Max-Age=0'">切换账号</a></div>
</div></body></html>'''
        self.send_response(200)
        self.send_header('Content-Type', 'text/html; charset=utf-8')
        self.send_header('Cache-Control', 'no-store')
        self.end_headers()
        self.wfile.write(pin_html.encode())

    def _check_setup(self):
        if not _setup.needs_setup(_EDITOR_DIR):
            return False
        path = self.path.split('?')[0].rstrip('/')
        if path == '/_setup':
            return False
        self.send_response(302)
        self.send_header('Location', '/_setup')
        self.end_headers()
        return True

    def _serve_login(self, error='', redirect_to=''):
        html = LOGIN_HTML
        html = html.replace('__ERROR_DISPLAY__', 'block' if error else 'none')
        html = html.replace('__ERROR_MSG__', error)
        html = html.replace('__REDIRECT__', redirect_to)
        # Build login action URL that works behind reverse proxy
        html = html.replace('__LOGIN_ACTION__', 'login')
        self.send_response(200)
        self.send_header('Content-Type', 'text/html; charset=utf-8')
        self.send_header('Cache-Control', 'no-store')
        self.end_headers()
        self.wfile.write(html.encode())

    def _redirect(self, location):
        self.send_response(302)
        self.send_header('Location', location)
        self.end_headers()

    def _auth_user_json(self, session):
        """Return JSON string for AUTH_USER JS variable."""
        if not session:
            return 'null'
        return json.dumps({'user': session['user'], 'name': session['name'], 'role': session['role'], 'paths': session['paths']})

    def _make_auth_buttons(self, session):
        """Return admin_btn and logout_btn HTML for toolbar."""
        if not session:
            return '', ''
        admin_btn = ''
        if session['role'] == 'admin':
            admin_btn = '<button class="admin-btn" onclick="showAdminPanel()" title="用户管理"><svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M16 21v-2a4 4 0 0 0-4-4H6a4 4 0 0 0-4 4v2"/><circle cx="9" cy="7" r="4"/><path d="M22 21v-2a4 4 0 0 0-3-3.87"/><path d="M16 3.13a4 4 0 0 1 0 7.75"/></svg></button>'
        logout_btn = f'<span class="user-info"><span class="username">{session["name"]}</span><button class="logout-btn" onclick="showPinDialog()" title="设置 PIN 码"><svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><rect width="18" height="11" x="3" y="11" rx="2" ry="2"/><path d="M7 11V7a5 5 0 0 1 10 0v4"/></svg></button><button class="logout-btn" onclick="doLogout()" title="登出"><svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M9 21H5a2 2 0 0 1-2-2V5a2 2 0 0 1 2-2h4"/><polyline points="16 17 21 12 16 7"/><line x1="21" y1="12" x2="9" y2="12"/></svg></button></span>'
        return admin_btn, logout_btn

    def _serve_html(self, mode_config):
        """Serve the editor HTML with given config dict."""
        html = HTML_TEMPLATE
        html = html.replace('__FULL_MODE__', mode_config.get('full_mode', 'false'))
        html = html.replace('__SHARE_FILE__', mode_config.get('share_file', 'null'))
        html = html.replace('__SIDEBAR_HTML__', mode_config.get('sidebar', ''))
        html = html.replace('__INITIAL_FNAME__', mode_config.get('fname', ''))
        html = html.replace('__SAVE_DISPLAY__', mode_config.get('save_display', ''))
        html = html.replace('__EMPTY_MSG__', mode_config.get('empty_msg', ''))
        html = html.replace('__IS_READONLY__', mode_config.get('is_readonly', 'false'))
        html = html.replace('__SHARE_BTN_DISPLAY__', mode_config.get('share_btn_display', ''))
        html = html.replace('__ANN_BTN_DISPLAY__', mode_config.get('ann_btn_display', ''))
        html = html.replace('__READONLY_BANNER__', mode_config.get('readonly_banner', ''))
        html = html.replace('__EDITOR_TITLE__', mode_config.get('editor_title', '📝 修荷 Editor'))
        html = html.replace('__DEFAULT_VIEW__', mode_config.get('default_view', 'split'))
        html = html.replace('__AUTH_USER__', mode_config.get('auth_user', 'null'))
        html = html.replace('__THEME_BTN__', '<button class="theme-btn" onclick="toggleTheme()" title="\xe5\x88\x87\xe6\x8d\xa2\xe4\xb8\xbb\xe9\xa2\x98"><svg id="themeIcon" width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M12 3a6 6 0 0 0 9 9 9 9 0 1 1-9-9Z"/></svg></button>')
        html = html.replace('__ADMIN_BTN__', mode_config.get('admin_btn', ''))
        html = html.replace('__LOGOUT_BTN__', mode_config.get('logout_btn', ''))
        self.send_response(200)
        self.send_header('Content-Type', 'text/html; charset=utf-8')
        self.send_header('Cache-Control', 'no-store, no-cache, must-revalidate, max-age=0')
        self.end_headers()
        self.wfile.write(html.encode())

    def do_GET(self):
        path = self.path.split('?')[0].rstrip('/')

        # ==================== AUTH MIDDLEWARE ====================
        # Public paths: /login, /s/{token}/*
        is_public = path in ('/login', '') or path.startswith('/s/')
        if path == '/login' or path == '/login/':
            # Serve login page
            session = self._check_auth()
            if session:
                self._redirect('./')
                return
            # Check if device is remembered → show PIN or auto-login
            dev_user, dev_info = self._check_device_token()
            if dev_user and dev_info:
                qs = parse_qs(urlparse(self.path).query)
                redirect_to = _safe_redirect(qs.get('redirect', ['./'])[0])
                self._serve_pin_page(dev_user, redirect_to=redirect_to)
                return
            self._serve_login()
            return

        # Share links are public — skip auth
        if not path.startswith('/s/'):
            session = self._check_auth()
            if not session:
                # Check device token — remembered device
                dev_user, dev_info = self._check_device_token()
                if dev_user and dev_info:
                    # Device remembered: show PIN or auto-login
                    redirect_path = _safe_redirect(self.path)
                    self._serve_pin_page(dev_user, redirect_to=redirect_path)
                    return
                # No device token → full login
                redirect_path = _safe_redirect(self.path)
                self._redirect('login?redirect=' + url_quote(redirect_path))
                return
        else:
            session = None  # share links don't need auth

        # Serve local images for preview
        if path == '/api/img':
            qs = parse_qs(urlparse(self.path).query)
            img_path = qs.get('p', [''])[0]
            if not img_path:
                self.send_error(400); return
            # Resolve: absolute path or relative to workspace
            if os.path.isabs(img_path):
                full = os.path.normpath(img_path)
            else:
                full = os.path.normpath(os.path.join(WORKSPACE, img_path))
            # Security: must be under workspace
            if not full.startswith(WORKSPACE) or not os.path.isfile(full):
                self.send_error(404); return
            mime = mimetypes.guess_type(full)[0] or 'application/octet-stream'
            with open(full, 'rb') as f:
                data = f.read()
            self.send_response(200)
            self.send_header('Content-Type', mime)
            self.send_header('Content-Length', str(len(data)))
            self.send_header('Cache-Control', 'max-age=3600')
            self.end_headers()
            self.wfile.write(data)
            return

        # Download exported file
        if path.startswith('/api/download/'):
            dl_id = path[len('/api/download/'):]
            with _downloads_lock:
                info = _downloads.pop(dl_id, None)
            if not info or not os.path.isfile(info['path']):
                self._json_resp(404, {'ok': False, 'error': 'Download not found or expired'})
                return
            fpath = info['path']
            fname = info['filename']
            mime = 'application/pdf' if fname.endswith('.pdf') else 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
            try:
                with open(fpath, 'rb') as f:
                    data = f.read()
                self.send_response(200)
                self.send_header('Content-Type', mime)
                self.send_header('Content-Disposition', f'attachment; filename*=UTF-8\'\'{url_quote(fname)}')
                self.send_header('Content-Length', str(len(data)))
                self.end_headers()
                self.wfile.write(data)
            finally:
                try: os.unlink(fpath)
                except: pass
            return

        # Share URL: /s/{token} or /s/{token}/api/*
        if path.startswith('/s/'):
            parts = path[3:].split('/', 1)
            token = parts[0]
            sub_path = '/' + parts[1] if len(parts) > 1 else ''
            share = get_share(token)
            if not share:
                self.send_error(404, 'Share link expired or invalid')
                return
            is_readonly = share['mode'] == 'readonly'
            share_type = share.get('type', 'file')
            
            if share_type == 'folder':
                # --- FOLDER SHARE ---
                folder_rel = share['file']  # relative folder path
                folder_full = os.path.normpath(os.path.join(WORKSPACE, folder_rel))
                allowed_files = share.get('files', [])  # relative to folder
                
                if sub_path == '/api/files':
                    # Return only allowed files
                    file_list = []
                    for f in allowed_files:
                        full = os.path.join(folder_full, f)
                        if os.path.isfile(full):
                            file_list.append({'name': os.path.basename(f), 'path': f})
                    self._json_resp(200, {'files': file_list})
                    return
                elif sub_path == '/api/load':
                    qs = parse_qs(urlparse(self.path).query)
                    rel = qs.get('path', [''])[0]
                    if rel not in allowed_files:
                        self._json_resp(403, {'error': 'File not in share'}); return
                    full = os.path.normpath(os.path.join(folder_full, rel))
                    if not full.startswith(folder_full) or not os.path.isfile(full):
                        self.send_error(404); return
                    with open(full, 'r', errors='replace') as f:
                        content = f.read()
                    self._json_resp(200, {'content': content, 'filename': os.path.basename(full)})
                    return
                elif sub_path == '/api/save':
                    pass  # handled in do_POST
                elif sub_path == '/api/img':
                    qs = parse_qs(urlparse(self.path).query)
                    img_path = qs.get('p', [''])[0]
                    if not img_path: self.send_error(400); return
                    full = os.path.normpath(img_path) if os.path.isabs(img_path) else os.path.normpath(os.path.join(WORKSPACE, img_path))
                    if not full.startswith(WORKSPACE) or not os.path.isfile(full): self.send_error(404); return
                    mime = mimetypes.guess_type(full)[0] or 'application/octet-stream'
                    with open(full, 'rb') as f: data = f.read()
                    self.send_response(200)
                    self.send_header('Content-Type', mime)
                    self.send_header('Content-Length', str(len(data)))
                    self.send_header('Cache-Control', 'max-age=3600')
                    self.end_headers()
                    self.wfile.write(data)
                    return
                elif sub_path.startswith('/api/download/'):
                    dl_id = sub_path[len('/api/download/'):]
                    with _downloads_lock:
                        info = _downloads.pop(dl_id, None)
                    if not info or not os.path.isfile(info['path']):
                        self._json_resp(404, {'ok': False, 'error': 'Download not found or expired'}); return
                    fpath, fname_dl = info['path'], info['filename']
                    mime = 'application/pdf' if fname_dl.endswith('.pdf') else 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
                    try:
                        with open(fpath, 'rb') as f: data = f.read()
                        self.send_response(200)
                        self.send_header('Content-Type', mime)
                        self.send_header('Content-Disposition', f'attachment; filename*=UTF-8\'\'{url_quote(fname_dl)}')
                        self.send_header('Content-Length', str(len(data)))
                        self.end_headers()
                        self.wfile.write(data)
                    finally:
                        try: os.unlink(fpath)
                        except: pass
                    return
                elif sub_path == '/api/config':
                    cfg = get_config()
                    safe_config = {'title': cfg.get('title', 'Editor'), 'signatures': {}, 'pdfThemes': cfg.get('pdfThemes', {}), 'pdfLayouts': cfg.get('pdfLayouts', {}), 'wordStyles': {}, 'wordMargins': {}, 'wordPages': {}}
                    for k, v in cfg.get('signatures', {}).items():
                        safe_config['signatures'][k] = None if v is None else {'name': v.get('name', k), 'hasAvatar': bool(v.get('avatar'))}
                    for section in ('wordStyles', 'wordMargins', 'wordPages'):
                        for k, v in cfg.get(section, {}).items():
                            safe_config[section][k] = {'label': v.get('label', k)}
                            if v.get('default'): safe_config[section][k]['default'] = True
                    self._json_resp(200, safe_config)
                    return
                elif sub_path.startswith('/api/'):
                    self.send_error(403); return
                elif sub_path and sub_path != '/':
                    self.send_error(404); return
                
                # Serve folder share HTML — uses FULL_MODE=true for sidebar but with limited file list
                folder_name = os.path.basename(folder_rel) or folder_rel
                banner_text = '👁 只读模式' if is_readonly else '✏️ 可编辑模式'
                banner_bg = '' if is_readonly else ' style="background:rgba(255,165,0,.1);border-color:var(--accent);color:var(--accent)"'
                self._serve_html({
                    'full_mode': 'true',  # enables sidebar + file tree
                    'share_file': 'null',
                    'sidebar': SIDEBAR_FULL.replace('🔍 搜索...', '🔍 搜索...').replace(
                        'onclick="showNewFileDialog()"', 'style="display:none"'
                    ),
                    'fname': '选择文件查看',
                    'save_display': 'style="display:none"',
                    'empty_msg': '📁 选择左侧文件查看',
                    'is_readonly': 'true' if is_readonly else 'false',
                    'share_btn_display': 'style="display:none"',
                    'ann_btn_display': 'style="display:none"',
                    'readonly_banner': f'<div class="readonly-banner"{banner_bg}>{banner_text} — 📁 {folder_name} · {len(allowed_files)} 个文件</div>',
                    'editor_title': '📁 ' + folder_name,
                    'default_view': 'preview-only' if is_readonly else 'split',
                })
                return
            
            else:
                # --- SINGLE FILE SHARE ---
                full_path = os.path.normpath(os.path.join(WORKSPACE, share['file']))
                if not os.path.isfile(full_path):
                    self.send_error(404, 'File not found')
                    return
                
                if sub_path == '/api/load':
                    with open(full_path, 'r', errors='replace') as f:
                        content = f.read()
                    self._json_resp(200, {'content': content, 'filename': os.path.basename(full_path)})
                    return
                if sub_path == '/api/save':
                    if is_readonly:
                        self._json_resp(403, {'ok': False, 'error': 'Read-only share'})
                        return
                    pass  # handled in do_POST
                if sub_path == '/api/img':
                    qs = parse_qs(urlparse(self.path).query)
                    img_path = qs.get('p', [''])[0]
                    if not img_path: self.send_error(400); return
                    full = os.path.normpath(img_path) if os.path.isabs(img_path) else os.path.normpath(os.path.join(WORKSPACE, img_path))
                    if not full.startswith(WORKSPACE) or not os.path.isfile(full): self.send_error(404); return
                    mime = mimetypes.guess_type(full)[0] or 'application/octet-stream'
                    with open(full, 'rb') as f: data = f.read()
                    self.send_response(200)
                    self.send_header('Content-Type', mime)
                    self.send_header('Content-Length', str(len(data)))
                    self.send_header('Cache-Control', 'max-age=3600')
                    self.end_headers()
                    self.wfile.write(data)
                    return
                if sub_path.startswith('/api/download/'):
                    dl_id = sub_path[len('/api/download/'):]
                    with _downloads_lock:
                        info = _downloads.pop(dl_id, None)
                    if not info or not os.path.isfile(info['path']):
                        self._json_resp(404, {'ok': False, 'error': 'Download not found or expired'}); return
                    fpath, fname_dl = info['path'], info['filename']
                    mime = 'application/pdf' if fname_dl.endswith('.pdf') else 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
                    try:
                        with open(fpath, 'rb') as f: data = f.read()
                        self.send_response(200)
                        self.send_header('Content-Type', mime)
                        self.send_header('Content-Disposition', f'attachment; filename*=UTF-8\'\'{url_quote(fname_dl)}')
                        self.send_header('Content-Length', str(len(data)))
                        self.end_headers()
                        self.wfile.write(data)
                    finally:
                        try: os.unlink(fpath)
                        except: pass
                    return
                if sub_path == '/api/config':
                    cfg = get_config()
                    safe_config = {'title': cfg.get('title', 'Editor'), 'signatures': {}, 'pdfThemes': cfg.get('pdfThemes', {}), 'pdfLayouts': cfg.get('pdfLayouts', {}), 'wordStyles': {}, 'wordMargins': {}, 'wordPages': {}}
                    for k, v in cfg.get('signatures', {}).items():
                        safe_config['signatures'][k] = None if v is None else {'name': v.get('name', k), 'hasAvatar': bool(v.get('avatar'))}
                    for section in ('wordStyles', 'wordMargins', 'wordPages'):
                        for k, v in cfg.get(section, {}).items():
                            safe_config[section][k] = {'label': v.get('label', k)}
                            if v.get('default'): safe_config[section][k]['default'] = True
                    self._json_resp(200, safe_config)
                    return
                if sub_path.startswith('/api/'):
                    self.send_error(403)
                    return
                if sub_path and sub_path != '/':
                    self.send_error(404)
                    return
                
                fname = os.path.basename(share['file'])
                self._serve_html({
                    'full_mode': 'false',
                    'share_file': json.dumps(full_path),
                    'sidebar': '',
                    'fname': fname,
                    'save_display': 'style="display:none"' if is_readonly else '',
                    'empty_msg': '加载中...',
                    'is_readonly': 'true' if is_readonly else 'false',
                    'share_btn_display': 'style="display:none"',
                    'ann_btn_display': 'style="display:none"' if is_readonly else '',
                    'readonly_banner': '<div class="readonly-banner">👁 只读模式 — 此文档由修荷团队分享</div>' if is_readonly else '<div class="readonly-banner" style="background:rgba(255,165,0,.1);border-color:var(--accent);color:var(--accent)">✏️ 可编辑模式 — 此文档由修荷团队分享</div>',
                    'editor_title': '📄 ' + fname,
                    'default_view': 'preview-only' if is_readonly else 'split',
                })
                return

        if path == '' or path == '/':
            admin_btn, logout_btn = self._make_auth_buttons(session)
            auth_user_json = self._auth_user_json(session)
            if FULL_MODE:
                cfg = get_config()
                editor_title = '📝 ' + cfg.get('title', 'Editor')
                self._serve_html({
                    'full_mode': 'true',
                    'share_file': 'null',
                    'sidebar': SIDEBAR_FULL,
                    'fname': '选择文件开始编辑',
                    'save_display': 'style="display:none"',
                    'empty_msg': '📄 选择一个文件开始编辑',
                    'is_readonly': 'false',
                    'share_btn_display': '',
                    'ann_btn_display': '',
                    'readonly_banner': '',
                    'editor_title': editor_title,
                    'auth_user': auth_user_json,
                    'admin_btn': admin_btn,
                    'logout_btn': logout_btn,
                })
            else:
                cfg = get_config()
                editor_title = '📝 ' + cfg.get('title', 'Editor')
                fname = os.path.basename(SHARE_FILE)
                self._serve_html({
                    'full_mode': 'false',
                    'share_file': json.dumps(SHARE_FILE),
                    'sidebar': '',
                    'fname': fname,
                    'save_display': '',
                    'empty_msg': '加载中...',
                    'is_readonly': 'false',
                    'share_btn_display': 'style="display:none"',
                    'ann_btn_display': '',
                    'readonly_banner': '',
                    'editor_title': editor_title,
                    'auth_user': auth_user_json,
                    'admin_btn': admin_btn,
                    'logout_btn': logout_btn,
                })
        elif path == '/api/config':
            # Return safe subset of config for the export UI (no AI keys or internals)
            cfg = get_config()
            safe_config = {
                'title': cfg.get('title', 'Editor'),
                'signatures': {},
                'pdfThemes': cfg.get('pdfThemes', {'light': {'label': '☀️ 浅色', 'default': True}, 'dark': {'label': '🌙 暗色'}, 'sepia': {'label': '📜 复古'}}),
                'pdfLayouts': cfg.get('pdfLayouts', {'a4': {'label': 'A4', 'default': True}, 'a4-compact': {'label': 'A4 紧凑'}, 'letter': {'label': 'Letter'}, 'single': {'label': '📜 单页'}}),
                'wordStyles': {},
                'wordMargins': {},
                'wordPages': {},
            }
            # Build signatures (strip avatar paths, just expose keys and names)
            for k, v in cfg.get('signatures', {}).items():
                if v is None:
                    safe_config['signatures'][k] = None
                else:
                    safe_config['signatures'][k] = {'name': v.get('name', k), 'hasAvatar': bool(v.get('avatar'))}
            # Build word configs (strip internal details, just labels and defaults)
            for section in ('wordStyles', 'wordMargins', 'wordPages'):
                for k, v in cfg.get(section, {}).items():
                    safe_config[section][k] = {'label': v.get('label', k)}
                    if v.get('default'):
                        safe_config[section][k]['default'] = True
            self._json_resp(200, safe_config)
        elif path == '/api/custom-signatures':
            # List all custom signatures for the current user
            if not session:
                self._json_resp(401, {'error': 'Not logged in'}); return
            username = session['user']
            users_data = get_users()
            user = users_data.get('users', {}).get(username, {})
            sigs = user.get('customSignatures', [])
            # Migrate from old single customSignature if exists
            old_cs = user.get('customSignature')
            if old_cs and not sigs:
                sigs = [{'id': secrets.token_urlsafe(6), 'name': old_cs.get('name', ''), 'avatar': old_cs.get('avatar')}]
                user['customSignatures'] = sigs
                user.pop('customSignature', None)
                _save_users(users_data)
            result = []
            for s in sigs:
                avatar_url = None
                if s.get('avatar'):
                    ap = os.path.join(_EDITOR_DIR, s['avatar']) if not os.path.isabs(s.get('avatar', '')) else s['avatar']
                    if os.path.isfile(ap):
                        avatar_url = 'api/custom-avatar?id=' + url_quote(s['id']) + '&u=' + url_quote(username) + '&t=' + str(int(os.path.getmtime(ap)))
                result.append({'id': s['id'], 'name': s.get('name', ''), 'avatarUrl': avatar_url})
            self._json_resp(200, {'ok': True, 'signatures': result})
        elif path.startswith('/api/custom-avatar'):
            # Serve custom avatar image
            qs = parse_qs(urlparse(self.path).query)
            username = qs.get('u', [''])[0]
            sig_id = qs.get('id', [''])[0]
            if not username:
                self.send_error(400); return
            users_data = get_users()
            user = users_data.get('users', {}).get(username, {})
            sigs = user.get('customSignatures', [])
            avatar_rel = None
            for s in sigs:
                if s.get('id') == sig_id:
                    avatar_rel = s.get('avatar', '')
                    break
            if not avatar_rel:
                self.send_error(404); return
            avatar_path = os.path.join(_EDITOR_DIR, avatar_rel) if not os.path.isabs(avatar_rel) else avatar_rel
            if not os.path.isfile(avatar_path):
                self.send_error(404); return
            mime = mimetypes.guess_type(avatar_path)[0] or 'image/jpeg'
            with open(avatar_path, 'rb') as f:
                data = f.read()
            self.send_response(200)
            self.send_header('Content-Type', mime)
            self.send_header('Content-Length', str(len(data)))
            self.send_header('Cache-Control', 'max-age=86400')
            self.end_headers()
            self.wfile.write(data)
            return
        elif path == '/api/shares':
            # List shares for a specific file
            qs = parse_qs(urlparse(self.path).query)
            file_filter = qs.get('file', [''])[0]
            shares = _load_shares()
            result = []
            for token, info in shares.items():
                if not file_filter or info.get('file') == file_filter:
                    result.append({**info, 'token': token})
            self._json_resp(200, {'shares': result})
        elif path == '/api/files':
            if not FULL_MODE:
                self.send_error(403); return
            skip = {'.git','node_modules','.next','dist','__pycache__','.cache','.trash','.turbo','.venv','.npm','.config','.local','.ssh'}
            VIEWABLE_EXTS = {'.md', '.pdf', '.docx', '.xlsx', '.xls', '.jpg', '.jpeg', '.png', '.gif', '.webp', '.txt', '.json', '.csv', '.html', '.htm'}
            md_files = []
            user_paths = session['paths'] if session else ['*']
            walk_roots = SCOPED_DIRS if SCOPED_DIRS else [WORKSPACE]
            for walk_root in walk_roots:
                for root, dirs, fnames in os.walk(walk_root):
                    dirs[:] = sorted([d for d in dirs if d not in skip])
                    for f in sorted(fnames):
                        ext = os.path.splitext(f)[1].lower()
                        if ext in VIEWABLE_EXTS:
                            full = os.path.join(root, f)
                            rel = os.path.relpath(full, WORKSPACE)
                            if check_path_access(user_paths, rel):
                                md_files.append({'name': f, 'path': rel, 'type': ext})
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'files': md_files}).encode())
        elif path == '/api/load':
            if FULL_MODE:
                qs = parse_qs(urlparse(self.path).query)
                rel = qs.get('path', [''])[0]
                full = os.path.normpath(os.path.join(WORKSPACE, rel))
                if not full.startswith(WORKSPACE) or not os.path.isfile(full):
                    self.send_error(404); return
                # Path access check
                if session and not check_path_access(session['paths'], rel):
                    self._json_resp(403, {'error': '无权访问此文件'}); return
            else:
                full = SHARE_FILE
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            with open(full, 'r', errors='replace') as f:
                content = f.read()
            self.wfile.write(json.dumps({'content': content, 'filename': os.path.basename(full)}).encode())
        elif path == '/api/preview':
            # Convert non-MD files to HTML for preview
            qs = parse_qs(urlparse(self.path).query)
            rel = qs.get('path', [''])[0]
            full = os.path.normpath(os.path.join(WORKSPACE, rel))
            if not full.startswith(WORKSPACE) or not os.path.isfile(full):
                self.send_error(404); return
            if session and not check_path_access(session['paths'], rel):
                self._json_resp(403, {'error': '无权访问此文件'}); return
            ext = os.path.splitext(full)[1].lower()
            html = ''
            try:
                if ext == '.pdf':
                    # Render PDF pages as base64 images using pymupdf
                    import fitz
                    doc = fitz.open(full)
                    pages = []
                    for i, page in enumerate(doc):
                        mat = fitz.Matrix(1.5, 1.5)  # 150 DPI
                        pix = page.get_pixmap(matrix=mat)
                        img_bytes = pix.tobytes("jpeg", jpg_quality=85)
                        b64 = base64.b64encode(img_bytes).decode()
                        pages.append(f'<div style="margin:12px 0;text-align:center"><img src="data:image/jpeg;base64,{b64}" style="max-width:100%;border:1px solid #30363d;border-radius:4px" alt="Page {i+1}"><div style="color:#7d8590;font-size:11px;margin-top:4px">Page {i+1}/{len(doc)}</div></div>')
                    doc.close()
                    html = '\n'.join(pages)
                elif ext == '.docx':
                    import mammoth
                    with open(full, 'rb') as f:
                        result = mammoth.convert_to_html(f)
                        html = f'<div class="docx-preview">{result.value}</div>'
                        if result.messages:
                            html += f'<div style="color:#7d8590;font-size:11px;margin-top:12px">⚠️ {len(result.messages)} warning(s)</div>'
                elif ext in ('.xlsx', '.xls'):
                    import openpyxl
                    wb = openpyxl.load_workbook(full, read_only=True, data_only=True)
                    sheets_html = []
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        rows = list(ws.iter_rows(values_only=True))
                        if not rows:
                            sheets_html.append(f'<h3 style="color:#ffa500">{sheet_name}</h3><p style="color:#7d8590">空表</p>')
                            continue
                        tbl = f'<h3 style="color:#ffa500;margin:16px 0 8px">{sheet_name}</h3><div style="overflow-x:auto"><table style="border-collapse:collapse;width:100%;font-size:13px">'
                        for ri, row in enumerate(rows):
                            tag = 'th' if ri == 0 else 'td'
                            cells = ''.join(f'<{tag} style="border:1px solid #30363d;padding:6px 10px;background:{"#161b22" if ri==0 else "transparent"}">{cell if cell is not None else ""}</{tag}>' for cell in row)
                            tbl += f'<tr>{cells}</tr>'
                        tbl += '</table></div>'
                        sheets_html.append(tbl)
                    wb.close()
                    html = '\n'.join(sheets_html)
                elif ext in ('.jpg', '.jpeg', '.png', '.gif', '.webp', '.svg'):
                    # Use direct URL instead of base64 for fast loading
                    from urllib.parse import quote as _q
                    img_url = 'api/raw?path=' + _q(rel, safe='')
                    size = os.path.getsize(full)
                    size_str = f'{size/1024:.0f} KB' if size < 1048576 else f'{size/1048576:.1f} MB'
                    html = f'<div style="text-align:center;padding:20px"><img src="__IMG_URL__" style="max-width:100%;border-radius:6px;border:1px solid #30363d" alt="{os.path.basename(full)}"><div style="color:#7d8590;font-size:12px;margin-top:8px">{os.path.basename(full)} ({size_str})</div></div>'.replace('__IMG_URL__', img_url)
                elif ext in ('.html', '.htm'):
                    # For HTML files, use an iframe — api() in JS adds correct prefix
                    from urllib.parse import quote
                    html = f'<iframe id="html-preview-frame" data-raw-path="{quote(rel)}" style="width:100%;min-height:80vh;border:1px solid #30363d;border-radius:6px;background:#0D0D0F"></iframe><script>document.getElementById("html-preview-frame").src=api("api/raw")+"?path="+encodeURIComponent("{rel}");</script>'
                elif ext in ('.txt', '.csv', '.json'):
                    with open(full, 'r', errors='replace') as f:
                        content = f.read()
                    escaped = content.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                    html = f'<pre style="background:#161b22;padding:14px;border-radius:6px;overflow-x:auto;border:1px solid #30363d;font-size:13px;line-height:1.6;white-space:pre-wrap">{escaped}</pre>'
                else:
                    html = '<p style="color:#7d8590">不支持预览此文件类型</p>'
            except Exception as e:
                html = f'<div style="color:#f85149;padding:20px">预览失败: {str(e)}</div>'
            self._json_resp(200, {'html': html, 'filename': os.path.basename(full)})
        elif path == '/api/raw':
            # Serve raw files (images, HTML, downloads)
            qs = parse_qs(urlparse(self.path).query)
            rel = qs.get('path', [''])[0]
            full = os.path.normpath(os.path.join(WORKSPACE, rel))
            if not full.startswith(WORKSPACE) or not os.path.isfile(full):
                self.send_error(404); return
            if session and not check_path_access(session['paths'], rel):
                self.send_error(403); return
            mime = mimetypes.guess_type(full)[0] or 'application/octet-stream'
            self.send_response(200)
            self.send_header('Content-Type', mime)
            # Download mode
            if 'dl' in qs:
                fname = os.path.basename(full)
                self.send_header('Content-Disposition', 'attachment; filename="' + fname + '"')
            size = os.path.getsize(full)
            self.send_header('Content-Length', str(size))
            self.send_header('Cache-Control', 'max-age=3600')
            self.end_headers()
            with open(full, 'rb') as f:
                self.wfile.write(f.read())
        elif path == '/api/annotations':
            # GET: list annotation files
            ann_files = sorted(os.listdir(ANN_DIR)) if os.path.isdir(ANN_DIR) else []
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'files': ann_files}).encode())
        elif path == '/api/references':
            # Scan all .md files for references to a given path
            if not FULL_MODE:
                self.send_error(403); return
            qs = parse_qs(urlparse(self.path).query)
            target = qs.get('path', [''])[0]
            if not target:
                self.send_error(400); return
            basename = os.path.basename(target)
            # Patterns to search: the basename, the full relative path, and ./relative path
            patterns = [basename, target, './' + target]
            # Also search without .md extension
            if basename.endswith('.md'):
                patterns.append(basename[:-3])
            skip = {'.git','node_modules','.next','dist','__pycache__','.cache','.trash','.turbo','.venv','.npm','.config','.local','.ssh'}
            user_paths = session['paths'] if session else ['*']
            refs = []
            for root, dirs, fnames in os.walk(WORKSPACE):
                dirs[:] = [d for d in dirs if d not in skip]
                for f in fnames:
                    if not f.endswith('.md'):
                        continue
                    full = os.path.join(root, f)
                    rel = os.path.relpath(full, WORKSPACE)
                    if rel == target:
                        continue  # skip self
                    if not check_path_access(user_paths, rel):
                        continue
                    try:
                        with open(full, 'r', errors='replace') as fh:
                            content = fh.read()
                        for pat in patterns:
                            if pat in content:
                                # Find the specific lines
                                lines = []
                                for i, line in enumerate(content.split('\n'), 1):
                                    if pat in line:
                                        lines.append({'line': i, 'text': line.strip()[:120]})
                                refs.append({'file': rel, 'matches': lines[:5]})
                                break
                    except:
                        pass
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'refs': refs, 'count': len(refs)}).encode())
        elif path == '/api/trash':
            # List trash items
            if not FULL_MODE:
                self.send_error(403); return
            items = list_trash()
            user_paths = session['paths'] if session else ['*']
            items = [it for it in items if check_path_access(user_paths, it.get('original', ''))]
            self._json_resp(200, {'ok': True, 'items': items})
            return

        elif path == '/api/activity':
            # Admin only: activity log
            if not session or session['role'] != 'admin':
                self._json_resp(403, {'ok': False, 'error': 'Forbidden'}); return
            qs = {}
            if '?' in self.path:
                for part in self.path.split('?', 1)[1].split('&'):
                    if '=' in part:
                        k, v = part.split('=', 1)
                        qs[k] = v
            limit = min(int(qs.get('limit', '50')), 200)
            offset = int(qs.get('offset', '0'))
            items, total = get_activity(limit, offset)
            self._json_resp(200, {'ok': True, 'items': items, 'total': total})
            return

        elif path == '/api/users':
            # Admin only: list users
            if not session or session['role'] != 'admin':
                self._json_resp(403, {'ok': False, 'error': '权限不足'})
                return
            users_data = get_users()
            # Return users without password hashes
            safe_users = {}
            for uname, uinfo in users_data.get('users', {}).items():
                safe_users[uname] = {k: v for k, v in uinfo.items() if k != 'password_hash'}
            self._json_resp(200, {'ok': True, 'users': safe_users})
            return
        elif path == '/api/logout':
            # Also handle GET logout (redirect)
            token = _get_auth_cookie(self)
            if token:
                delete_session(token)
            self.send_response(302)
            self.send_header('Set-Cookie', 'auth_token=; Path=/; HttpOnly; Max-Age=0')
            self.send_header('Location', 'login')
            self.end_headers()
            return
        elif path == '/api/dirs':
            # List directories for smart new-file
            if not FULL_MODE:
                self.send_error(403); return
            skip = {'.git','node_modules','.next','dist','__pycache__','.cache','.trash','.turbo','.venv','.npm','.config','.local','.ssh'}
            user_paths = session['paths'] if session else ['*']
            dir_list = []
            for root, dirs, fnames in os.walk(WORKSPACE):
                dirs[:] = sorted([d for d in dirs if d not in skip])
                rel = os.path.relpath(root, WORKSPACE)
                if rel != '.' and check_path_access(user_paths, rel):
                    dir_list.append(rel)
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'dirs': sorted(dir_list)[:500]}).encode())
        else:
            self.send_error(404)

    def do_DELETE(self):
        path = self.path.split('?')[0].rstrip('/')
        
        # Auth check
        if not path.startswith('/s/'):
            session = self._check_auth()
            if not session:
                self._json_resp(401, {'ok': False, 'error': '未登录'}); return
        
        length = int(self.headers.get('Content-Length', 0))
        body = json.loads(self.rfile.read(length)) if length else {}
        
        # DELETE /api/users/{username}
        if path.startswith('/api/users/'):
            if not session or session['role'] != 'admin':
                self._json_resp(403, {'ok': False, 'error': '权限不足'}); return
            username = path[len('/api/users/'):]
            username = url_unquote(username)
            if username == session['user']:
                self._json_resp(400, {'ok': False, 'error': '不能删除自己'}); return
            users_data = get_users()
            if username not in users_data.get('users', {}):
                self._json_resp(404, {'ok': False, 'error': '用户不存在'}); return
            del users_data['users'][username]
            _save_users(users_data)
            log_activity(session['user'], 'user_delete', username)
            self._json_resp(200, {'ok': True})
            return
        elif path == '/api/share':
            if session and session['role'] == 'viewer':
                self._json_resp(403, {'ok': False, 'error': 'No permission'}); return
            token = body.get('token', '')
            if delete_share(token):
                self._json_resp(200, {'ok': True})
            else:
                self._json_resp(404, {'ok': False, 'error': 'Token not found'})
        else:
            self.send_error(404)

    def do_PUT(self):
        path = self.path.split('?')[0].rstrip('/')
        length = int(self.headers.get('Content-Length', 0))
        body = json.loads(self.rfile.read(length)) if length else {}

        # PUT /api/users/{username} — update user (admin only)
        if path.startswith('/api/users/'):
            session = self._check_auth()
            if not session or session['role'] != 'admin':
                self._json_resp(403, {'ok': False, 'error': '权限不足'}); return
            username = path[len('/api/users/'):]
            username = url_unquote(username)
            users_data = get_users()
            if username not in users_data.get('users', {}):
                self._json_resp(404, {'ok': False, 'error': '用户不存在'}); return
            user = users_data['users'][username]
            if 'name' in body: user['name'] = body['name']
            if 'role' in body and body['role'] in ('admin', 'editor', 'viewer'): user['role'] = body['role']
            if 'paths' in body: user['paths'] = body['paths']
            if body.get('password'): user['password_hash'] = hash_password(body['password'])
            if 'pin' in body:
                if body['pin']:
                    user['pin'] = str(body['pin'])
                else:
                    user.pop('pin', None)
            _save_users(users_data)
            self._json_resp(200, {'ok': True})
            return

        self.send_error(404)

    def do_POST(self):
        path = self.path.split('?')[0].rstrip('/')
        length = int(self.headers.get('Content-Length', 0))
        raw_body = self.rfile.read(length)

        # ==================== SETUP ====================
        if path == '/_setup' and _setup.needs_setup(_EDITOR_DIR):
            body_data = json.loads(raw_body) if raw_body else {}
            result = _setup.handle_setup_post(body_data, _EDITOR_DIR)
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps(result).encode())
            return

        # ==================== LOGIN ====================
        if path == '/login' or path == '/login/':
            # Parse form data or JSON
            content_type = self.headers.get('Content-Type', '')
            if 'application/x-www-form-urlencoded' in content_type:
                form = parse_qs(raw_body.decode(), keep_blank_values=True)
                username = form.get('username', [''])[0].strip()
                password = form.get('password', [''])[0]
                pin_input = form.get('pin', [''])[0]
                redirect_to = _safe_redirect(form.get('redirect', [''])[0]) or './'
                remember = form.get('remember', [''])[0] == '1'
                is_pin_login = form.get('pin_login', [''])[0] == '1'
            else:
                body = json.loads(raw_body) if raw_body else {}
                username = body.get('username', '').strip()
                password = body.get('password', '')
                pin_input = body.get('pin', '')
                redirect_to = _safe_redirect(body.get('redirect', '')) or './'
                remember = body.get('remember', False)
                is_pin_login = body.get('pin_login', False)
            
            # PIN login (device remembered, just need PIN or auto-pass)
            if is_pin_login and username:
                users_data = get_users()
                user_info = users_data.get('users', {}).get(username)
                if user_info and verify_pin(username, pin_input):
                    token = create_session(username, user_info)
                    expire_hours = users_data.get('settings', {}).get('sessionExpireHours', 72)
                    max_age = int(expire_hours * 3600)
                    self.send_response(302)
                    self.send_header('Set-Cookie', f'auth_token={token}; Path=/; HttpOnly; Max-Age={max_age}')
                    self.send_header('Location', redirect_to)
                    self.end_headers()
                    return
                else:
                    self._serve_pin_page(username, redirect_to, error='PIN 错误')
                    return
            
            # Normal password login
            users_data = get_users()
            user_info = users_data.get('users', {}).get(username)
            if not user_info or not verify_password(password, user_info.get('password_hash', '')):
                self._serve_login(error='用户名或密码错误', redirect_to=redirect_to)
                return
            
            token = create_session(username, user_info)
            expire_hours = users_data.get('settings', {}).get('sessionExpireHours', 72)
            max_age = int(expire_hours * 3600)
            
            self.send_response(302)
            self.send_header('Set-Cookie', f'auth_token={token}; Path=/; HttpOnly; Max-Age={max_age}')
            # Set device cookie if "remember" checked
            if remember:
                dev_token = create_device_token(username)
                if dev_token:
                    self.send_header('Set-Cookie', f'device_token={dev_token}; Path=/; Max-Age={_DEVICE_COOKIE_MAX_AGE}')
            self.send_header('Location', redirect_to)
            self.end_headers()
            return

        # ==================== AUTH CHECK FOR POST ====================
        # Share links are public
        if not path.startswith('/s/'):
            session = self._check_auth()
            if path == '/api/logout':
                # Logout
                token = _get_auth_cookie(self)
                if token:
                    delete_session(token)
                self.send_response(200)
                self.send_header('Set-Cookie', 'auth_token=; Path=/; HttpOnly; Max-Age=0')
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({'ok': True}).encode())
                return
            if not session:
                # Allow /api/pin without session if we can identify by other means
                if path != '/api/pin':
                    self._json_resp(401, {'ok': False, 'error': '未登录'})
                    return
        else:
            session = None

        _ct = self.headers.get('Content-Type', '')
        body = json.loads(raw_body) if raw_body and 'multipart' not in _ct and 'x-www-form-urlencoded' not in _ct else {}

        # ==================== SET OWN PIN ====================
        if path == '/api/pin':
            if not session:
                self._json_resp(401, {'ok': False, 'error': '未登录'}); return
            pin = body.get('pin', '')
            if pin and (not str(pin).isdigit() or len(str(pin)) != 4):
                self._json_resp(400, {'ok': False, 'error': 'PIN 必须是 4 位数字'}); return
            set_user_pin(session['user'], str(pin) if pin else None)
            self._json_resp(200, {'ok': True, 'hasPin': bool(pin)})
            return

        # ==================== USER MANAGEMENT ====================
        if path == '/api/users':
            # Create user (admin only)
            if not session or session['role'] != 'admin':
                self._json_resp(403, {'ok': False, 'error': '权限不足'}); return
            username = body.get('username', '').strip()
            password = body.get('password', '')
            name = body.get('name', '')
            role = body.get('role', 'viewer')
            paths = body.get('paths', ['*'])
            if not username or not password:
                self._json_resp(400, {'ok': False, 'error': '用户名和密码不能为空'}); return
            if role not in ('admin', 'editor', 'viewer'):
                self._json_resp(400, {'ok': False, 'error': '无效角色'}); return
            users_data = get_users()
            if username in users_data.get('users', {}):
                self._json_resp(409, {'ok': False, 'error': '用户已存在'}); return
            users_data.setdefault('users', {})[username] = {
                'password_hash': hash_password(password),
                'name': name,
                'role': role,
                'paths': paths,
                'created': time.strftime('%Y-%m-%d')
            }
            _save_users(users_data)
            log_activity(session['user'], 'user_create', username, role)
            self._json_resp(200, {'ok': True})
            return

        # Handle share token API calls
        if path.startswith('/s/'):
            parts = path[3:].split('/', 1)
            token = parts[0]
            sub_path = '/' + parts[1] if len(parts) > 1 else ''
            share = get_share(token)
            if not share:
                self.send_error(404); return
            if sub_path == '/api/save':
                if share['mode'] == 'readonly':
                    self._json_resp(403, {'ok': False, 'error': 'Read-only'}); return
                share_type = share.get('type', 'file')
                if share_type == 'folder':
                    rel = body.get('path', '')
                    allowed = share.get('files', [])
                    if rel not in allowed:
                        self._json_resp(403, {'ok': False, 'error': 'File not in share'}); return
                    folder_full = os.path.normpath(os.path.join(WORKSPACE, share['file']))
                    full = os.path.normpath(os.path.join(folder_full, rel))
                else:
                    full = os.path.normpath(os.path.join(WORKSPACE, share['file']))
                if not full.startswith(WORKSPACE):
                    self.send_error(403); return
                with open(full, 'w') as f:
                    f.write(body['content'])
                self._json_resp(200, {'ok': True})
            elif sub_path == '/api/annotations':
                ts = time.strftime('%Y%m%d-%H%M%S')
                safe_name = share['file'].replace('/', '_').replace('\\', '_')
                ann_file = os.path.join(ANN_DIR, f'{ts}_{safe_name}.json')
                with open(ann_file, 'w') as f:
                    json.dump(body, f, ensure_ascii=False, indent=2)
                self._json_resp(200, {'ok': True, 'file': os.path.basename(ann_file)})
            elif sub_path == '/api/export':
                content = body.get('content', '')
                fmt = body.get('format', 'pdf')
                theme = body.get('theme', 'dark')
                signature = body.get('signature', 'none')
                title = body.get('title', 'document')
                layout = body.get('layout', 'a4')
                if not content:
                    self._json_resp(400, {'ok': False, 'error': 'No content'}); return
                word_style = body.get('wordStyle', 'modern')
                word_margin = body.get('wordMargin', 'normal')
                word_page = body.get('wordPage', 'a4')
                custom_sig = None
                if signature == 'custom':
                    custom_sig = {'name': body.get('customName', ''), 'avatar': body.get('customAvatar')}
                try:
                    if fmt == 'word':
                        out_path, filename, dl_id = export_word(content, signature=signature, title=title, word_style=word_style, word_margin=word_margin, word_page=word_page, custom_sig=custom_sig)
                    else:
                        out_path, filename, dl_id = export_pdf(content, theme=theme, signature=signature, title=title, layout=layout, custom_sig=custom_sig)
                    with _downloads_lock:
                        _downloads[dl_id] = {'path': out_path, 'filename': filename, 'created': time.time()}
                    self._json_resp(200, {'ok': True, 'downloadId': dl_id, 'filename': filename})
                except Exception as e:
                    self._json_resp(500, {'ok': False, 'error': str(e)})
            else:
                self.send_error(403)
            return
        
        if path == '/api/share':
            # Create a share link (file or folder)
            if not FULL_MODE:
                self.send_error(403); return
            if session and session['role'] == 'viewer':
                self._json_resp(403, {'ok': False, 'error': 'No permission'}); return
            file_path = body.get('file', '')
            mode = body.get('mode', 'readonly')
            share_type = body.get('type', 'file')
            if mode not in ('readonly', 'editable'):
                self._json_resp(400, {'ok': False, 'error': 'Invalid mode'}); return
            if session and not check_path_access(session['paths'], file_path):
                self._json_resp(403, {'ok': False, 'error': 'No access'}); return
            if share_type == 'folder':
                folder_files = body.get('files', [])
                full = os.path.normpath(os.path.join(WORKSPACE, file_path))
                if not full.startswith(WORKSPACE) or not os.path.isdir(full):
                    self._json_resp(404, {'ok': False, 'error': 'Folder not found'}); return
                token = create_share(file_path, mode, share_type='folder', files=folder_files)
                if session:
                    log_activity(session['user'], 'share', file_path, mode + ' (folder)')
                self._json_resp(200, {'ok': True, 'token': token, 'mode': mode, 'type': 'folder', 'fileCount': len(folder_files)})
            else:
                full = os.path.normpath(os.path.join(WORKSPACE, file_path))
                if not full.startswith(WORKSPACE) or not os.path.isfile(full):
                    self._json_resp(404, {'ok': False, 'error': 'File not found'}); return
                token = create_share(file_path, mode)
                if session:
                    log_activity(session['user'], 'share', file_path, mode)
                self._json_resp(200, {'ok': True, 'token': token, 'mode': mode})


        elif path == '/api/upload':
            if not FULL_MODE:
                self._json_resp(403, {'ok': False, 'error': 'Not available'}); return
            if session and session['role'] == 'viewer':
                self._json_resp(403, {'ok': False, 'error': 'No permission'}); return
            content_type = self.headers.get('Content-Type', '')
            if 'multipart/form-data' not in content_type:
                self._json_resp(400, {'ok': False, 'error': 'Multipart required'}); return
            # Parse boundary
            boundary = None
            for part in content_type.split(';'):
                part = part.strip()
                if part.startswith('boundary='):
                    boundary = part[9:].strip('"')
            if not boundary:
                self._json_resp(400, {'ok': False, 'error': 'No boundary'}); return
            # Parse multipart manually
            sep = ('--' + boundary).encode()
            parts = raw_body.split(sep)
            dest_dir = ''
            file_parts = []
            for p in parts:
                p = p.strip()
                if not p or p == b'--':
                    continue
                # Split headers from body
                hdr_end = p.find(b'\r\n\r\n')
                if hdr_end < 0:
                    continue
                headers_raw = p[:hdr_end].decode('utf-8', errors='replace')
                body_data = p[hdr_end+4:]
                if body_data.endswith(b'\r\n'):
                    body_data = body_data[:-2]
                # Parse Content-Disposition
                fname = None
                field_name = None
                for line in headers_raw.split('\r\n'):
                    if 'Content-Disposition' in line:
                        for kv in line.split(';'):
                            kv = kv.strip()
                            if kv.startswith('name='):
                                field_name = kv[5:].strip('"')
                            elif kv.startswith('filename='):
                                fname = kv[9:].strip('"')
                if field_name == 'dir' and not fname:
                    dest_dir = body_data.decode('utf-8', errors='replace').strip()
                elif fname:
                    file_parts.append((fname, body_data))
            uploaded = []
            failed = []
            for fname, data in file_parts:
                fname = os.path.basename(fname)
                if not fname:
                    continue
                rel = os.path.join(dest_dir, fname) if dest_dir else fname
                full = os.path.normpath(os.path.join(WORKSPACE, rel))
                if not full.startswith(WORKSPACE):
                    failed.append(fname); continue
                if session and not check_path_access(session['paths'], rel):
                    failed.append(fname); continue
                os.makedirs(os.path.dirname(full), exist_ok=True)
                with open(full, 'wb') as f:
                    f.write(data)
                uploaded.append(rel)
                if session:
                    log_activity(session['user'], 'upload', rel)
            self._json_resp(200, {'ok': True, 'uploaded': uploaded, 'failed': failed})
            return
        elif path == '/api/rename':
            if not FULL_MODE:
                self.send_error(403); return
            old_path = body.get('path', '').strip()
            new_name = body.get('name', '').strip()
            if not old_path or not new_name:
                self._json_resp(400, {'ok': False, 'error': 'Missing path or name'}); return
            if '/' in new_name or '\\' in new_name:
                self._json_resp(400, {'ok': False, 'error': 'Invalid name'}); return
            old_full = os.path.normpath(os.path.join(WORKSPACE, old_path))
            if not old_full.startswith(WORKSPACE) or not os.path.exists(old_full):
                self._json_resp(404, {'ok': False, 'error': 'Not found'}); return
            if session:
                if session['role'] == 'viewer':
                    self._json_resp(403, {'ok': False, 'error': 'No permission'}); return
                if not check_path_access(session['paths'], old_path):
                    self._json_resp(403, {'ok': False, 'error': 'No access'}); return
            new_full = os.path.join(os.path.dirname(old_full), new_name)
            if os.path.exists(new_full):
                self._json_resp(409, {'ok': False, 'error': 'Name already exists'}); return
            os.rename(old_full, new_full)
            new_rel = os.path.relpath(new_full, WORKSPACE)
            if session:
                log_activity(session['user'], 'rename', old_path, new_rel)
            self._json_resp(200, {'ok': True, 'newPath': new_rel})
            return
        elif path == '/api/save':
            if FULL_MODE:
                rel = body.get('path', '')
                full = os.path.normpath(os.path.join(WORKSPACE, rel))
                if not full.startswith(WORKSPACE):
                    self.send_error(403); return
                # Path + role access check
                if session:
                    if session['role'] == 'viewer':
                        self._json_resp(403, {'ok': False, 'error': '查看者无法编辑文件'}); return
                    if not check_path_access(session['paths'], rel):
                        self._json_resp(403, {'ok': False, 'error': '无权编辑此文件'}); return
            else:
                full = SHARE_FILE
            with open(full, 'w') as f:
                f.write(body['content'])
            if session:
                log_activity(session['user'], 'save', body.get('path', 'shared'))
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'ok': True}).encode())
        elif path == '/api/move':
            if not FULL_MODE:
                self.send_error(403); return
            if session and session['role'] == 'viewer':
                self._json_resp(403, {'ok': False, 'error': 'No permission'}); return
            src = body.get('from', '').strip()
            dest_dir = body.get('toDir', '').strip()
            if not src:
                self._json_resp(400, {'ok': False, 'error': '缺少源文件路径'}); return
            src_full = os.path.normpath(os.path.join(WORKSPACE, src))
            if not src_full.startswith(WORKSPACE) or not os.path.isfile(src_full):
                self._json_resp(404, {'ok': False, 'error': '源文件不存在'}); return
            if session and not check_path_access(session['paths'], src):
                self._json_resp(403, {'ok': False, 'error': 'No access to this file'}); return
            # Destination
            fname = os.path.basename(src)
            if dest_dir:
                dest_full = os.path.normpath(os.path.join(WORKSPACE, dest_dir, fname))
            else:
                dest_full = os.path.normpath(os.path.join(WORKSPACE, fname))
            if not dest_full.startswith(WORKSPACE):
                self._json_resp(403, {'ok': False, 'error': '目标路径越界'}); return
            if os.path.exists(dest_full):
                self._json_resp(409, {'ok': False, 'error': '目标已存在: ' + os.path.relpath(dest_full, WORKSPACE)}); return
            os.makedirs(os.path.dirname(dest_full), exist_ok=True)
            os.rename(src_full, dest_full)
            new_rel = os.path.relpath(dest_full, WORKSPACE)
            if session:
                log_activity(session['user'], 'move', src, new_rel)
            self._json_resp(200, {'ok': True, 'newPath': new_rel})
            return
        elif path == '/api/delete':
            if not FULL_MODE:
                self.send_error(403); return
            if session and session['role'] == 'viewer':
                self._json_resp(403, {'ok': False, 'error': 'No permission'}); return
            rel = body.get('path', '').strip()
            if not rel:
                self._json_resp(400, {'ok': False, 'error': '缺少路径'}); return
            if session and not check_path_access(session['paths'], rel):
                self._json_resp(403, {'ok': False, 'error': 'No access to this file'}); return
            entry = trash_item(rel)
            if session and entry:
                log_activity(session['user'], 'delete', rel)
            if entry:
                self._json_resp(200, {'ok': True, 'entry': entry})
            else:
                self._json_resp(400, {'ok': False, 'error': '无法删除（路径不存在或受保护）'})
            return
        elif path == '/api/restore':
            if not FULL_MODE:
                self.send_error(403); return
            if session and session['role'] == 'viewer':
                self._json_resp(403, {'ok': False, 'error': 'No permission'}); return
            trash_id = body.get('id', '').strip()
            if not trash_id:
                self._json_resp(400, {'ok': False, 'error': '缺少 id'}); return
            entry, err = restore_item(trash_id)
            if entry:
                if session:
                    log_activity(session['user'], 'restore', entry['original'])
                self._json_resp(200, {'ok': True, 'restored': entry['original']})
            else:
                self._json_resp(400, {'ok': False, 'error': err})
            return
        elif path == '/api/trash-empty':
            if not FULL_MODE:
                self.send_error(403); return
            if not session or session['role'] != 'admin':
                self._json_resp(403, {'ok': False, 'error': 'Admin only'}); return
            empty_trash()
            if session:
                log_activity(session['user'], 'empty_trash', '')
            self._json_resp(200, {'ok': True})
            return
        elif path == '/api/new':
            if not FULL_MODE:
                self.send_error(403); return
            if session and session['role'] == 'viewer':
                self._json_resp(403, {'ok': False, 'error': 'No permission'}); return
            rel = body.get('path', '').strip()
            if not rel or '..' in rel:
                self.send_response(400)
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({'ok': False, 'error': '无效路径'}).encode())
                return
            if not rel.endswith('.md'):
                rel += '.md'
            if session and not check_path_access(session['paths'], rel):
                self._json_resp(403, {'ok': False, 'error': 'No access to this path'}); return
            full = os.path.normpath(os.path.join(WORKSPACE, rel))
            if not full.startswith(WORKSPACE):
                self.send_response(403)
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({'ok': False, 'error': '路径越界'}).encode())
                return
            if os.path.exists(full):
                self.send_response(409)
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({'ok': False, 'error': '文件已存在'}).encode())
                return
            os.makedirs(os.path.dirname(full), exist_ok=True)
            with open(full, 'w') as f:
                f.write(f'# {os.path.splitext(os.path.basename(rel))[0]}\n\n')
            if session:
                log_activity(session['user'], 'create', rel)
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'ok': True, 'path': rel}).encode())
        elif path == '/api/annotations':
            # Save annotations as JSON file
            ts = time.strftime('%Y%m%d-%H%M%S')
            safe_name = body.get('file', 'unknown').replace('/', '_').replace('\\', '_')
            ann_file = os.path.join(ANN_DIR, f'{ts}_{safe_name}.json')
            with open(ann_file, 'w') as f:
                json.dump(body, f, ensure_ascii=False, indent=2)
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'ok': True, 'file': os.path.basename(ann_file)}).encode())
        elif path == '/api/custom-signatures':
            # CRUD for custom signatures (array)
            session = self._check_auth()
            if not session:
                self._json_resp(401, {'ok': False, 'error': 'Not logged in'}); return
            username = session['user']
            action = body.get('action', 'create')  # create, update, delete
            users_data = get_users()
            user = users_data.get('users', {}).get(username)
            if not user:
                self._json_resp(404, {'ok': False, 'error': 'User not found'}); return
            sigs = user.setdefault('customSignatures', [])
            # Max 10 custom signatures per user
            if action == 'create':
                if len(sigs) >= 10:
                    self._json_resp(400, {'ok': False, 'error': '最多 10 个自定义署名'}); return
                sig_name = body.get('name', '').strip()
                if not sig_name:
                    self._json_resp(400, {'ok': False, 'error': '署名不能为空'}); return
                sig_id = secrets.token_urlsafe(6)
                avatar_rel = self._save_custom_avatar(username, sig_id, body.get('avatar'))
                sigs.append({'id': sig_id, 'name': sig_name, 'avatar': avatar_rel})
                _save_users(users_data)
                self._json_resp(200, {'ok': True, 'id': sig_id})
            elif action == 'update':
                sig_id = body.get('id', '')
                target = next((s for s in sigs if s.get('id') == sig_id), None)
                if not target:
                    self._json_resp(404, {'ok': False, 'error': '署名不存在'}); return
                if 'name' in body:
                    new_name = body['name'].strip()
                    if new_name: target['name'] = new_name
                avatar_data = body.get('avatar')
                if avatar_data and avatar_data != '__keep__':
                    avatar_rel = self._save_custom_avatar(username, sig_id, avatar_data)
                    if avatar_rel: target['avatar'] = avatar_rel
                elif avatar_data is None:
                    # Explicitly clear avatar
                    old_av = target.get('avatar')
                    if old_av:
                        try: os.unlink(os.path.join(_EDITOR_DIR, old_av))
                        except: pass
                    target['avatar'] = None
                _save_users(users_data)
                self._json_resp(200, {'ok': True})
            elif action == 'delete':
                sig_id = body.get('id', '')
                target = next((s for s in sigs if s.get('id') == sig_id), None)
                if target:
                    # Remove avatar file
                    if target.get('avatar'):
                        try: os.unlink(os.path.join(_EDITOR_DIR, target['avatar']))
                        except: pass
                    sigs.remove(target)
                    _save_users(users_data)
                self._json_resp(200, {'ok': True})
            else:
                self._json_resp(400, {'ok': False, 'error': 'Unknown action'})
            return
        elif path == '/api/export':
            content = body.get('content', '')
            fmt = body.get('format', 'pdf')
            theme = body.get('theme', 'dark')
            signature = body.get('signature', 'none')
            title = body.get('title', 'document')
            layout = body.get('layout', 'a4')
            if not content:
                self._json_resp(400, {'ok': False, 'error': 'No content provided'})
                return
            word_style = body.get('wordStyle', 'modern')
            word_margin = body.get('wordMargin', 'normal')
            word_page = body.get('wordPage', 'a4')
            custom_sig = None
            if signature == 'custom':
                custom_sig = {'name': body.get('customName', ''), 'avatar': body.get('customAvatar')}
            try:
                if fmt == 'word':
                    out_path, filename, dl_id = export_word(content, signature=signature, title=title, word_style=word_style, word_margin=word_margin, word_page=word_page, custom_sig=custom_sig)
                else:
                    out_path, filename, dl_id = export_pdf(content, theme=theme, signature=signature, title=title, layout=layout, custom_sig=custom_sig)
                with _downloads_lock:
                    _downloads[dl_id] = {'path': out_path, 'filename': filename, 'created': time.time()}
                self._json_resp(200, {'ok': True, 'downloadId': dl_id, 'filename': filename})
            except Exception as e:
                self._json_resp(500, {'ok': False, 'error': str(e)})
            return
        elif path == '/api/ai-edit':
            # AI-assisted content editing via Claude
            if not FULL_MODE:
                # Also allow in editable share mode
                token_match = re.match(r'/s/([^/]+)', urlparse(self.headers.get('Referer', '')).path)
                if token_match:
                    share = get_share(token_match.group(1))
                    if not share or share['mode'] == 'readonly':
                        self._json_resp(403, {'ok': False, 'error': 'Read-only share'})
                        return
            instruction = body.get('instruction', '')
            content = body.get('content', '')
            selection = body.get('selection', '')
            if not instruction or not content:
                self._json_resp(400, {'ok': False, 'error': 'Missing instruction or content'})
                return
            try:
                import anthropic
                _ai_key = os.environ.get('ANTHROPIC_API_KEY', '')
                if not _ai_key:
                    # Try loading from .env.local
                    _env_path = os.path.join(WORKSPACE, 'Projects/Asyre/web/.env.local')
                    if os.path.isfile(_env_path):
                        with open(_env_path) as ef:
                            for line in ef:
                                if line.startswith('ANTHROPIC_API_KEY='):
                                    _ai_key = line.split('=', 1)[1].strip()
                if not _ai_key:
                    self._json_resp(500, {'ok': False, 'error': 'No API key configured'})
                    return
                client = anthropic.Anthropic(api_key=_ai_key)
                # Read AI config (hot-reloadable)
                ai_cfg = get_config().get('ai', {})
                system_prompt = ai_cfg.get('systemPrompt',
                    "你是一个 Markdown 文档编辑助手。用户会给你一篇文档和修改指令。\n"
                    "你的任务是按照指令修改文档内容，返回修改后的完整文档。\n"
                    "规则：\n"
                    "1. 只返回修改后的完整 Markdown 文档，不要加任何解释或包裹标记\n"
                    "2. 保留原文的格式（标题、列表、加粗等）\n"
                    "3. 只修改指令要求的部分，其余保持不变\n"
                    "4. 如果指令不清楚，按最合理的理解执行"
                )
                ai_model = ai_cfg.get('model', 'claude-sonnet-4-5-20250514')
                user_msg = f"## 修改指令\n{instruction}\n\n"
                if selection:
                    user_msg += f"## 选中的文本（重点修改这部分）\n```\n{selection}\n```\n\n"
                user_msg += f"## 文档内容\n{content}"
                
                response = client.messages.create(
                    model=ai_model,
                    max_tokens=8192,
                    system=system_prompt,
                    messages=[{"role": "user", "content": user_msg}]
                )
                result = response.content[0].text
                # Generate a brief summary of changes
                summary_resp = client.messages.create(
                    model="claude-haiku-4-5-20250514",
                    max_tokens=200,
                    messages=[{"role": "user", "content": f"用一句话概括这个修改指令做了什么（中文，不超过50字）：\n指令：{instruction}"}]
                )
                summary = summary_resp.content[0].text
                self._json_resp(200, {'ok': True, 'content': result, 'summary': summary})
            except Exception as e:
                self._json_resp(500, {'ok': False, 'error': str(e)})
        else:
            self.send_error(404)

    def log_message(self, *args): pass

if __name__ == '__main__':
    if '--version' in sys.argv:
        print(f'Asyre File v{__version__}')
        sys.exit(0)
    if '--setup' in sys.argv:
        _setup.cli_setup(_EDITOR_DIR)
        sys.exit(0)
    _load_sessions_from_file()
    start_cleanup_thread()
    mode = 'SHARE: ' + SHARE_FILE if SHARE_FILE else 'FULL (all files)'
    print(f'Asyre File v{__version__} [{mode}] → http://{_cfg.get("server.host")}:{PORT}')
    HTTPServer((_cfg.get('server.host'), PORT), Handler).serve_forever()
